# ===============================
# Future imports
# ===============================
from __future__ import annotations

# ===============================
# Standard Library Imports
# ===============================
import io
import json
import re
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from typing import Optional, Tuple, Any
from urllib.parse import urlencode

# ===============================
# Django Core Imports
# ===============================
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import models, transaction
from django.db.models import (
    Case, F, Q, Sum, Value, Value as V, 
    ExpressionWrapper, DecimalField, CharField,
    When, OuterRef, Subquery, Prefetch, Window,
    Count
)
from django.db.models.functions import TruncDate, Coalesce
from django.http import HttpRequest, HttpResponse, JsonResponse, Http404
from django.middleware.csrf import get_token
from django.shortcuts import render, get_object_or_404, redirect
from django.templatetags.static import static
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.decorators import method_decorator
from django.utils.timezone import is_naive, make_aware, get_current_timezone
from django.views.decorators.csrf import csrf_exempt

# ===============================
# Django View Imports
# ===============================
from django.views import View
from django.views.decorators.http import require_GET, require_http_methods, require_POST
from django.views.generic import (
    CreateView, UpdateView, DeleteView, ListView, 
    DetailView, TemplateView, FormView
)

# ===============================
# Local App Imports (Models)
# ===============================
from .models import (
    Business, BusinessStock, Party, Staff,
    ProductCategory, Product, ProductImage,
    BankAccount, BankMovement, CashFlow,
    PurchaseOrder, PurchaseOrderItem, PurchaseOrderPayment,
    PurchaseReturn, PurchaseReturnItem, PurchaseReturnRefund,
    SalesOrder, SalesOrderItem, SalesInvoice, SalesInvoiceItem,
    SalesInvoiceReceipt, SalesOrderReceipt, SalesReturn, 
    SalesReturnItem, SalesReturnRefund,
    Payment, Expense, ExpenseCategory,
    StockTransaction, StockMove, Warehouse, WarehouseStock,
    UserSettings, BusinessSummary
)

# ===============================
# Local App Imports (Forms & Utils)
# ===============================
from .forms import (
    BusinessForm, PartyForm, ProductCategoryForm, 
    ProductForm, StaffForm, BankAccountForm, 
    BankMovementForm, CashFlowFilterForm,
    PurchaseOrderForm, PurchaseOrderItemFormSet,
    PurchaseReturnForm, PurchaseReturnItemFormSet,
    SalesOrderForm, SalesOrderItemForm, SalesOrderItemFormSet,
    SalesInvoiceForm, SalesInvoiceItemFormSet,
    SalesReturnForm, SalesReturnItemFormSet,
    ExpenseForm, PurchaseOrderExpenseFormSet, ProductImageFormSet,
    WarehouseForm, QuickReceiptForm,
    UserSettingsForm
)
from .ledger import build_ledger
from django.core.management import call_command

# ---------- Dashboard / Businesses ----------
@method_decorator(login_required, name="dispatch")
class BusinessesView(ListView):
    template_name = "barkat/dashboard/businesses.html"
    context_object_name = "businesses"
    paginate_by = 25  # optional

    def get_queryset(self):
        return Business.objects.filter(is_deleted=False).select_related("summary").order_by("-id")

    # ---- helper to build overall party summary (all businesses) ----
    def _build_party_summary(self, kind: str, q: str | None,
                             date_from, date_to):
        """
        Returns list of dicts:
          {
            "party": Party instance,
            "balance_abs": Decimal,
            "balance_side": "Dr" or "Cr",
          }
        Only nonzero balances. merged across all active businesses.
        """
        kind = (kind or "customer").strip().lower()
        if kind not in ("customer", "supplier"):
            kind = "customer"

        # base party queryset
        if kind == "customer":
            party_qs = Party.objects.filter(
                type__in=[Party.CUSTOMER, Party.BOTH]
            )
        else:
            party_qs = Party.objects.filter(
                type__in=[Party.VENDOR, Party.BOTH]
            )

        party_qs = party_qs.filter(is_deleted=False)

        if q:
            party_qs = party_qs.filter(
                Q(display_name__icontains=q)
                | Q(phone__icontains=q)
                | Q(email__icontains=q)
            )

        party_qs = party_qs.order_by("display_name", "id")

        biz_list = list(
            Business.objects.filter(is_deleted=False, is_active=True)
            .order_by("name", "id")
        )

        rows = []
        for p in party_qs:
            total_dr = Decimal("0.00")
            total_cr = Decimal("0.00")
            opening_counted = False

            for b in biz_list:
                ledger_rows, _totals_b, _entity = build_ledger(
                    kind=kind,
                    business_id=b.id,
                    entity_id=p.id,
                    date_from=date_from,
                    date_to=date_to,
                )

                if not ledger_rows:
                    continue

                # remove pending cheque effects
                filtered_b = _filter_cheque_payments_from_rows(
                    ledger_rows,
                    b.id,
                    p.id,
                    exclude_pending=True,
                )

                if not filtered_b:
                    continue

                # separate opening rows for this business
                cleaned_rows, ob_dr, ob_cr = _extract_opening(filtered_b)

                if not opening_counted:
                    stats = _recalculate_totals_excluding_pending(
                        cleaned_rows,
                        ob_dr,
                        ob_cr,
                    )
                    if ob_dr > 0 or ob_cr > 0:
                        opening_counted = True
                else:
                    # later businesses. ignore their opening lines
                    stats = _recalculate_totals_excluding_pending(
                        cleaned_rows,
                        Decimal("0.00"),
                        Decimal("0.00"),
                    )

                total_dr += stats.get("total_dr", Decimal("0.00"))
                total_cr += stats.get("total_cr", Decimal("0.00"))

            balance = total_dr - total_cr  # Dr minus Cr

            if balance == 0:
                # skip zero remaining parties
                continue

            rows.append(
                {
                    "party": p,
                    "balance_abs": abs(balance),
                    "balance_side": "Dr" if balance >= 0 else "Cr",
                }
            )

        # sort by largest remaining
        rows.sort(key=lambda r: r["balance_abs"], reverse=True)
        return rows

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # party summary filters from querystring
        request = self.request
        party_kind = (request.GET.get("kind") or "customer").strip().lower()
        if party_kind not in ("customer", "supplier"):
            party_kind = "customer"

        party_q = (request.GET.get("q") or "").strip()
        date_from = _parse_date(request.GET.get("date_from"))
        date_to = _parse_date(request.GET.get("date_to"))

        if date_from and date_to and date_from > date_to:
            date_from, date_to = date_to, date_from

        party_rows = self._build_party_summary(
            kind=party_kind,
            q=party_q,
            date_from=date_from,
            date_to=date_to,
        )

        ctx["csrf_token"] = get_token(self.request)
        ctx["party_kind"] = party_kind
        ctx["party_rows"] = party_rows
        ctx["party_q"] = party_q
        ctx["party_date_from"] = date_from
        ctx["party_date_to"] = date_to
        return ctx

@login_required
def recalculate_all_totals_view(request):
    """
    Trigger the recalculate_stats management command logic via UI.
    Requires user to be logged in. 
    Ideally this should be protected by a permission or only for staff/admin.
    """
    try:
        call_command('recalculate_stats')
        messages.success(request, "All financial totals and balances have been successfully recalculated.")
    except Exception as e:
        messages.error(request, f"Error during recalculation: {str(e)}")
    
    return redirect(request.META.get('HTTP_REFERER', 'business'))


@login_required
def financial_summary_view(request):
    """
    Financial Period Summary Report - Report Style
    Matches the lo-fi mockup with series ranges and tabular bank details.
    """
    date_str = request.GET.get('date')
    if date_str:
        try:
            today = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            today = timezone.localdate()
    else:
        today = timezone.localdate()

    # --- SECTION 1: Performance Headers (Report Style) ---
    
    # Total Sales
    sales_qs = SalesOrder.objects.filter(
        business__is_deleted=False,
        is_deleted=False,
        created_at__date=today
    )
    fulfilled_sales = sales_qs.filter(status=SalesOrder.Status.FULFILLED)
    total_sales = fulfilled_sales.aggregate(s=Sum('net_total'))['s'] or Decimal('0.00')
    sales_count = fulfilled_sales.count()
    
    # Receipt Series (e.g., "Receipt #2 to #4")
    sales_ids = fulfilled_sales.values_list('id', flat=True).order_by('id')
    if sales_ids:
        sales_series = f"SO #{sales_ids[0]} to #{sales_ids[len(sales_ids)-1]}"
    else:
        sales_series = "—"
    
    # Total Receipt (sum of all receipts applied to sales)
    total_receipt = SalesOrderReceipt.objects.filter(
        sales_order__in=fulfilled_sales,
        sales_order__created_at__date=today
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    # Cancelled Sales
    cancelled_sales = sales_qs.filter(status=SalesOrder.Status.CANCELLED)
    total_cancelled = cancelled_sales.aggregate(s=Sum('net_total'))['s'] or Decimal('0.00')
    cancelled_count = cancelled_sales.count()
    
    # Total Purchase
    po_qs = PurchaseOrder.objects.filter(
        business__is_deleted=False,
        is_deleted=False,
        is_active=True,
        created_at__date=today
    )
    total_purchase = po_qs.aggregate(s=Sum('net_total'))['s'] or Decimal('0.00')
    
    # Purchase Series (e.g., "PO #1 to #4")
    po_ids = po_qs.values_list('id', flat=True).order_by('id')
    if po_ids:
        po_series = f"PO #{po_ids[0]} to #{po_ids[len(po_ids)-1]}"
    else:
        po_series = "—"
    
    # Pending POs
    pending_po_count = po_qs.filter(status='pending').count()
    
    # Total Expenses (split into Landed PO vs Operating)
    expenses_qs = Expense.objects.filter(
        business__is_deleted=False,
        is_deleted=False,
        date=today
    )
    
    # Landed PO: Expenses linked to Purchase Orders
    landed_po_expense = expenses_qs.filter(purchase_order__isnull=False).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    # Operating: All other expenses
    operating_expense = expenses_qs.filter(purchase_order__isnull=True).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    total_expenses = landed_po_expense + operating_expense

    # --- SECTION 2: Amount IN (Formula Bar) ---
    
    # Cash Via Sales Deposited (Cash sales)
    cash_sales = Payment.objects.filter(
        Q(direction=Payment.IN),
        Q(payment_method=Payment.PaymentMethod.CASH),
        Q(applied_sales_orders__isnull=False) | Q(applied_sales_invoices__isnull=False),
        date=today,
        is_deleted=False
    ).distinct().aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    # Cash Via Receipt (Other cash receipts/collections)
    cash_receipt = Payment.objects.filter(
        direction=Payment.IN,
        payment_method=Payment.PaymentMethod.CASH,
        applied_sales_orders__isnull=True,
        applied_sales_invoices__isnull=True,
        applied_purchase_returns__isnull=True,
        date=today,
        is_deleted=False
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    # Total Cash In
    total_cash_in = cash_sales + cash_receipt
    
    # Sales Via Bank (Bank sales)
    bank_sales = Payment.objects.filter(
        Q(direction=Payment.IN),
        Q(payment_method__in=[Payment.PaymentMethod.BANK, Payment.PaymentMethod.CARD]),
        Q(applied_sales_orders__isnull=False) | Q(applied_sales_invoices__isnull=False),
        date=today,
        is_deleted=False
    ).distinct().aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    # Receipt Via Bank (Other bank receipts)
    bank_receipt = Payment.objects.filter(
        direction=Payment.IN,
        payment_method__in=[Payment.PaymentMethod.BANK, Payment.PaymentMethod.CARD],
        applied_sales_orders__isnull=True,
        applied_sales_invoices__isnull=True,
        applied_purchase_returns__isnull=True,
        date=today,
        is_deleted=False
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    # Total Bank deposits
    total_bank_deposits = bank_sales + bank_receipt

    # --- SECTION 3: Amount OUT (Formula Bar) ---
    
    # Purchase Order payments
    po_payments = Payment.objects.filter(
        direction=Payment.OUT,
        applied_purchase_orders__isnull=False,
        date=today,
        is_deleted=False
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    # General Payments (not tied to PO or SR)
    general_payments = Payment.objects.filter(
        direction=Payment.OUT,
        applied_purchase_orders__isnull=True,
        applied_sales_returns__isnull=True,
        date=today,
        is_deleted=False
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    # Sale Return Refunds
    sr_refunds = Payment.objects.filter(
        direction=Payment.OUT,
        applied_sales_returns__isnull=False,
        date=today,
        is_deleted=False
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    # Total Cash Out
    total_cash_out = po_payments + general_payments + sr_refunds + total_expenses

    # --- SECTION 4: Bank Details by Account (Table) ---
    
    bank_accounts = BankAccount.objects.filter(is_active=True, is_deleted=False)
    bank_details = []
    
    for acc in bank_accounts:
        # Bank Sales Amount (sales deposited to this bank)
        bank_sales_amount = Payment.objects.filter(
            bank_account=acc,
            direction=Payment.IN,
            date=today,
            is_deleted=False
        ).filter(
            Q(applied_sales_orders__isnull=False) | Q(applied_sales_invoices__isnull=False)
        ).distinct().aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
        
        # Deposited (Cash) - manual cash deposits to this bank (CashFlow IN not from payments)
        # This captures direct deposits to bank that aren't from sales/receipts
        total_bank_in = CashFlow.objects.filter(
            bank_account=acc,
            flow_type=CashFlow.IN,
            date=today,
            is_deleted=False,
            linked_payment__isnull=True  # Not from a payment transaction
        ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
        
        # Manual deposits = Total IN - Bank Sales Amount
        cash_deposited = max(Decimal('0.00'), total_bank_in - bank_sales_amount)
        
        # Cheque Deposited (cleared cheques to this bank)
        cheque_deposited = Payment.objects.filter(
            bank_account=acc,
            payment_method=Payment.PaymentMethod.CHEQUE,
            cheque_status=Payment.ChequeStatus.DEPOSITED,
            date=today,
            is_deleted=False
        ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
        
        # Total Deposited
        total_deposited = bank_sales_amount + cash_deposited + cheque_deposited
        
        # Current Balance (from CashFlow)
        total_flow = CashFlow.objects.filter(
            bank_account=acc,
            date__lte=today,
            is_deleted=False
        ).aggregate(
            t=Sum(Case(
                When(flow_type=CashFlow.IN, then=F('amount')),
                When(flow_type=CashFlow.OUT, then=-F('amount')),
                default=Decimal('0.00'),
                output_field=DecimalField()
            ))
        )['t'] or Decimal('0.00')
        current_balance = acc.opening_balance + total_flow
        
        bank_details.append({
            'account': acc,
            'bank_sales_amount': bank_sales_amount,
            'cash_deposited': cash_deposited,
            'cheque_deposited': cheque_deposited,
            'total_deposited': total_deposited,
            'current_balance': current_balance
        })
    
    # Grand Total of All Banks
    grand_total_banks = sum(b['current_balance'] for b in bank_details)
    
    # --- SECTION 5: Amount In Hand Summary ---
    
    # Cash in Hand (physical cash not deposited)
    cash_in_hand = total_cash_in - total_cash_out
    
    # Cheque in Hand (pending cheques)
    cheques_pending = Payment.objects.filter(
        payment_method=Payment.PaymentMethod.CHEQUE,
        cheque_status=Payment.ChequeStatus.PENDING,
        is_deleted=False
    ).aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    
    # All Bank Balance
    all_bank_balance = grand_total_banks

    context = {
        'today': today,
        
        # Sales Card
        'total_sales': total_sales,
        'sales_count': sales_count,
        'sales_series': sales_series,
        'total_receipt': total_receipt,
        'total_cancelled': total_cancelled,
        'cancelled_count': cancelled_count,
        
        # Purchase Card
        'total_purchase': total_purchase,
        'po_series': po_series,
        'pending_po_count': pending_po_count,
        
        # Expense Card
        'landed_po_expense': landed_po_expense,
        'operating_expense': operating_expense,
        'total_expenses': total_expenses,
        
        # Amount IN
        'cash_sales': cash_sales,
        'cash_receipt': cash_receipt,
        'total_cash_in': total_cash_in,
        'bank_sales': bank_sales,
        'bank_receipt': bank_receipt,
        'total_bank_deposits': total_bank_deposits,
        
        # Amount OUT
        'po_payments': po_payments,
        'general_payments': general_payments,
        'sr_refunds': sr_refunds,
        'total_cash_out': total_cash_out,
        
        # Bank Details
        'bank_details': bank_details,
        'grand_total_banks': grand_total_banks,
        
        # Amount In Hand
        'cash_in_hand': cash_in_hand,
        'cheques_pending': cheques_pending,
        'all_bank_balance': all_bank_balance,
    }
    
    return render(request, "barkat/finance/financial_summary.html", context)

# ---------- LIST ----------
@method_decorator(login_required, name="dispatch")
class BusinessesListView(ListView):
    template_name = "barkat/dashboard/businesses.html"
    context_object_name = "businesses"
    paginate_by = 25  # optional

    def get_queryset(self):
        return Business.objects.filter(is_deleted=False).order_by("-id")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["csrf_token"] = get_token(self.request)
        return ctx

# ---------- ADD PAGE (GET shows form, POST saves & redirects) ----------
@login_required
def business_add_page(request):
    if request.method == "POST":
        form = BusinessForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.save()
            messages.success(request, "Business created successfully.")
            return redirect("businesses_list")
    else:
        form = BusinessForm()
    return render(
        request,
        "barkat/dashboard/add_edit.html",
        {"form": form, "mode": "create", "title": "New Business", "submit_label": "Create"},
    )

# ---------- EDIT PAGE (GET shows form, POST updates & redirects) ----------
@login_required
def business_edit_page(request, pk):
    biz = get_object_or_404(Business, pk=pk, is_deleted=False)
    if request.method == "POST":
        form = BusinessForm(request.POST, instance=biz)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = request.user
            obj.save()
            messages.success(request, "Business updated successfully.")
            return redirect("businesses_list")
    else:
        form = BusinessForm(instance=biz)
    return render(
        request,
        "barkat/dashboard/add_edit.html",
        {"form": form, "mode": "edit", "title": "Edit Business", "submit_label": "Save Changes"},
    )

# ---------- JSON (still available if needed) ----------
@login_required
def business_json(request, pk):
    biz = get_object_or_404(Business, pk=pk, is_deleted=False)
    return JsonResponse({
        "id": biz.id,
        "code": biz.code,
        "name": biz.name,
        "legal_name": biz.legal_name,
        "ntn": biz.ntn,
        "sales_tax_reg": biz.sales_tax_reg,
        "phone": biz.phone,
        "email": biz.email,
        "address": biz.address,
        "is_active": biz.is_active,
    })

# ---------- CREATE (supports AJAX JSON or normal POST redirect) ----------
@login_required
@require_http_methods(["POST"])
def business_create(request):
    form = BusinessForm(request.POST)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.created_by = request.user
        obj.save()
        messages.success(request, "Business created successfully.")
        # If AJAX, return JSON; else redirect
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"ok": True, "id": obj.id})
        return redirect("businesses_list")
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"ok": False, "errors": form.errors}, status=400)
    return render(request, "barkat/dashboard/add_edit.html", {
        "form": form, "mode": "create", "title": "New Business", "submit_label": "Create"
    })

# ---------- UPDATE (supports AJAX JSON or normal POST redirect) ----------

@require_http_methods(["POST"])
def business_update(request, pk):
    biz = get_object_or_404(Business, pk=pk, is_deleted=False)
    form = BusinessForm(request.POST, instance=biz)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.updated_by = request.user
        obj.save()
        messages.success(request, "Business updated successfully.")
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"ok": True})
        return redirect("businesses_list")
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"ok": False, "errors": form.errors}, status=400)
    return render(request, "barkat/dashboard/add_edit.html", {
        "form": form, "mode": "edit", "title": "Edit Business", "submit_label": "Save Changes"
    })

# ---------- DELETE (soft) — supports AJAX JSON or normal POST redirect ----------
@require_POST
@login_required
def business_delete(request, pk):
    biz = get_object_or_404(Business, pk=pk, is_deleted=False)
    biz.is_deleted = True
    biz.updated_by = request.user
    biz.save(update_fields=["is_deleted", "updated_by", "updated_at"])
    messages.success(request, "Business deleted.")

    # Support ajax delete
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"ok": True})

    return redirect("businesses_list")

# ---------- Parties ----------
class CustomersListView(ListView):
    template_name = "barkat/parties/customers_list.html"
    context_object_name = "customers"

    def get_queryset(self):
        qs = Party.objects.filter(type__in=[Party.CUSTOMER, Party.BOTH], is_deleted=False).order_by("display_name")
        q = self.request.GET.get("q")
        if q:
          qs = qs.filter(Q(display_name__icontains=q) | Q(phone__icontains=q) | Q(email__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["businesses"] = Business.objects.filter(is_deleted=False).order_by("name")
        return ctx

@method_decorator(login_required, name="dispatch")
class BusinessCustomersListView(ListView):
    """
    Lists ONLY the customers that belong to a single Business (via Party.default_business).
    Renders templates/barkat/parties/business_customer.html
    """
    template_name = "barkat/parties/business_customer.html"
    context_object_name = "customers"

    def get_queryset(self):
        business_id = self.kwargs["business_id"]
        self.business = get_object_or_404(Business, pk=business_id, is_deleted=False)

        qs = Party.objects.filter(
            type__in=[Party.CUSTOMER, Party.BOTH],
            is_deleted=False,
            default_business_id=self.business.id,
        ).order_by("display_name")

        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(
                Q(display_name__icontains=q)
                | Q(phone__icontains=q)
                | Q(email__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # current business (for heading) + all businesses (to render the colored nav links)
        ctx["business"] = self.business
        ctx["businesses"] = Business.objects.filter(is_deleted=False).order_by("name")
        return ctx

class VendorsListView(LoginRequiredMixin, ListView):
    template_name = "barkat/parties/vendors_list.html"   # <-- fixed
    context_object_name = "vendors"
    login_url = "login"          # or rely on settings.LOGIN_URL
    redirect_field_name = "next"

    def get(self, request, *args, **kwargs):
        from barkat.utils.auth_helpers import user_has_cancellation_password
        if (
            user_has_cancellation_password(request)
            and not request.session.get("supplier_ledger_unlocked")
        ):
            return render(
                request,
                "barkat/finance/password_gate.html",
                {
                    "gate_title": "Vendors / suppliers list",
                    "gate_message": "Viewing the vendors list requires your cancellation password (User Settings). Enter it to continue, or Cancel to go back to customers.",
                    "cancel_url": reverse("customers_list"),
                    "next_url": reverse("vendors_list"),
                    "action": "supplier_ledger",
                },
            )
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = Party.objects.filter(
            type__in=[Party.VENDOR, Party.BOTH],
            is_deleted=False
        ).order_by("display_name")
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(
                Q(display_name__icontains=q) |
                Q(phone__icontains=q) |
                Q(email__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["businesses"] = Business.objects.filter(
            is_deleted=False, is_active=True
        ).order_by("name")
        return ctx

@method_decorator(login_required, name="dispatch")
class BusinessVendorsView(ListView):
    template_name = "barkat/parties/business_vendor.html"
    context_object_name = "vendors"

    def dispatch(self, request, *args, **kwargs):
        self.business = get_object_or_404(Business, pk=kwargs["business_id"], is_deleted=False)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        from barkat.utils.auth_helpers import user_has_cancellation_password
        if (
            user_has_cancellation_password(request)
            and not request.session.get("supplier_ledger_unlocked")
        ):
            return render(
                request,
                "barkat/finance/password_gate.html",
                {
                    "gate_title": "Vendors / suppliers list",
                    "gate_message": "Viewing the vendors list requires your cancellation password (User Settings). Enter it to continue, or Cancel to go back to customers.",
                    "cancel_url": reverse("customers_list"),
                    "next_url": reverse("business_vendors", args=[kwargs["business_id"]]),
                    "action": "supplier_ledger",
                },
            )
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = Party.objects.filter(
            type__in=[Party.VENDOR, Party.BOTH],
            default_business=self.business,     # filter to that business
            is_deleted=False
        ).order_by("display_name")
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(
                Q(display_name__icontains=q) |
                Q(phone__icontains=q) |
                Q(email__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["business"] = self.business
        ctx["businesses"] = Business.objects.filter(is_deleted=False, is_active=True).order_by("name")
        return ctx

class PartyDetailView(DetailView):
    model = Party
    template_name = "barkat/parties/party_detail.html"
    context_object_name = "party"

class PartyCreateView(CreateView):
    model = Party
    form_class = PartyForm
    template_name = "barkat/parties/party_form.html"
    success_url = reverse_lazy("customers_list")  # final fallback

    # --- Helpers ---
    def _get_requested_type(self):
        t = (self.request.GET.get("type") or "").lower()
        if t in ("vendor", "v"):
            return Party.VENDOR
        return Party.CUSTOMER

    def _type_label(self, party_type):
        return "Vendor" if party_type == Party.VENDOR else "Customer"

    # new. send fixed_type to form so it can disable the field
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["fixed_type"] = self._get_requested_type()
        return kwargs

    # --- Initial form data ---
    def get_initial(self):
        initial = super().get_initial()
        biz_id = self.request.GET.get("business")
        if biz_id:
            initial["default_business"] = biz_id
        initial["type"] = self._get_requested_type()
        return initial

    # --- Save ---
    def form_valid(self, form):
        obj = form.save(commit=False)

        # enforce type from query regardless of posted value
        obj.type = self._get_requested_type()

        if not obj.default_business:
            biz_id = self.request.GET.get("business")
            if biz_id:
                try:
                    obj.default_business = Business.objects.get(pk=biz_id)
                except Business.DoesNotExist:
                    pass

        user = self.request.user if self.request.user.is_authenticated else None
        obj.created_by = user
        obj.updated_by = user

        obj.save()
        self.object = obj

        messages.success(
            self.request,
            f"{self._type_label(obj.type)} created successfully."
        )
        return redirect(self.get_success_url())

class PartyUpdateView(UpdateView):
    model = Party
    form_class = PartyForm
    template_name = "barkat/parties/party_form.html"
    success_url = reverse_lazy("customers_list")

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.updated_by = self.request.user if self.request.user.is_authenticated else None
        obj.save()
        messages.success(self.request, "Party updated successfully.")
        return super().form_valid(form)

@login_required
@require_POST
def party_delete(request, pk):
    party = get_object_or_404(Party, pk=pk, is_deleted=False)
    party.is_deleted = True
    # If you have these fields, uncomment:
    # party.deleted_by = request.user
    # party.deleted_at = timezone.now()
    party.save()

    messages.success(request, f"Deleted customer: {party.display_name}")
    next_url = request.POST.get("next") or request.GET.get("next")
    return redirect(next_url or "customers_list")

# ---------- Catalog ----------
class ProductFilterMixin:
    """Consolidate product filtering and valuation logic."""
    def get_product_queryset(self, request, base_qs=None):
        if base_qs is None:
            base_qs = Product.objects.filter(is_deleted=False)
            
        qs = base_qs.select_related("business", "category", "uom", "bulk_uom").annotate(
            total_stock_value=ExpressionWrapper(
                Coalesce(F("purchase_price"), V(0)) * Coalesce(F("stock_qty"), V(0)),
                output_field=DecimalField(max_digits=18, decimal_places=2)
            )
        ).order_by("-id")

        q = request.GET.get("q")
        biz_id = request.GET.get("business")
        
        if q:
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(sku__icontains=q) |
                Q(barcode__icontains=q) |
                Q(category__name__icontains=q) |
                Q(business__name__icontains=q) |
                Q(company_name__icontains=q)
            )
        
        # Only apply global business filter if biz_id is present and we're not already filtered
        if biz_id and not hasattr(self, 'business'):
            qs = qs.filter(business_id=biz_id)
            
        # Price filter
        price_op = request.GET.get("price_op")
        price_val = request.GET.get("price_val")
        if price_op and price_val:
            try:
                price_decimal = Decimal(price_val)
                if price_op == "gte":
                    qs = qs.filter(sale_price__gte=price_decimal)
                elif price_op == "lte":
                    qs = qs.filter(sale_price__lte=price_decimal)
                elif price_op == "eq":
                    qs = qs.filter(sale_price=price_decimal)
            except (ValueError, InvalidOperation):
                pass
        
        # Stock filter
        stock_op = request.GET.get("stock_op")
        stock_val = request.GET.get("stock_val")
        if stock_op and stock_val:
            try:
                stock_decimal = Decimal(stock_val)
                if stock_op == "gte":
                    qs = qs.filter(stock_qty__gte=stock_decimal)
                elif stock_op == "lte":
                    qs = qs.filter(stock_qty__lte=stock_decimal)
                elif stock_op == "eq":
                    qs = qs.filter(stock_qty=stock_decimal)
            except (ValueError, InvalidOperation):
                pass
        
        return qs

    def get_grand_total_stock_value(self, qs):
        return qs.aggregate(total=Sum("total_stock_value"))["total"] or Decimal("0.00")

# All categories
class ProductCategoriesListView(LoginRequiredMixin, ListView):
    template_name = "barkat/catalog/categories_list.html"
    context_object_name = "categories"
    login_url = "login"

    def get_queryset(self):
        qs = (ProductCategory.objects
              .filter(is_deleted=False)
              .select_related("business", "parent")
              .order_by("business__name", "name"))
        q   = self.request.GET.get("q")
        biz = self.request.GET.get("business")
        if q:
            qs = qs.filter(
                Q(code__icontains=q) |
                Q(name__icontains=q) |
                Q(business__name__icontains=q) |
                Q(parent__name__icontains=q)
            )
        if biz:
            qs = qs.filter(business_id=biz)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["businesses"] = Business.objects.filter(is_deleted=False, is_active=True).order_by("name")
        return ctx

class ProductCategoryCreateView(LoginRequiredMixin, CreateView):
    model = ProductCategory
    form_class = ProductCategoryForm
    template_name = "barkat/catalog/category_form.html"
    success_url = reverse_lazy("product_categories")  # final fallback
    login_url = "login"

    def get_initial(self):
        initial = super().get_initial()
        biz_id = self.request.GET.get("business")
        if biz_id:
            # Assumes your ProductCategory has a FK named "business"
            initial["business"] = biz_id
        return initial

    def form_valid(self, form):
        obj = form.save(commit=False)

        # If "business" wasn't selected in the form, take it from ?business=
        if not getattr(obj, "business_id", None):
            biz_id = self.request.GET.get("business")
            if biz_id:
                try:
                    obj.business = Business.objects.get(pk=biz_id)
                except Business.DoesNotExist:
                    pass

        # Audit fields
        user = self.request.user if self.request.user.is_authenticated else None
        if hasattr(obj, "created_by"):
            obj.created_by = user
        if hasattr(obj, "updated_by"):
            obj.updated_by = user

        obj.save()
        self.object = obj  # make sure get_success_url can use it

        messages.success(self.request, "Category created successfully.")
        return redirect(self.get_success_url())

    def get_success_url(self):
        # 1) honor ?next=
        nxt = self.request.GET.get("next")
        if nxt:
            return nxt

        # 2) if ?business= provided, go to that business’s category list
        biz_id = self.request.GET.get("business")
        if biz_id:
            return reverse("business_categories", args=[biz_id])

        # 3) fallback
        return super().get_success_url()

class ProductCategoryUpdateView(LoginRequiredMixin, UpdateView):
    model = ProductCategory
    form_class = ProductCategoryForm
    template_name = "barkat/catalog/category_form.html"
    login_url = "login"
    success_url = reverse_lazy("product_categories")

    def get_queryset(self):
        # Don’t allow editing deleted categories
        return ProductCategory.objects.filter(is_deleted=False).select_related("business", "parent")

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.updated_by = self.request.user
        obj.save()
        messages.success(self.request, "Category updated successfully.")
        return super().form_valid(form)

# ---------- DELETE (SOFT) ----------
class ProductCategoryDeleteView(LoginRequiredMixin, View):
    login_url = "login"

    def get(self, request, pk):
        category = get_object_or_404(ProductCategory, pk=pk, is_deleted=False)
        return render(
            request,
            "barkat/catalog/category_confirm_delete.html",
            {"category": category},
        )

    def post(self, request, pk):
        category = get_object_or_404(ProductCategory, pk=pk, is_deleted=False)
        category.is_deleted = True
        category.updated_by = request.user
        category.save(update_fields=["is_deleted", "updated_by", "updated_at"])
        messages.success(request, f"Category “{category.name}” deleted.")
        return redirect("product_categories")

# Categories limited to one business
class BusinessCategoriesListView(LoginRequiredMixin, ListView):
    template_name = "barkat/catalog/business_categories_list.html"
    context_object_name = "categories"
    login_url = "login"

    def dispatch(self, request, *args, **kwargs):
        self.business = get_object_or_404(Business, pk=kwargs["business_id"], is_deleted=False)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        qs = (ProductCategory.objects
              .filter(is_deleted=False, business=self.business)
              .select_related("business", "parent")
              .order_by("name"))
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(
                Q(code__icontains=q) |
                Q(name__icontains=q) |
                Q(parent__name__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["business"] = self.business
        ctx["businesses"] = Business.objects.filter(is_deleted=False, is_active=True).order_by("name")
        return ctx

# PRODUCT Create/Update (as you had)
class ProductsListView(LoginRequiredMixin, ProductFilterMixin, ListView):
    template_name = "barkat/catalog/products_list.html"
    context_object_name = "products"
    login_url = "login"
    paginate_by = None  # Disable server-side pagination, use client-side pagination

    def get_queryset(self):
        base_qs = Product.objects.filter(is_deleted=False)
        return self.get_product_queryset(self.request, base_qs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["businesses"] = Business.objects.filter(is_deleted=False, is_active=True).order_by("name")
        ctx["q"] = self.request.GET.get("q", "")
        ctx["business_filter"] = self.request.GET.get("business", "")
        
        # Calculate grand total for the filtered queryset
        ctx["grand_total_stock_value"] = self.get_grand_total_stock_value(self.object_list)
        return ctx


@require_GET
@login_required
def export_products_csv(request):
    """Export products to CSV with filters applied."""
    import csv
    from django.http import HttpResponse
    
    mixin = ProductFilterMixin()
    qs = mixin.get_product_queryset(request)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="products_export.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([
        'ID', 'Name', 'Company', 'SKU', 'Barcode', 'Business', 
        'Category', 'Purchase Price', 'Sale Price', 'Current Stock', 
        'UOM', 'Stock Value', 'Is Active'
    ])
    
    # Write data rows
    for p in qs:
        writer.writerow([
            p.id,
            p.name,
            p.company_name or '',
            p.sku or '',
            p.barcode or '',
            p.business.name if p.business else '',
            p.category.name if p.category else '',
            str(p.purchase_price or '0.00'),
            str(p.sale_price or '0.00'),
            str(p.stock_qty or '0.00'),
            p.uom.code if p.uom else '',
            str(p.total_stock_value or '0.00'),
            'Yes' if p.is_active else 'No'
        ])
    
    # Add grand total row
    grand_total = mixin.get_grand_total_stock_value(qs)
    writer.writerow([])
    writer.writerow(['', '', '', '', '', '', '', '', '', '', 'GRAND TOTAL STOCK VALUE', str(grand_total)])
    
    return response


@require_GET
@login_required
def export_business_products_csv(request, business_id):
    """Export business products to CSV with filters applied."""
    import csv
    from django.http import HttpResponse
    
    business = get_object_or_404(Business, pk=business_id, is_deleted=False)
    mixin = ProductFilterMixin()
    base_qs = Product.objects.filter(is_deleted=False, business=business)
    qs = mixin.get_product_queryset(request, base_qs=base_qs)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"products_{business.name.replace(' ', '_')}_export.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([
        'ID', 'Name', 'Company', 'SKU', 'Barcode', 
        'Category', 'Purchase Price', 'Sale Price', 'Current Stock', 
        'UOM', 'Stock Value', 'Is Active'
    ])
    
    # Write data rows
    for p in qs:
        writer.writerow([
            p.id,
            p.name,
            p.company_name or '',
            p.sku or '',
            p.barcode or '',
            p.category.name if p.category else '',
            str(p.purchase_price or '0.00'),
            str(p.sale_price or '0.00'),
            str(p.stock_qty or '0.00'),
            p.uom.code if p.uom else '',
            str(p.total_stock_value or '0.00'),
            'Yes' if p.is_active else 'No'
        ])
    
    # Add grand total row
    grand_total = mixin.get_grand_total_stock_value(qs)
    writer.writerow([])
    writer.writerow(['', '', '', '', '', '', '', '', '', 'GRAND TOTAL STOCK VALUE', str(grand_total)])
    
    return response

class BusinessProductsListView(LoginRequiredMixin, ProductFilterMixin, ListView):
    template_name = "barkat/catalog/business_products.html"
    context_object_name = "products"
    login_url = "login"
    paginate_by = None  # Disable server-side pagination, use client-side pagination

    def dispatch(self, request, *args, **kwargs):
        self.business = get_object_or_404(Business, pk=kwargs["business_id"], is_deleted=False)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        base_qs = Product.objects.filter(is_deleted=False, business=self.business)
        return self.get_product_queryset(self.request, base_qs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["business"] = self.business
        ctx["businesses"] = Business.objects.filter(is_deleted=False, is_active=True).order_by("name")
        ctx["q"] = self.request.GET.get("q", "")
        
        # Calculate grand total for the filtered queryset
        ctx["grand_total_stock_value"] = self.get_grand_total_stock_value(self.object_list)
        return ctx

class ProductCreateView(LoginRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = "barkat/catalog/product_form.html"
    login_url = "login"
    success_url = reverse_lazy("products_list")

    def get_initial(self):
        initial = super().get_initial()
        biz_id = self.request.GET.get("business")
        if biz_id:
            initial["business"] = biz_id
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        
        # Products Data for JS (Updated with Bulk Logic)
        products_qs = Product.objects.filter(is_active=True, is_deleted=False).select_related("uom", "bulk_uom", "category")
        products_cards = []
        for p in products_qs:
            products_cards.append({
                "id": p.id,
                "name": p.name,
                "sale_price": str(p.sale_price),
                "category_id": p.category_id or "",
                "image_url": _product_image_url(p),
                "stock": str(p.stock_qty or 0),
                "barcode": getattr(p, "barcode", "") or "",
                "base_uom_id": p.uom_id,
                "base_uom_name": p.uom.code if p.uom else "",
                "bulk_uom_id": p.bulk_uom_id or "",
                "bulk_uom_name": p.bulk_uom.code if p.bulk_uom else "",
                "default_bulk_size": str(p.default_bulk_size or 1),
            })
        ctx["products_cards"] = products_cards

        if self.request.method == "POST":
            ctx["image_formset"] = ProductImageFormSet(self.request.POST, self.request.FILES, prefix="productimage_set")
        else:
            ctx["image_formset"] = ProductImageFormSet(prefix="productimage_set")
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        obj = form.save(commit=False)
        if not getattr(obj, "business_id", None):
            biz_id = self.request.GET.get("business")
            if biz_id:
                obj.business = Business.objects.filter(pk=biz_id).first()

        obj.created_by = self.request.user
        obj.updated_by = self.request.user
        obj.save()
        self.object = obj
        form.save_m2m()

        # Handle Images
        image_formset = ProductImageFormSet(self.request.POST, self.request.FILES, instance=self.object, prefix="productimage_set")
        if image_formset.is_valid():
            image_formset.save()

        # For ProductCreateView, always redirect back to form for "Save & Add Another"
        # Since we only have one button for new products, always go back to the form
        messages.success(self.request, f"Product '{obj.name}' created successfully. You can add another product below.")
        # Get business ID from GET, POST, or from the saved object
        biz_id = self.request.GET.get("business") or self.request.POST.get("business") or (obj.business_id if hasattr(obj, 'business_id') and obj.business_id else None)
        
        # Build redirect URL - ensure we're using the correct URL pattern
        redirect_url = reverse('product_create')
        if biz_id:
            redirect_url = f"{redirect_url}?business={biz_id}"
        
        # Use HttpResponseRedirect to ensure proper redirect handling
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(redirect_url)

    def get_success_url(self):
        biz_id = self.request.GET.get("business")
        if biz_id: return reverse("business_products", args=[biz_id])
        return super().get_success_url()

class ProductUpdateView(LoginRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = "barkat/catalog/product_form.html"
    login_url = "login"
    success_url = reverse_lazy("products_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Same products_cards logic as CreateView...
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.updated_by = self.request.user
        obj.save()
        form.save_m2m()

        image_formset = ProductImageFormSet(self.request.POST, self.request.FILES, instance=self.object, prefix="productimage_set")
        if image_formset.is_valid():
            image_formset.save()

        messages.success(self.request, "Product updated successfully.")
        return redirect(self.get_success_url())

class ProductDeleteView(LoginRequiredMixin, View):
    login_url = "login"

    def get(self, request, pk):
        product = get_object_or_404(Product, pk=pk, is_deleted=False)
        return render(
            request,
            "barkat/catalog/product_confirm_delete.html",
            {"product": product},
        )

    def post(self, request, pk):
        product = get_object_or_404(Product, pk=pk, is_deleted=False)
        product.is_deleted = True
        product.updated_by = request.user
        product.save(update_fields=["is_deleted", "updated_by", "updated_at"])
        messages.success(request, f'Product “{product.name}” deleted.')
        return redirect("products_list")


# STAFF
class StaffListView(LoginRequiredMixin, ListView):
    model = Staff
    template_name = "barkat/staff/staff_list.html"
    context_object_name = "staff_list"

    def get_queryset(self):
        # All staff (no business filter)
        qs = Staff.objects.select_related("business", "user").order_by("business__code", "full_name")
        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(
                Q(full_name__icontains=q) |
                Q(phone__icontains=q) |
                Q(cnic__icontains=q) |
                Q(user__username__icontains=q) |
                Q(user__email__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["businesses"] = Business.objects.order_by("code")
        ctx["current_business"] = None  # explicitly none on "All" page
        ctx["q"] = self.request.GET.get("q", "")
        ctx["roles"] = Staff.Roles.choices
        return ctx

class BusinessStaffListView(LoginRequiredMixin, ListView):
    model = Staff
    template_name = "barkat/staff/business_staff.html"
    context_object_name = "staff_list"

    def dispatch(self, request, *args, **kwargs):
        self.business = get_object_or_404(Business, id=self.kwargs["business_id"])
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        qs = (
            Staff.objects
            .select_related("business", "user")
            .filter(business=self.business)
            .order_by("full_name")
        )
        q = self.request.GET.get("q", "").strip()
        role = self.request.GET.get("role", "").strip()
        if q:
            qs = qs.filter(
                Q(full_name__icontains=q) |
                Q(phone__icontains=q) |
                Q(cnic__icontains=q) |
                Q(user__username__icontains=q) |
                Q(user__email__icontains=q)
            )
        if role:
            qs = qs.filter(role=role)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["business"] = self.business
        ctx["businesses"] = Business.objects.order_by("code")
        ctx["current_business"] = self.business.id
        ctx["q"] = self.request.GET.get("q", "")
        ctx["current_role"] = self.request.GET.get("role", "")
        ctx["roles"] = Staff.Roles.choices
        return ctx

class StaffCreateView(LoginRequiredMixin, CreateView):
    model = Staff
    form_class = StaffForm
    template_name = "barkat/staff/staff_form.html"
    success_url = reverse_lazy("staff_list")

    def get_initial(self):
        initial = super().get_initial()
        if self.request.GET.get("business"):
            initial["business"] = self.request.GET["business"]
        return initial

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.created_by = self.request.user
        obj.updated_by = self.request.user
        obj.save()
        self.object = obj  # important for get_success_url
        messages.success(self.request, "Staff member created successfully.")
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        messages.error(self.request, "Please fix the errors below.")
        return super().form_invalid(form)

class StaffUpdateView(LoginRequiredMixin, UpdateView):
    model = Staff
    form_class = StaffForm
    template_name = "barkat/staff/staff_form.html"
    success_url = reverse_lazy("staff_list")

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.updated_by = self.request.user
        obj.save()
        self.object = obj
        messages.success(self.request, "Staff member updated successfully.")
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        messages.error(self.request, "Please fix the errors below.")
        return super().form_invalid(form)

class StaffDeleteView(LoginRequiredMixin, DeleteView):
    model = Staff
    template_name = "barkat/staff/staff_confirm_delete.html"
    success_url = reverse_lazy("staff_list")

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Staff member deleted.")
        return super().delete(request, *args, **kwargs)
    
# Bank Acount



class BankAccountListView(LoginRequiredMixin, ListView):
    model = BankAccount
    template_name = "barkat/finance/bankaccount_list.html"
    context_object_name = "accounts"

    def get_queryset(self):
        q = (self.request.GET.get("q") or "").strip()

        # Keep your existing logic for bank account balances (those are bank-specific)
        qs = (
            BankAccount.objects
            .annotate(
                net=Coalesce(
                    Sum(
                        Case(
                            When(
                                cashflows__flow_type=CashFlow.IN,
                                then=F("cashflows__amount"),
                            ),
                            When(
                                cashflows__flow_type=CashFlow.OUT,
                                then=-F("cashflows__amount"),
                            ),
                            default=Decimal("0.00"),
                            output_field=DecimalField(
                                max_digits=12,
                                decimal_places=2,
                            ),
                        )
                    ),
                    Decimal("0.00"),
                ),
                current_balance_annotated=F("opening_balance") + F("net"),
            )
            .order_by("-is_active", "bank_name", "name")
        )

        if q:
            qs = qs.filter(
                Q(name__icontains=q)
                | Q(bank_name__icontains=q)
                | Q(account_number__icontains=q)
                | Q(branch__icontains=q)
            )

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        
        # ---------------------------------------------------------------------
        # STANDARDIZED CASH IN HAND LOGIC (Ledger-based)
        # ---------------------------------------------------------------------
        D0 = Decimal("0.00")

        # 1. Cash from unified ledger (null bank_account or CASH type bank_account)
        cash_flows = CashFlow.objects.filter(
            Q(bank_account__isnull=True) | Q(bank_account__account_type=BankAccount.CASH)
        ).aggregate(
            t=Sum(Case(
                When(flow_type=CashFlow.IN, then=F('amount')),
                When(flow_type=CashFlow.OUT, then=-F('amount')),
                default=D0,
                output_field=models.DecimalField()
            ))
        )['t'] or D0
        
        # 2. Add opening balances of all active CASH type BankAccounts
        cash_acc_opening = BankAccount.objects.filter(
            account_type=BankAccount.CASH,
            is_active=True
        ).aggregate(s=Sum('opening_balance'))['s'] or D0
        
        ctx["cash_in_hand"] = cash_flows + cash_acc_opening
        return ctx
    
class BankAccountDetailView(LoginRequiredMixin, ListView):
    """
    Shows all CashFlow transactions for a single bank account.
    Supports date range filter and print mode.
    """
    model = CashFlow
    template_name = "barkat/finance/bankaccount_detail.html"
    context_object_name = "flows"
    paginate_by = 50

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)
        self.account = get_object_or_404(BankAccount, pk=self.kwargs["pk"])
        self.print_mode = self.request.GET.get("print") == "1"

        self.date_from = parse_date(self.request.GET.get("date_from") or "") or None
        self.date_to = parse_date(self.request.GET.get("date_to") or "") or None

        # this will be filled in get_queryset
        self.opening_for_range = None

    def get_paginate_by(self, queryset):
        if self.print_mode:
            return None
        return self.paginate_by

    def get_queryset(self):
        signed_amount = Case(
            When(flow_type=CashFlow.IN, then=F("amount")),
            When(flow_type=CashFlow.OUT, then=-F("amount")),
            default=Decimal("0.00"),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )

        base_opening = self.account.opening_balance or Decimal("0.00")

        # compute opening balance at start of range including all flows before date_from
        if self.date_from:
            prev_qs = (
                CashFlow.objects
                .filter(bank_account=self.account, date__lt=self.date_from)
                .annotate(signed_amount=signed_amount)
            )
            prev_total = prev_qs.aggregate(
                s=Coalesce(Sum("signed_amount"), Decimal("0.00"))
            )["s"] or Decimal("0.00")
            opening_for_range = base_opening + prev_total
        else:
            opening_for_range = base_opening

        self.opening_for_range = opening_for_range

        qs = CashFlow.objects.filter(bank_account=self.account)

        if self.date_from:
            qs = qs.filter(date__gte=self.date_from)
        if self.date_to:
            qs = qs.filter(date__lte=self.date_to)

        qs = (
            qs.select_related("bank_account")
            .select_related(
                "linked_payment",
                "linked_expense",
                "movement_in",
                "movement_out",
            )
            .annotate(
                signed_amount=signed_amount,
                running_balance=Value(
                    opening_for_range,
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
                + Window(
                    expression=Sum("signed_amount"),
                    order_by=["date", "id"],
                ),
                display_reference=Case(
                    When(linked_payment__isnull=False, then=F("linked_payment__reference")),
                    When(linked_expense__isnull=False, then=F("linked_expense__reference")),
                    When(movement_in__isnull=False, then=F("movement_in__reference_no")),
                    When(movement_out__isnull=False, then=F("movement_out__reference_no")),
                    default=Value(""),
                    output_field=CharField(),
                ),
                display_description=Case(
                    When(linked_payment__isnull=False, then=F("linked_payment__description")),
                    When(linked_expense__isnull=False, then=F("linked_expense__description")),
                    When(movement_in__isnull=False, then=F("movement_in__notes")),
                    When(movement_out__isnull=False, then=F("movement_out__notes")),
                    default=F("description"),
                    output_field=CharField(),
                ),
                source_label=Case(
                    When(movement_in__isnull=False, then=Value("Bank movement")),
                    When(movement_out__isnull=False, then=Value("Bank movement")),
                    When(linked_payment__isnull=False, then=Value("Payment")),
                    When(linked_expense__isnull=False, then=Value("Expense")),
                    default=Value("Manual entry"),
                    output_field=CharField(),
                ),
            )
            .order_by("date", "id")
        )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        qs = self.get_queryset()

        agg = qs.aggregate(
            total_in=Coalesce(
                Sum("amount", filter=Q(flow_type=CashFlow.IN)),
                Decimal("0.00"),
            ),
            total_out=Coalesce(
                Sum("amount", filter=Q(flow_type=CashFlow.OUT)),
                Decimal("0.00"),
            ),
        )

        opening = self.opening_for_range or (self.account.opening_balance or Decimal("0.00"))
        current = opening + agg["total_in"] - agg["total_out"]

        ctx["account"] = self.account
        ctx["totals"] = {
            "opening_balance": opening,  # opening at start of selected range
            "total_in": agg["total_in"],
            "total_out": agg["total_out"],
            "current_balance": current,
        }
        ctx["date_from"] = self.date_from
        ctx["date_to"] = self.date_to
        ctx["print_mode"] = self.print_mode
        return ctx

class BankAccountCreateView(LoginRequiredMixin, CreateView):
    model = BankAccount
    form_class = BankAccountForm
    template_name = "barkat/finance/bankaccount_form.html"
    success_url = reverse_lazy("bankaccount_list")

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.created_by = self.request.user
        obj.updated_by = self.request.user
        obj.save()
        messages.success(self.request, "Bank account created.")
        return redirect(self.success_url)

class BankAccountUpdateView(LoginRequiredMixin, UpdateView):
    model = BankAccount
    form_class = BankAccountForm
    template_name = "barkat/finance/bankaccount_form.html"
    success_url = reverse_lazy("bankaccount_list")

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.updated_by = self.request.user
        obj.save()
        messages.success(self.request, "Bank account updated.")
        return redirect(self.success_url)

class BankAccountDeleteView(LoginRequiredMixin, DeleteView):
    model = BankAccount
    template_name = "barkat/finance/bankaccount_confirm_delete.html"
    success_url = reverse_lazy("bankaccount_list")

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Bank account deleted.")
        return super().delete(request, *args, **kwargs)


# ---------- BankMovement CRUD ----------

def compute_cash_in_hand_for_range(date_from, date_to, business=None) -> Decimal:
    """
    Compute physical cash in hand (IN - OUT) for a given date range.
    Only counts rows where bank_account is NULL (pure cash).

    NOTE: CashFlow is currently global (no business FK). so `business`
    is accepted but not used yet. Once CashFlow is per business.
    you can add a filter here.
    """
    # Normalise ordering
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    qs = CashFlow.objects.filter(bank_account__isnull=True)

    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    agg = qs.aggregate(
        net=Coalesce(
            Sum(
                Case(
                    When(flow_type=CashFlow.IN,  then=F("amount")),
                    When(flow_type=CashFlow.OUT, then=-F("amount")),
                    default=Decimal("0.00"),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            Decimal("0.00"),
        )
    )
    return agg["net"] or Decimal("0.00")


class BankMovementListView(LoginRequiredMixin, ListView):
    model = BankMovement
    template_name = "barkat/finance/movement_list.html"
    context_object_name = "movements"

    def get_queryset(self):
        qs = (
            BankMovement.objects
            .select_related(
                "from_bank",
                "to_bank",
                "party",
                "purchase_order",
                "purchase_order__supplier",
            )
            .order_by("-date", "-id")
        )

        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(
                models.Q(method__icontains=q)
                | models.Q(reference_no__icontains=q)
                | models.Q(notes__icontains=q)
                | models.Q(from_bank__name__icontains=q)
                | models.Q(to_bank__name__icontains=q)
                | models.Q(from_bank__bank_name__icontains=q)
                | models.Q(to_bank__bank_name__icontains=q)
                | models.Q(party__display_name__icontains=q)
                | models.Q(purchase_order__id__icontains=q)
            )

        mt = (self.request.GET.get("type") or "").strip()
        if mt:
            qs = qs.filter(movement_type=mt)

        return qs


class BankMovementCreateView(LoginRequiredMixin, CreateView):
    model = BankMovement
    form_class = BankMovementForm
    template_name = "barkat/finance/movement_form.html"
    success_url = reverse_lazy("movement_list")

    def get_business_for_request(self):
        business_id = (self.request.GET.get("business") or "").strip()
        if business_id.isdigit():
            return Business.objects.filter(pk=int(business_id)).first()
        return None

    def get_initial(self):
        initial = super().get_initial()
        if not initial.get("date"):
            initial["date"] = timezone.localdate()
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        form = ctx.get("form")
        business = self.get_business_for_request()
        
        # Calculate CUMULATIVE balance until TODAY
        today = timezone.localdate()
        D0 = Decimal("0.00")

        # 1. TOTAL CASH IN (Cumulative Sales + Receipts)
        cash_in_qs = Payment.objects.filter(
            direction=Payment.IN, 
            payment_source=Payment.CASH,
            date__lte=today
        )
        if business:
            cash_in_qs = cash_in_qs.filter(business=business)
        
        total_cash_in = cash_in_qs.aggregate(
            s=Coalesce(Sum("amount", output_field=models.DecimalField(max_digits=18, decimal_places=2)), D0)
        )["s"]

        # 2. TOTAL CASH OUT (Payments to Suppliers/Others)
        cash_out_qs = Payment.objects.filter(
            direction=Payment.OUT, 
            payment_source=Payment.CASH,
            date__lte=today
        )
        if business:
            cash_out_qs = cash_out_qs.filter(business=business)
        
        total_cash_out = cash_out_qs.aggregate(
            s=Coalesce(Sum("amount", output_field=models.DecimalField(max_digits=18, decimal_places=2)), D0)
        )["s"]

        # 3. EXTRA CASH EXPENSES
        # Note: Removed the exclude(expense_id) line to fix the FieldError
        exp_cash_qs = Expense.objects.filter(date__lte=today)
        if business:
            exp_cash_qs = exp_cash_qs.filter(business=business)

        cash_source_val = getattr(Expense, "PAYMENT_SOURCE_CASH", None) or getattr(Expense, "SOURCE_CASH", None)
        if cash_source_val is not None and hasattr(Expense, "payment_source"):
            exp_cash_qs = exp_cash_qs.filter(payment_source=cash_source_val)
        elif hasattr(Expense, "payment_source"):
            exp_cash_qs = exp_cash_qs.filter(payment_source__in=["cash", "CASH"])

        total_extra_expenses = exp_cash_qs.aggregate(
            s=Coalesce(Sum("amount", output_field=models.DecimalField(max_digits=18, decimal_places=2)), D0)
        )["s"]

        # 4. BANK MOVEMENT ADJUSTMENTS
        bm_qs = BankMovement.objects.filter(date__lte=today)
        if business:
            bm_qs = bm_qs.filter(business=business)
        
        cash_delta_from_bm = D0
        for mv in bm_qs:
            amt = mv.amount or D0
            mtype = (mv.movement_type or "").lower()
            if mtype in ("deposit", "cash_deposit"):
                cash_delta_from_bm -= amt 
            elif mtype in ("withdraw", "withdrawal", "cash_withdrawal"):
                cash_delta_from_bm += amt 

        # FINAL FORMULA
        cash_in_hand = total_cash_in - total_cash_out - total_extra_expenses + cash_delta_from_bm

        ctx["cash_in_hand"] = cash_in_hand
        ctx["cash_in_hand_display"] = f"{cash_in_hand:.2f}"

        # --- Dropdown logic ---
        po_qs = PurchaseOrder.objects.select_related("supplier").order_by("-created_at", "-id")
        if business:
            po_qs = po_qs.filter(business=business)
        ctx["purchase_orders"] = po_qs

        if form and "cheques" in form.fields:
            pending_qs = form.fields["cheques"].queryset
            ctx["cheque_amounts"] = {str(p.pk): float(p.amount or 0) for p in pending_qs}
        else:
            ctx["cheque_amounts"] = {}

        return ctx

    @transaction.atomic
    def form_valid(self, form):
        obj: BankMovement = form.save(commit=False)
        obj.created_by = self.request.user
        obj.updated_by = self.request.user
        obj.save()

        mv_type = (obj.movement_type or "").lower()
        cheque_deposit_codes = {"cheque_deposit", getattr(BankMovement, "CHEQUE_DEPOSIT", "").lower()}
        cheque_payment_codes = {"cheque_payment", getattr(BankMovement, "CHEQUE_PAYMENT", "").lower()}

        if mv_type in cheque_deposit_codes:
            cheques = form.cleaned_data.get("cheques")
            if cheques:
                total = Decimal("0.00")
                for pay in cheques.select_for_update():
                    total += pay.amount or Decimal("0.00")
                if obj.amount != total:
                    obj.amount = total
                    obj.save(update_fields=["amount", "updated_by"])
                for pay in cheques.select_for_update():
                    if (getattr(pay, "direction", None) == Payment.IN
                        and getattr(pay, "payment_method", None) == Payment.PaymentMethod.CHEQUE
                        and getattr(pay, "cheque_status", None) == Payment.ChequeStatus.PENDING):
                        pay.cheque_status = Payment.ChequeStatus.DEPOSITED
                        if getattr(obj, "to_bank_id", None):
                            pay.bank_account = obj.to_bank
                            pay.payment_source = Payment.BANK
                        pay.updated_by = self.request.user
                        pay.save(update_fields=["cheque_status", "bank_account", "payment_source", "updated_by"])

        if mv_type in cheque_payment_codes:
            po = obj.purchase_order
            party = obj.party
            if po and party:
                remaining = po.balance_due or Decimal("0.00")
                if remaining < 0: remaining = 0
                amount = obj.amount or Decimal("0.00")
                if amount > 0:
                    if amount > remaining and remaining > 0:
                        amount = remaining
                        obj.amount = amount
                        obj.save(update_fields=["amount", "updated_by"])
                    
                    pay_kwargs = {
                        "business": po.business,
                        "date": obj.date,
                        "party": party,
                        "amount": amount,
                        "payment_source": Payment.BANK,
                        "created_by": self.request.user,
                        "updated_by": self.request.user,
                        "direction": Payment.OUT,
                    }
                    payment = Payment.objects.create(**pay_kwargs)
                    po.apply_payment(payment, amount)

        messages.success(self.request, "Bank movement created.")
        return redirect(self.success_url)


class BankMovementUpdateView(LoginRequiredMixin, UpdateView):
    model = BankMovement
    form_class = BankMovementForm
    template_name = "barkat/finance/movement_form.html"
    success_url = reverse_lazy("movement_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        form = ctx.get("form")

        # all purchase orders for dropdown
        po_qs = PurchaseOrder.objects.select_related("supplier").order_by("-created_at", "-id")
        ctx["purchase_orders"] = po_qs

        # cheque_amounts mapping for cheque deposit
        if form and "cheques" in form.fields:
            pending_qs = form.fields["cheques"].queryset
            ctx["cheque_amounts"] = {str(p.pk): float(p.amount or 0) for p in pending_qs}
        else:
            ctx["cheque_amounts"] = {}

        return ctx

    @transaction.atomic
    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.updated_by = self.request.user
        obj.save()
        messages.success(self.request, "Movement updated and ledger synced.")
        return redirect(self.success_url)

class BankMovementDeleteView(LoginRequiredMixin, DeleteView):
    model = BankMovement
    template_name = "barkat/finance/movement_confirm_delete.html"
    success_url = reverse_lazy("movement_list")

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Movement deleted.")
        return super().delete(request, *args, **kwargs)



# -------- Lists --------
# -------------------------
# Purchase Orders (LISTS)
# -------------------------

class PurchaseOrderListView(LoginRequiredMixin, ListView):
    model = PurchaseOrder
    template_name = "barkat/purchases/purchase_order_list.html"
    context_object_name = "orders"
    paginate_by = 20

    def _base_filtered_qs(self):
        dec = DecimalField(max_digits=12, decimal_places=2)
        zero_dec = Value(Decimal("0.00"), output_field=dec)

        qs = (
            PurchaseOrder.objects
            .select_related("business", "supplier", "created_by")
        )

        # text search (Existing logic preserved)
        q = (self.request.GET.get("q") or "").strip()
        if q:
            # We explicitly list searchable fields to avoid FieldErrors
            cond = Q()
            cond |= Q(supplier__display_name__icontains=q)
            cond |= Q(business__name__icontains=q)
            cond |= Q(status__icontains=q)
            cond |= Q(notes__icontains=q)

            if q.isdigit():
                cond |= Q(id=int(q))
            qs = qs.filter(cond)

        # date range filter (Existing logic preserved)
        d_from = (self.request.GET.get("from") or "").strip()
        d_to = (self.request.GET.get("to") or "").strip()
        if d_from:
            qs = qs.filter(created_at__date__gte=d_from)
        if d_to:
            qs = qs.filter(created_at__date__lte=d_to)

        self._date_from = d_from
        self._date_to = d_to

        # --- REAL PAID AND REMAINING CALCULATION ---
        # 1. We use your related_name 'payment_applications' to sum the bridge table amounts.
        # 2. We use 'net_total' which is the recomputed final amount in your model.
        # 3. We use 'total_cost' which is the subtotal (before tax and discount).
        qs = qs.annotate(
            annotated_subtotal=Coalesce(F("total_cost"), zero_dec),
            annotated_net_total=Coalesce(F("net_total"), zero_dec),
            paid_amount=Coalesce(
                Sum("payment_applications__amount"), 
                zero_dec
            ),
        ).annotate(
            # We calculate remaining based on the two annotations above
            remaining=F("annotated_net_total") - F("paid_amount")
        )

        return qs

    def get_queryset(self):
        # Generate the filtered and annotated queryset
        self.filtered_qs = self._base_filtered_qs()
        return self.filtered_qs.order_by("-created_at", "-id")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        dec = DecimalField(max_digits=12, decimal_places=2)
        zero_dec = Value(Decimal("0.00"), output_field=dec)

        # Aggregate totals for the summary cards at top/bottom of list
        # We use the same field names we annotated in _base_filtered_qs
        totals = self.filtered_qs.aggregate(
            total_subtotal=Coalesce(Sum("annotated_subtotal"), zero_dec),
            total_net=Coalesce(Sum("annotated_net_total"), zero_dec),
            total_paid=Coalesce(Sum("paid_amount"), zero_dec),
            total_remaining=Coalesce(Sum("remaining"), zero_dec),
        )

        ctx.update({
            "businesses": Business.objects.filter(is_active=True, is_deleted=False).order_by("name"),
            "business": None,
            "totals": totals,
            "q": (self.request.GET.get("q") or "").strip(),
            "date_from": getattr(self, "_date_from", ""),
            "date_to": getattr(self, "_date_to", ""),
        })
        return ctx

class BusinessPurchaseOrderListView(LoginRequiredMixin, ListView):
    model = PurchaseOrder
    template_name = "barkat/purchases/business_purchase_order.html"
    context_object_name = "orders"
    paginate_by = 20

    def dispatch(self, request, *args, **kwargs):
        self.business = get_object_or_404(Business, pk=kwargs.get("business_id"), is_deleted=False)
        return super().dispatch(request, *args, **kwargs)

    def _base_filtered_qs(self):
        dec = DecimalField(max_digits=12, decimal_places=2)
        zero_dec = Value(Decimal("0.00"), output_field=dec)

        # Filter by current business and select related for performance
        qs = (
            PurchaseOrder.objects
            .filter(business=self.business)
            .select_related("business", "supplier", "created_by")
        )

        # Text search (Existing logic preserved)
        q = (self.request.GET.get("q") or "").strip()
        if q:
            cond = Q()
            cond |= Q(supplier__display_name__icontains=q)
            cond |= Q(supplier_name__icontains=q)
            cond |= Q(supplier_phone__icontains=q)
            cond |= Q(status__icontains=q)
            cond |= Q(notes__icontains=q)

            if q.isdigit():
                cond |= Q(id=int(q))
            qs = qs.filter(cond)

        # Date range filter (Existing logic preserved)
        d_from = (self.request.GET.get("from") or "").strip()
        d_to = (self.request.GET.get("to") or "").strip()
        if d_from:
            qs = qs.filter(created_at__date__gte=d_from)
        if d_to:
            qs = qs.filter(created_at__date__lte=d_to)

        self._date_from = d_from
        self._date_to = d_to

        # --- REAL PAID AND REMAINING CALCULATION ---
        # We use 'payment_applications' which is the related_name for PurchaseOrderPayment bridge
        qs = qs.annotate(
            # Subtotal (before tax and discount)
            annotated_subtotal=Coalesce(F("total_cost"), zero_dec),
            # Final cost of the order
            annotated_net_total=Coalesce(F("net_total"), zero_dec),
            # Total sum of all payments applied to this order
            paid_amount=Coalesce(
                Sum("payment_applications__amount"), 
                zero_dec
            ),
        ).annotate(
            # Balance due calculation
            remaining=F("annotated_net_total") - F("paid_amount")
        )

        return qs

    def get_queryset(self):
        self.filtered_qs = self._base_filtered_qs()
        return self.filtered_qs.order_by("-created_at", "-id")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        dec = DecimalField(max_digits=12, decimal_places=2)
        zero_dec = Value(Decimal("0.00"), output_field=dec)

        # Aggregate totals for the top summary cards
        totals = self.filtered_qs.aggregate(
            total_subtotal=Coalesce(Sum("annotated_subtotal"), zero_dec),
            total_net=Coalesce(Sum("annotated_net_total"), zero_dec),
            total_paid=Coalesce(Sum("paid_amount"), zero_dec),
            total_remaining=Coalesce(Sum("remaining"), zero_dec),
        )

        ctx.update({
            "businesses": Business.objects.filter(is_active=True, is_deleted=False).order_by("name"),
            "business": self.business,
            "totals": totals,
            "q": (self.request.GET.get("q") or "").strip(),
            "date_from": getattr(self, "_date_from", ""),
            "date_to": getattr(self, "_date_to", ""),
        })
        return ctx
# ----------------------------------------
# Create / Update with inline items + PAY
# ----------------------------------------
from decimal import Decimal


# views.py
# In views.py
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from datetime import datetime
from django.utils import timezone


@require_GET
def product_detail_api(request):
    """Return product UOM details for PO form"""
    product_id = request.GET.get('product_id')
    if not product_id:
        return JsonResponse({'ok': False}, status=400)
    
    try:
        product = Product.objects.select_related('uom', 'bulk_uom').get(pk=product_id)
        data = {
            'ok': True,
            'product_id': product.id,
            'name': product.name,
            'uom_id': product.uom_id,
            'uom_code': product.uom.code if product.uom else '',
            'bulk_uom_id': product.bulk_uom_id,
            'bulk_uom_code': product.bulk_uom.code if product.bulk_uom else '',
            'default_bulk_size': str(product.default_bulk_size),
            'purchase_price': str(product.purchase_price),
        }
        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({'ok': False}, status=404)


def check_barcode_exists_api(request):
    """
    API endpoint to check if a barcode exists in other businesses.
    Returns JSON with existence info.
    """
    from django.http import JsonResponse
    
    if request.method != 'GET':
        return JsonResponse({"ok": False, "error": "Only GET method allowed"}, status=405)
    
    barcode = request.GET.get("barcode", "").strip()
    business_id = request.GET.get("business_id", "").strip()
    
    if not barcode:
        return JsonResponse({"ok": False, "error": "Barcode parameter required"}, status=400)
    
    # Find product with this barcode
    existing_product = Product.objects.filter(
        barcode=barcode
    ).exclude(
        barcode=""
    ).select_related("business").first()
    
    if not existing_product:
        return JsonResponse({
            "ok": True,
            "exists": False,
            "message": "Barcode is available"
        })
    
    # Barcode exists - check if it's in a different business
    if business_id and str(existing_product.business_id) == str(business_id):
        # Same business - that's OK (might be editing existing product)
        return JsonResponse({
            "ok": True,
            "exists": True,
            "same_business": True,
            "business_id": str(existing_product.business_id),
            "business_name": existing_product.business.name,
            "product_name": existing_product.name,
            "message": f"Barcode exists in your business for product '{existing_product.name}'"
        })
    else:
        # Different business - this is a duplicate!
        return JsonResponse({
            "ok": True,
            "exists": True,
            "same_business": False,
            "business_id": str(existing_product.business_id),
            "business_name": existing_product.business.name,
            "product_name": existing_product.name,
            "message": f"Barcode already exists in business '{existing_product.business.name}' for product '{existing_product.name}'"
        })


def generate_barcode_api(request):
    """Generate a unique barcode for a product"""
    business_id = request.GET.get('business_id')
    product_id = request.GET.get('product_id')  # Optional, for existing products
    
    if not business_id:
        return JsonResponse({'ok': False, 'error': 'business_id is required'}, status=400)
    
    try:
        business = Business.objects.get(pk=business_id, is_deleted=False)
    except Business.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Business not found'}, status=404)
    
    # Convert product_id to int if provided
    pid = None
    if product_id and str(product_id).isdigit():
        try:
            pid = int(product_id)
            # Verify product exists and belongs to the business
            Product.objects.get(pk=pid, business=business)
        except (ValueError, Product.DoesNotExist):
            pid = None
    
    # Generate barcode
    barcode = Product.generate_barcode(business=business, product_id=pid)
    
    return JsonResponse({
        'ok': True,
        'barcode': barcode
    })


class PurchaseOrderCreateView(LoginRequiredMixin, CreateView):
    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = "barkat/purchases/purchase_order_form.html"

    # -----------------------------
    # helper to get selected business
    # -----------------------------
    def _get_selected_business(self, request, form=None):
        bid = request.POST.get("business") or request.GET.get("business")
        if bid:
            return Business.objects.filter(pk=bid).first()
        if form is not None:
            bval = form.initial.get("business")
            if isinstance(bval, Business):
                return bval
            if bval:
                return Business.objects.filter(pk=bval).first()
        return None

    def get_initial(self):
        initial = super().get_initial()
        bid = self.request.GET.get("business")
        if bid and Business.objects.filter(pk=bid).exists():
            initial["business"] = bid
        return initial

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        biz = self._get_selected_business(self.request)
        if biz:
            kwargs["fixed_business"] = biz
        return kwargs

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['supplier'].queryset = Party.objects.filter(
            is_active=True, 
            is_deleted=False
        ).order_by('display_name')
        return form  # <--- THIS WAS MISSING
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        business = self._get_selected_business(self.request, form=ctx["form"])
        ctx["formset"] = PurchaseOrderItemFormSet(
            self.request.POST or None,
            form_kwargs={"business": business},
        )
        ctx["expense_formset"] = PurchaseOrderExpenseFormSet(
            self.request.POST or None,
            prefix="expenses",
        )
        
        # Products for quick add with UOM data - filtered by business if selected
        products_qs = Product.objects.filter(
            is_active=True,
            is_deleted=False
        ).select_related("uom", "bulk_uom", "category").order_by("name")
        
        # Filter by business if one is selected
        if business:
            products_qs = products_qs.filter(business=business)
        
        products_cards = []
        for p in products_qs:
            product_data = {
                "id": p.id,
                "name": p.name,
                "company_name": getattr(p, "company_name", "") or "",
                "sale_price": str(p.sale_price),
                "purchase_price": str(p.purchase_price or 0),
                "category_id": p.category_id or "",
                "stock": str(p.stock_qty or 0),
                "barcode": getattr(p, "barcode", "") or "",
                "uom_id": p.uom_id or "",
                "uom_code": p.uom.code if p.uom else "",
                "has_bulk": bool(p.bulk_uom_id and p.default_bulk_size > 0),
                "business_id": str(p.business_id),  # Add business_id for JavaScript filtering
            }
            
            # Add bulk unit info if available
            if p.bulk_uom_id and p.default_bulk_size and p.default_bulk_size > 0:
                product_data["bulk_uom_id"] = p.bulk_uom_id
                product_data["bulk_uom_code"] = p.bulk_uom.code
                product_data["bulk_size"] = str(p.default_bulk_size)
            else:
                product_data["bulk_uom_id"] = ""
                product_data["bulk_uom_code"] = ""
                product_data["bulk_size"] = "1"
                
            products_cards.append(product_data)
        
        ctx["products_cards"] = products_cards
        ctx["business"] = business
        
        # Add UOMs and Categories for product registration modal
        from barkat.models import UnitOfMeasure, ProductCategory
        ctx["uoms"] = UnitOfMeasure.objects.all().order_by("code")
        if business:
            ctx["categories"] = ProductCategory.objects.filter(business=business, is_deleted=False).order_by("name")
        else:
            ctx["categories"] = ProductCategory.objects.filter(is_deleted=False).order_by("business__name", "name")
        
        return ctx


    @transaction.atomic
    def form_valid(self, form):
        ctx = self.get_context_data(form=form)
        formset = ctx["formset"]
        expense_formset = ctx["expense_formset"]

        if not formset.is_valid() or not expense_formset.is_valid():
            return self.form_invalid(form)

        po: PurchaseOrder = form.save(commit=False)
        biz = self._get_selected_business(self.request, form=form)
        if biz:
            po.business = biz

        po_date = form.cleaned_data.get("po_date")
        if po_date:
            po_datetime = timezone.make_aware(
                datetime.combine(po_date, datetime.min.time())
            )
            po.created_at = po_datetime

        po.created_by = self.request.user
        po.updated_by = self.request.user
        po.save()


        # Save formset items with proper uom/size handling and sale price conversion
        for item_form in formset:
            if item_form.cleaned_data and not item_form.cleaned_data.get('DELETE'):
                item = item_form.save(commit=False)
                item.purchase_order = po
                
                # ✅ NEW: Only set defaults if TRULY missing (shouldn't happen with proper form)
                if item.uom_id is None:
                    # This should not happen if form is working correctly
                    item.uom = item.product.uom
                    item.size_per_unit = Decimal("1.000000")
                    print(f"⚠️ WARNING: UOM was not set for item {item.product.name}, defaulting to lowest unit")
                
                # Handle sale price conversion: if sale_price is provided and unit is bulk, convert to lower unit
                sale_price = item_form.cleaned_data.get('sale_price')
                if sale_price is not None and sale_price > 0:
                    # Check if the selected UOM is the bulk unit
                    if (item.product.bulk_uom_id and 
                        item.uom_id == item.product.bulk_uom_id and 
                        item.size_per_unit and 
                        item.size_per_unit > Decimal("1")):
                        # Sale price is in bulk unit - convert to lower unit
                        # Example: 5000 (bag) / 50 (kg per bag) = 100 (per kg)
                        lower_unit_sale_price = sale_price / item.size_per_unit
                        # Update product's sale_price with the converted lower unit price
                        Product.objects.filter(pk=item.product_id).update(
                            sale_price=lower_unit_sale_price
                        )
                    else:
                        # Sale price is already in lower unit - save directly
                        Product.objects.filter(pk=item.product_id).update(
                            sale_price=sale_price
                        )
                    
                item.save()

        # Save expenses
        expenses = expense_formset.save(commit=False)
        for expense in expenses:
            expense.purchase_order = po
            if biz:
                expense.business = biz
            expense.created_by = self.request.user
            expense.updated_by = self.request.user
            expense.save()

            # --- INSTANT PAYMENT LOGIC ---
            if expense.is_paid and not expense.payment:
                pay_method = expense.payment_source  # 'cash' or 'bank'
                pay_source = Payment.CASH if pay_method == "cash" else Payment.BANK
                
                payment_kwargs = {
                    "business": biz,
                    "date": po_date or timezone.localdate(),
                    "party": po.supplier,
                    "amount": expense.amount,
                    "description": f"Instant payment for PO #{po.id} expense: {expense.get_category_display()}",
                    "reference": f"PO-{po.id}-EXP",
                    "payment_source": pay_source,
                    "payment_method": "bank" if pay_method == "bank" else "cash",
                    "direction": Payment.OUT,
                    "created_by": self.request.user,
                    "updated_by": self.request.user,
                }
                if pay_method == "bank":
                    payment_kwargs["bank_account"] = expense.bank_account
                
                pay = Payment.objects.create(**payment_kwargs)

                
                # Link payment to expense and save again (to trigger model logic that skips CashFlow)
                expense.payment = pay
                expense.save(update_fields=["payment", "updated_at", "updated_by"])
                
                # ALSO Link payment to PO so it shows up in "Paid So Far"
                po.apply_payment(pay, expense.amount)
            
        # Handle deleted expenses
        for deleted_expense in expense_formset.deleted_objects:
            # If deleted expense had a linked payment, maybe delete the payment too?
            # For now, let's keep it simple.
            deleted_expense.delete()

        # Recompute and Distribute
        if hasattr(po, "distribute_expenses"):
            po.distribute_expenses()
            
        if hasattr(po, "recompute_totals"):
            po.recompute_totals()
            po.save(
                update_fields=[
                    "total_cost",
                    "net_total",
                    "updated_at",
                    "updated_by",
                ]
            )

        # Stock update logic (only when status == "received")
        status = (po.status or "").lower()
        if status == "received":
            item_qs = po.items.all()
            # (filtered logic remains same...)
            item_qs = item_qs.filter(product__isnull=False, quantity__gt=0)
            
            for item in item_qs:
                actual_qty = (item.quantity or Decimal("0")) * (item.size_per_unit or Decimal("1"))
                if actual_qty > 0:
                    Product.objects.filter(pk=item.product_id).update(
                        stock_qty=F("stock_qty") + actual_qty
                    )
                    # Update purchase price using LANDING COST
                    price_to_store = item.landing_unit_price or item.unit_price
                    if price_to_store is not None:
                        Product.objects.filter(pk=item.product_id).update(
                            purchase_price=price_to_store
                        )

        # Payment logic
        method = form.cleaned_data.get("payment_method") or "none"
        bank = form.cleaned_data.get("bank_account")
        paid = (form.cleaned_data.get("paid_amount") or Decimal("0.00")).quantize(
            Decimal("0.01")
        )

        if paid > 0:
            pay_source = None
            if method == "cash":
                pay_source = Payment.CASH
            elif method in ("bank", "cheque"):
                pay_source = Payment.BANK

            if pay_source:
                # NEW: Use po_date for payment date as well
                payment_date = po_date or timezone.localdate()
                
                payment_kwargs = {
                    "business": po.business,
                    "date": payment_date,  # Use PO date
                    "party": po.supplier,
                    "amount": paid,
                    "description": f"Payment for PO #{po.id}",
                    "reference": f"PO-{po.id}",
                    "payment_source": pay_source,
                    "created_by": self.request.user,
                    "updated_by": self.request.user,
                }

                if _model_has_field(Payment, "direction"):
                    payment_kwargs["direction"] = Payment.OUT

                if _model_has_field(Payment, "payment_method"):
                    payment_kwargs["payment_method"] = method

                if (
                    method in ("bank", "cheque")
                    and _model_has_field(Payment, "bank_account")
                    and bank
                ):
                    payment_kwargs["bank_account"] = bank

                payment = Payment.objects.create(**payment_kwargs)


                po.apply_payment(payment, paid)
                messages.success(
                    self.request,
                    f"Recorded payment ₨ {paid} for PO #{po.id}.",
                )

        messages.success(self.request, f"Purchase Order #{po.id} created.")
        self.object = po
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("po_list")


class PurchaseOrderUpdateView(LoginRequiredMixin, UpdateView):
    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = "barkat/purchases/purchase_order_form.html"

    def get_initial(self):
        initial = super().get_initial()
        po: PurchaseOrder = self.get_object()
        remaining = po.balance_due
        if remaining < Decimal("0.00"):
            remaining = Decimal("0.00")
        initial["paid_amount"] = remaining
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        po: PurchaseOrder = self.object

        applications = (
            PurchaseOrderPayment.objects
            .filter(purchase_order=po)
            .select_related("payment", "payment__bank_account")
            .order_by("created_at", "id")
        )

        paid_so_far = applications.aggregate(s=Sum("amount")).get("s") or Decimal("0.00")
        remaining = (po.net_total or Decimal("0.00")) - paid_so_far
        if remaining < Decimal("0.00"):
            remaining = Decimal("0.00")

        ctx["previous_payments"] = applications
        ctx["paid_so_far"] = paid_so_far
        ctx["remaining"] = remaining

        if self.request.POST:
            ctx["formset"] = PurchaseOrderItemFormSet(
                self.request.POST,
                instance=po,
                form_kwargs={"business": po.business},
            )
        else:
            ctx["formset"] = PurchaseOrderItemFormSet(
                instance=po,
                form_kwargs={"business": po.business},
            )
        
        ctx["expense_formset"] = PurchaseOrderExpenseFormSet(
            self.request.POST or None,
            instance=po,
            prefix="expenses",
        )
        
        # NEW: Build item unit data for JavaScript to restore saved units
        item_unit_data = {}
        for item in po.items.select_related('product', 'product__uom', 'product__bulk_uom', 'uom').all():
            if item.product_id:
                # Determine which unit is currently selected
                current_unit = 'lowest'
                if item.uom_id and item.product.bulk_uom_id and item.uom_id == item.product.bulk_uom_id:
                    current_unit = 'bulk'
                
                item_unit_data[str(item.id)] = {
                    'product_id': item.product_id,
                    'uom_id': item.uom_id,
                    'size_per_unit': str(item.size_per_unit or '1.000000'),
                    'current_unit': current_unit,
                }
        
        ctx['item_unit_data'] = item_unit_data
        
        # Products for quick add with UOM data - filtered by business
        products_qs = Product.objects.filter(
            is_active=True,
            is_deleted=False,
            business=po.business  # Filter by purchase order's business
        ).select_related("uom", "bulk_uom", "category").order_by("name")
        
        products_cards = []
        for p in products_qs:
            product_data = {
                "id": p.id,
                "name": p.name,
                "company_name": getattr(p, "company_name", "") or "",
                "sale_price": str(p.sale_price),
                "purchase_price": str(p.purchase_price or 0),
                "category_id": p.category_id or "",
                "stock": str(p.stock_qty or 0),
                "barcode": getattr(p, "barcode", "") or "",
                "uom_id": p.uom_id or "",
                "uom_code": p.uom.code if p.uom else "",
                "has_bulk": bool(p.bulk_uom_id and p.default_bulk_size > 0),
                "business_id": str(p.business_id),  # Add business_id for JavaScript filtering
            }
            
            # Add bulk unit info if available
            if p.bulk_uom_id and p.default_bulk_size and p.default_bulk_size > 0:
                product_data["bulk_uom_id"] = p.bulk_uom_id
                product_data["bulk_uom_code"] = p.bulk_uom.code
                product_data["bulk_size"] = str(p.default_bulk_size)
            else:
                product_data["bulk_uom_id"] = ""
                product_data["bulk_uom_code"] = ""
                product_data["bulk_size"] = "1"
                
            products_cards.append(product_data)
        
        ctx["products_cards"] = products_cards
        ctx["business"] = po.business
        
        # Add UOMs and Categories for product registration modal
        from barkat.models import UnitOfMeasure, ProductCategory
        ctx["uoms"] = UnitOfMeasure.objects.all().order_by("code")
        ctx["categories"] = ProductCategory.objects.filter(business=po.business, is_deleted=False).order_by("name")
        
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        ctx = self.get_context_data(form=form)
        formset = ctx["formset"]
        expense_formset = ctx["expense_formset"]
        
        if not formset.is_valid() or not expense_formset.is_valid():
            return self.form_invalid(form)

        # Snapshot old PO and items
        po_db: PurchaseOrder = (
            PurchaseOrder.objects.select_for_update()
            .prefetch_related("items")
            .get(pk=self.object.pk)
        )
        old_status = (po_db.status or "").lower()

        po: PurchaseOrder = form.save(commit=False)
        po.updated_by = self.request.user
        po.save()

        # OLD: Track old quantities in BASE UNIT (considering size_per_unit)
        old_qty_by_product = defaultdict(Decimal)
        for it in po_db.items.all():
            if not it.product_id:
                continue
            q = it.quantity or Decimal("0")
            size = it.size_per_unit or Decimal("1")
            if q <= 0:
                continue
            # Convert to base unit
            base_qty = q * size
            old_qty_by_product[it.product_id] += base_qty

        # Save new PO
        po: PurchaseOrder = form.save(commit=False)
        po.updated_by = self.request.user
        po.save()

        # Save items with proper uom/size handling and sale price conversion
        formset.instance = po
        instances = formset.save(commit=False)
        deleted = list(formset.deleted_objects)

        for idx, inst in enumerate(instances):
            inst.purchase_order = po
            
            # ✅ NEW: Only set defaults if TRULY missing (shouldn't happen with proper form)
            if inst.uom_id is None:
                # This should not happen if form is working correctly
                inst.uom = inst.product.uom
                inst.size_per_unit = Decimal("1.000000")
                print(f"⚠️ WARNING: UOM was not set for item {inst.product.name}, defaulting to lowest unit")
            
            # Handle sale price conversion: get sale_price from instance (set by formset.save(commit=False))
            sale_price = getattr(inst, 'sale_price', None)
            
            if sale_price is not None and sale_price > 0:
                # Check if the selected UOM is the bulk unit
                if (inst.product.bulk_uom_id and 
                    inst.uom_id == inst.product.bulk_uom_id and 
                    inst.size_per_unit and 
                    inst.size_per_unit > Decimal("1")):
                    # Sale price is in bulk unit - convert to lower unit
                    # Example: 5000 (bag) / 50 (kg per bag) = 100 (per kg)
                    lower_unit_sale_price = sale_price / inst.size_per_unit
                    # Update product's sale_price with the converted lower unit price
                    Product.objects.filter(pk=inst.product_id).update(
                        sale_price=lower_unit_sale_price
                    )
                else:
                    # Sale price is already in lower unit - save directly
                    Product.objects.filter(pk=inst.product_id).update(
                        sale_price=sale_price
                    )
                
            inst.save()

        for inst in deleted:
            inst.delete()

        formset.save_m2m()

        # New status and NEW items map (in BASE UNIT)
        new_status = (po.status or "").lower()

        new_qty_by_product = defaultdict(Decimal)
        for it in po.items.select_related("product").all():
            if not it.product_id:
                continue
            q = it.quantity or Decimal("0")
            size = it.size_per_unit or Decimal("1")
            if q <= 0:
                continue
            # Convert to base unit
            base_qty = q * size
            new_qty_by_product[it.product_id] += base_qty

        # Compute deltas - only status "received" contributes to stock
        all_product_ids = set(old_qty_by_product.keys()) | set(new_qty_by_product.keys())

        for pid in all_product_ids:
            old_q = old_qty_by_product.get(pid, Decimal("0"))
            new_q = new_qty_by_product.get(pid, Decimal("0"))

            # Old effect: stock added previously (if status was received)
            old_effect = old_q if old_status == "received" else Decimal("0")
            # New effect: stock to add now (if status is received)
            new_effect = new_q if new_status == "received" else Decimal("0")

            # Delta: difference between new and old effect
            delta = new_effect - old_effect
            
            if delta:
                Product.objects.filter(pk=pid).update(
                    stock_qty=F("stock_qty") + delta
                )

        # Save expenses
        expenses = expense_formset.save(commit=False)
        for expense in expenses:
            expense.purchase_order = po
            expense.business = po.business
            expense.created_by = self.request.user
            expense.updated_by = self.request.user
            expense.save()

            # --- INSTANT PAYMENT LOGIC ---
            if expense.is_paid and not expense.payment:
                pay_method = expense.payment_source  # 'cash' or 'bank'
                pay_source = Payment.CASH if pay_method == "cash" else Payment.BANK
                
                payment_kwargs = {
                    "business": po.business,
                    "date": po.created_at.date() if po.created_at else timezone.localdate(),
                    "party": po.supplier,
                    "amount": expense.amount,
                    "description": f"Instant payment for PO #{po.id} expense: {expense.get_category_display()}",
                    "reference": f"PO-{po.id}-EXP",
                    "payment_source": pay_source,
                    "payment_method": "bank" if pay_method == "bank" else "cash",
                    "direction": Payment.OUT,
                    "created_by": self.request.user,
                    "updated_by": self.request.user,
                }
                if pay_method == "bank":
                    payment_kwargs["bank_account"] = expense.bank_account
                
                pay = Payment.objects.create(**payment_kwargs)

                
                # Link payment to expense and save again (to trigger model logic that skips CashFlow)
                expense.payment = pay
                expense.save(update_fields=["payment", "updated_at", "updated_by"])

                # ALSO Link payment to PO so it shows up in "Paid So Far"
                po.apply_payment(pay, expense.amount)
            
        for deleted_expense in expense_formset.deleted_objects:
            deleted_expense.delete()

        # Recompute and Distribute
        if hasattr(po, "distribute_expenses"):
            po.distribute_expenses()
            
        if hasattr(po, "recompute_totals"):
            po.recompute_totals()
            po.save(update_fields=["total_cost", "net_total", "updated_at", "updated_by"])

        # Update last purchase price using LANDING COST
        for pid in new_qty_by_product.keys():
            last_item = (
                po.items.filter(product_id=pid)
                .exclude(unit_price__isnull=True)
                .order_by("id")
                .last()
            )
            if last_item:
                price_to_store = last_item.landing_unit_price or last_item.unit_price
                if price_to_store is not None:
                    Product.objects.filter(pk=pid).update(
                        purchase_price=price_to_store
                    )

        # Recompute totals
        if hasattr(po, "recompute_totals"):
            po.recompute_totals()
            po.save(update_fields=["total_cost", "net_total", "updated_at", "updated_by"])

        # Optional payment - clamped to remaining
        remaining = po.balance_due.quantize(Decimal("0.01"))
        method = form.cleaned_data.get("payment_method") or "none"
        bank = form.cleaned_data.get("bank_account")
        paid = (form.cleaned_data.get("paid_amount") or Decimal("0.00")).quantize(
            Decimal("0.01")
        )
        if paid > remaining:
            paid = remaining

        if paid > 0:
            pay_source = None
            if method == "cash":
                pay_source = Payment.CASH
            elif method in ("bank", "cheque"):
                pay_source = Payment.BANK

            if pay_source:
                payment_kwargs = {
                    "business": po.business,
                    "date": timezone.localdate(),
                    "party": po.supplier,
                    "amount": paid,
                    "description": f"Payment for PO #{po.id}",
                    "reference": f"PO-{po.id}",
                    "payment_source": pay_source,
                    "created_by": self.request.user,
                    "updated_by": self.request.user,
                }

                if _model_has_field(Payment, "direction"):
                    payment_kwargs["direction"] = Payment.OUT

                if _model_has_field(Payment, "payment_method"):
                    payment_kwargs["payment_method"] = method

                if (
                    method in ("bank", "cheque")
                    and _model_has_field(Payment, "bank_account")
                    and bank
                ):
                    payment_kwargs["bank_account"] = bank

                payment = Payment.objects.create(**payment_kwargs)


                po.apply_payment(payment, paid)
                messages.success(
                    self.request,
                    f"Recorded payment ₨ {paid} for PO #{po.id}.",
                )

        messages.success(self.request, f"Purchase Order #{po.id} updated.")
        self.object = po
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("po_list")



class PurchaseOrderDeleteView(LoginRequiredMixin, DeleteView):
    model = PurchaseOrder
    template_name = "barkat/purchases/purchase_order_confirm_delete.html"
    success_url = reverse_lazy("po_list")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Purchase Order deleted.")
        return super().delete(request, *args, **kwargs)

class PurchaseReturnListView(LoginRequiredMixin, ListView):
    model = PurchaseReturn
    template_name = "barkat/purchases/purchase_return_list.html"
    context_object_name = "returns"

    def get_queryset(self):
        qs = (
            PurchaseReturn.objects
            .select_related("business", "supplier", "created_by")
            .order_by("-created_at", "-id")
        )

        q = (self.request.GET.get("q") or "").strip()
        if q:
            cond = (
                Q(supplier__display_name__icontains=q) |
                Q(business__name__icontains=q) |
                Q(status__icontains=q)
            )
            if q.isdigit():
                cond |= Q(id=int(q))
            qs = qs.filter(cond)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["businesses"] = Business.objects.order_by("name")
        ctx["business"] = None  # "All" tab
        return ctx

class BusinessPurchaseReturnListView(LoginRequiredMixin, ListView):
    model = PurchaseReturn
    template_name = "barkat/purchases/business_purchase_return.html"
    context_object_name = "returns"

    def dispatch(self, request, *args, **kwargs):
        self.business = get_object_or_404(Business, pk=kwargs.get("business_id"))
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        qs = (
            PurchaseReturn.objects
            .filter(business=self.business)
            .select_related("business", "supplier", "created_by")
            .order_by("-created_at", "-id")
        )

        q = (self.request.GET.get("q") or "").strip()
        if q:
            cond = Q(supplier__display_name__icontains=q) | Q(status__icontains=q)
            if q.isdigit():
                cond |= Q(id=int(q))
            qs = qs.filter(cond)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["businesses"] = Business.objects.order_by("name")
        ctx["business"] = self.business
        return ctx


# ----------------------------------------
# Create / Update with inline items + REFUND
# ----------------------------------------


class PurchaseReturnCreateView(LoginRequiredMixin, CreateView):
    model = PurchaseReturn
    form_class = PurchaseReturnForm
    template_name = "barkat/purchases/purchase_return_form.html"

    # helper: find selected business early to pass into formset
    def _get_selected_business(self, request, form=None):
        bid = request.POST.get("business") or request.GET.get("business")
        if bid:
            return Business.objects.filter(pk=bid).first()
        if form is not None:
            bval = form.initial.get("business")
            if isinstance(bval, Business):
                return bval
            if bval:
                return Business.objects.filter(pk=bval).first()
        return None

    def get_initial(self):
        initial = super().get_initial()
        bid = self.request.GET.get("business")
        if bid and Business.objects.filter(pk=bid).exists():
            initial["business"] = bid
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        business = self._get_selected_business(self.request, form=ctx["form"])
        ctx["formset"] = PurchaseReturnItemFormSet(
            self.request.POST or None,
            form_kwargs={"business": business},
        )
        
        # Products for quick add with UOM data - filtered by business if selected
        products_qs = Product.objects.filter(
            is_active=True,
            is_deleted=False
        ).select_related("uom", "bulk_uom", "category").order_by("name")
        
        # Filter by business if one is selected
        if business:
            products_qs = products_qs.filter(business=business)
        
        products_cards = []
        for p in products_qs:
            product_data = {
                "id": p.id,
                "name": p.name,
                "company_name": getattr(p, "company_name", "") or "",
                "sale_price": str(p.sale_price),
                "purchase_price": str(p.purchase_price or 0),
                "category_id": p.category_id or "",
                "stock": str(p.stock_qty or 0),
                "barcode": getattr(p, "barcode", "") or "",
                "uom_id": p.uom_id or "",
                "uom_code": p.uom.code if p.uom else "",
                "has_bulk": bool(p.bulk_uom_id and p.default_bulk_size > 0),
                "business_id": str(p.business_id),  # Add business_id for JavaScript filtering
            }
            
            # Add bulk unit info if available
            if p.bulk_uom_id and p.default_bulk_size and p.default_bulk_size > 0:
                product_data["bulk_uom_id"] = p.bulk_uom_id
                product_data["bulk_uom_code"] = p.bulk_uom.code
                product_data["bulk_size"] = str(p.default_bulk_size)
            else:
                product_data["bulk_uom_id"] = ""
                product_data["bulk_uom_code"] = ""
                product_data["bulk_size"] = "1"
                
            products_cards.append(product_data)
        
        ctx["products_cards"] = products_cards
        ctx["business"] = business
        
        # Add UOMs and Categories for product registration modal
        from barkat.models import UnitOfMeasure, ProductCategory
        ctx["uoms"] = UnitOfMeasure.objects.all().order_by("code")
        if business:
            ctx["categories"] = ProductCategory.objects.filter(business=business, is_deleted=False).order_by("name")
        else:
            ctx["categories"] = ProductCategory.objects.filter(is_deleted=False).order_by("business__name", "name")
        
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        ctx = self.get_context_data(form=form)
        formset = ctx["formset"]
        if not formset.is_valid():
            return self.form_invalid(form)

        pr: PurchaseReturn = form.save(commit=False)
        pr.created_by = self.request.user
        pr.updated_by = self.request.user
        pr.save()

        # Save items first so we can read them for stock movement
        formset.instance = pr
        formset.save()

        # STOCK DELTA (CREATE)
        # Purchase Return reduces on-hand stock by returned qty.
        # Convert to product's base unit using size_per_unit
        for it in pr.items.select_related("product").all():
            prod = it.product
            if not prod:
                continue
            # Calculate actual quantity in base unit
            actual_qty = (it.quantity or Decimal("0")) * (it.size_per_unit or Decimal("1"))
            if actual_qty > 0:
                Product.objects.filter(pk=prod.pk).update(
                    stock_qty=F("stock_qty") - actual_qty
                )

        # Totals
        if hasattr(pr, "recompute_totals"):
            pr.recompute_totals()
            pr.save(update_fields=["total_cost", "net_total", "updated_at", "updated_by"])

        # Optional initial refund (money IN from supplier)
        method = form.cleaned_data.get("refund_method") or "none"
        bank   = form.cleaned_data.get("bank_account")
        received = (form.cleaned_data.get("received_amount") or Decimal("0.00")).quantize(Decimal("0.01"))

        if received > 0:
            pay_source = Payment.CASH if method == "cash" else (Payment.BANK if method == "bank" else None)
            if pay_source:
                payment = Payment(
                    business=pr.business,
                    date=timezone.localdate(),
                    party=pr.supplier,
                    direction=Payment.IN,  # refund coming IN
                    amount=received,
                    description=f"Refund for PR #{pr.id}",
                    reference=f"PR-{pr.id}",
                    payment_source=pay_source,
                    bank_account=bank if pay_source == Payment.BANK else None,
                    created_by=self.request.user,
                    updated_by=self.request.user,
                )
                try:
                    payment.full_clean()
                    payment.save()
                except ValidationError as e:
                    form.add_error(None, e)
                    return self.form_invalid(form)

                pr.apply_refund(payment, received)
                messages.success(self.request, f"Recorded refund ₨ {received} for PR #{pr.id}.")

        messages.success(self.request, f"Purchase Return #{pr.id} created.")
        self.object = pr
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("pr_list")

class PurchaseReturnUpdateView(LoginRequiredMixin, UpdateView):
    model = PurchaseReturn
    form_class = PurchaseReturnForm
    template_name = "barkat/purchases/purchase_return_form.html"

    def get_initial(self):
        initial = super().get_initial()
        pr: PurchaseReturn = self.get_object()
        remaining = pr.refund_remaining
        if remaining < Decimal("0.00"):
            remaining = Decimal("0.00")
        initial["received_amount"] = remaining
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        pr: PurchaseReturn = self.object

        applications = (
            PurchaseReturnRefund.objects
            .filter(purchase_return=pr)
            .select_related("payment", "payment__bank_account")
            .order_by("created_at", "id")
        )

        refunded_so_far = applications.aggregate(s=Sum("amount")).get("s") or Decimal("0.00")
        remaining = (pr.net_total or Decimal("0.00")) - refunded_so_far
        if remaining < Decimal("0.00"):
            remaining = Decimal("0.00")

        ctx["previous_refunds"] = applications
        ctx["refunded_so_far"] = refunded_so_far
        ctx["remaining"] = remaining

        if self.request.POST:
            ctx["formset"] = PurchaseReturnItemFormSet(
                self.request.POST,
                instance=pr,
                form_kwargs={"business": pr.business},
            )
        else:
            ctx["formset"] = PurchaseReturnItemFormSet(
                instance=pr,
                form_kwargs={"business": pr.business},
            )
        
        # Item unit data for edit mode (to restore unit selections)
        item_unit_data = {}
        for item in pr.items.select_related('product', 'product__uom', 'product__bulk_uom', 'uom').all():
            if item.uom_id:
                # Determine if it's bulk or lowest unit
                current_unit = 'lowest'
                if item.product.bulk_uom_id and item.uom_id == item.product.bulk_uom_id:
                    current_unit = 'bulk'
                
                item_unit_data[str(item.id)] = {
                    'current_unit': current_unit,
                    'uom_id': str(item.uom_id),
                    'size_per_unit': str(item.size_per_unit or '1.000000'),
                }
        ctx['item_unit_data'] = item_unit_data
        
        # Products for quick add with UOM data - filtered by business
        products_qs = Product.objects.filter(
            is_active=True,
            is_deleted=False,
            business=pr.business
        ).select_related("uom", "bulk_uom", "category").order_by("name")
        
        products_cards = []
        for p in products_qs:
            product_data = {
                "id": p.id,
                "name": p.name,
                "company_name": getattr(p, "company_name", "") or "",
                "sale_price": str(p.sale_price),
                "purchase_price": str(p.purchase_price or 0),
                "category_id": p.category_id or "",
                "stock": str(p.stock_qty or 0),
                "barcode": getattr(p, "barcode", "") or "",
                "uom_id": p.uom_id or "",
                "uom_code": p.uom.code if p.uom else "",
                "has_bulk": bool(p.bulk_uom_id and p.default_bulk_size > 0),
                "business_id": str(p.business_id),
            }
            
            # Add bulk unit info if available
            if p.bulk_uom_id and p.default_bulk_size and p.default_bulk_size > 0:
                product_data["bulk_uom_id"] = p.bulk_uom_id
                product_data["bulk_uom_code"] = p.bulk_uom.code
                product_data["bulk_size"] = str(p.default_bulk_size)
            else:
                product_data["bulk_uom_id"] = ""
                product_data["bulk_uom_code"] = ""
                product_data["bulk_size"] = "1"
                
            products_cards.append(product_data)
        
        ctx["products_cards"] = products_cards
        ctx["business"] = pr.business
        
        # Add UOMs and Categories for product registration modal
        from barkat.models import UnitOfMeasure, ProductCategory
        ctx["uoms"] = UnitOfMeasure.objects.all().order_by("code")
        ctx["categories"] = ProductCategory.objects.filter(business=pr.business, is_deleted=False).order_by("name")
        
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        ctx = self.get_context_data(form=form)
        formset = ctx["formset"]
        if not formset.is_valid():
            return self.form_invalid(form)

        pr: PurchaseReturn = form.save(commit=False)
        pr.updated_by = self.request.user
        pr.save()

        # Snapshot old items (in case you later add custom stock logic)
        old_items = {
            it.pk: {"product_id": it.product_id, "qty": (it.quantity or Decimal("0"))}
            for it in pr.items.all()
        }

        # Save new and changed, gather deleted
        formset.instance = pr
        instances = formset.save(commit=False)
        deleted = list(formset.deleted_objects)

        for inst in instances:
            inst.purchase_return = pr
            inst.save()

        for inst in deleted:
            inst.delete()

        formset.save_m2m()

        # Totals
        if hasattr(pr, "recompute_totals"):
            pr.recompute_totals()
            pr.save(update_fields=["total_cost", "net_total", "updated_at", "updated_by"])

        # Optional refund (clamped to remaining)
        remaining = pr.refund_remaining.quantize(Decimal("0.01"))
        method = form.cleaned_data.get("refund_method") or "none"
        bank   = form.cleaned_data.get("bank_account")
        received = (form.cleaned_data.get("received_amount") or Decimal("0.00")).quantize(Decimal("0.01"))
        if received > remaining:
            received = remaining

        if received > 0:
            pay_source = Payment.CASH if method == "cash" else (Payment.BANK if method == "bank" else None)
            if pay_source:
                payment = Payment(
                    business=pr.business,
                    date=timezone.localdate(),
                    party=pr.supplier,
                    direction=Payment.IN,
                    amount=received,
                    description=f"Refund for PR #{pr.id}",
                    reference=f"PR-{pr.id}",
                    payment_source=pay_source,
                    bank_account=bank if pay_source == Payment.BANK else None,
                    created_by=self.request.user,
                    updated_by=self.request.user,
                )
                try:
                    payment.full_clean()
                    payment.save()
                except ValidationError as e:
                    form.add_error(None, e)
                    return self.form_invalid(form)

                pr.apply_refund(payment, received)
                messages.success(self.request, f"Recorded refund ₨ {received} for PR #{pr.id}.")

        messages.success(self.request, f"Purchase Return #{pr.id} updated.")
        self.object = pr
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("pr_list")

class PurchaseReturnDeleteView(LoginRequiredMixin, DeleteView):
    model = PurchaseReturn
    template_name = "barkat/purchases/purchase_return_confirm_delete.html"
    success_url = reverse_lazy("pr_list")

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Purchase Return deleted.")
        return super().delete(request, *args, **kwargs)

class ExpenseBusinessMixin:
    """
    Reuse helper to get currently selected business from GET/POST
    """
    def _get_selected_business(self, request, form=None):
        bid = request.POST.get("business") or request.GET.get("business")
        if bid:
            return Business.objects.filter(pk=bid).first()
        if form is not None:
            bval = form.initial.get("business")
            if isinstance(bval, Business):
                return bval
            if bval:
                return Business.objects.filter(pk=bval).first()
        return None



class ExpensesListView(LoginRequiredMixin, ListView):
    template_name = "barkat/finance/expense_list.html"
    context_object_name = "expenses"
    login_url = "login"

    def get_queryset(self):
        queryset = Expense.objects.select_related("business", "party", "staff", "bank_account").order_by("-date", "-id")
        
        # 1. Filter by Category
        category = self.request.GET.get("category")
        if category:
            queryset = queryset.filter(category=category)

        # 2. Date Filtering (Quick Buttons and Custom Range)
        date_filter = self.request.GET.get("date_filter")
        today = timezone.localdate()

        if date_filter == "today":
            queryset = queryset.filter(date=today)
        elif date_filter == "yesterday":
            queryset = queryset.filter(date=today - timedelta(days=1))
        elif date_filter == "this_week":
            start_week = today - timedelta(days=today.weekday())
            queryset = queryset.filter(date__gte=start_week)
        elif date_filter == "this_month":
            queryset = queryset.filter(date__year=today.year, date__month=today.month)
        elif date_filter == "this_year":
            queryset = queryset.filter(date__year=today.year)
        
        # 3. Custom Date Range
        start_date = self.request.GET.get("start_date")
        end_date = self.request.GET.get("end_date")
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])

        return queryset
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        
        # Calculate Total for the current filtered queryset
        ctx["total_expense_amount"] = qs.aggregate(total=Sum("amount"))["total"] or 0
        
        # Pass constants for the filter UI
        ctx["businesses"] = Business.objects.filter(is_deleted=False, is_active=True).order_by("name")
        
        # ExpenseCategory is likely an Enum or TextChoices from your models
        from .models import ExpenseCategory 
        ctx["expense_categories"] = ExpenseCategory.choices
        
        # FIX: Use .dict() instead of .to_dict()
        ctx["current_filters"] = self.request.GET.dict() 
        
        return ctx
    
class BusinessExpensesListView(LoginRequiredMixin, ListView):
    template_name = "barkat/finance/business_expense.html"
    context_object_name = "expenses"
    login_url = "login"

    def dispatch(self, request, *args, **kwargs):
        self.business = get_object_or_404(Business, pk=kwargs["business_id"], is_deleted=False)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        # Start with business filter
        queryset = Expense.objects.filter(business=self.business).select_related(
            "business", "party", "staff", "bank_account"
        ).order_by("-date", "-id")
        
        # Reuse the same filtering logic as above
        category = self.request.GET.get("category")
        if category:
            queryset = queryset.filter(category=category)

        date_filter = self.request.GET.get("date_filter")
        today = timezone.localdate()

        if date_filter == "today":
            queryset = queryset.filter(date=today)
        elif date_filter == "yesterday":
            queryset = queryset.filter(date=today - timedelta(days=1))
        elif date_filter == "this_week":
            start_week = today - timedelta(days=today.weekday())
            queryset = queryset.filter(date__gte=start_week)
        elif date_filter == "this_month":
            queryset = queryset.filter(date__year=today.year, date__month=today.month)
        elif date_filter == "this_year":
            queryset = queryset.filter(date__year=today.year)
        
        start_date = self.request.GET.get("start_date")
        end_date = self.request.GET.get("end_date")
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])

        return queryset

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        
        ctx["business"] = self.business
        ctx["total_expense_amount"] = qs.aggregate(total=Sum("amount"))["total"] or 0
        ctx["businesses"] = Business.objects.filter(is_deleted=False, is_active=True).order_by("name")
        ctx["expense_categories"] = ExpenseCategory.choices
        return ctx
    
def _map_payment_source(expense: Expense) -> str | None:
    if expense.payment_source == Expense.CASH:
        return Payment.CASH
    if expense.payment_source == Expense.BANK:
        return Payment.BANK
    return None

@transaction.atomic
def upsert_payment_for_expense(expense: Expense) -> Payment | None:
    """
    Ensure a Payment.OUT exists for this Expense. Then apply it to the linked PO if possible.
    Reference used: EXP-<expense.id>
    """
    if not expense.pk or not expense.business_id:
        return None

    # Prefer explicit party. Otherwise use PO supplier if provided.
    party = expense.party or (expense.purchase_order.supplier if expense.purchase_order_id else None)
    if not party:
        return None

    pay_source = _map_payment_source(expense)
    if pay_source is None:
        return None

    ref = f"EXP-{expense.id}"
    desc = f"Expense {expense.get_category_display()}"

    payment, created = Payment.objects.get_or_create(
        business=expense.business,
        reference=ref,
        defaults=dict(
            date=expense.date,
            party=party,
            direction=Payment.OUT,
            amount=expense.amount,
            description=desc,
            payment_source=pay_source,
            bank_account=expense.bank_account if pay_source == Payment.BANK else None,
            created_by=expense.created_by,
            updated_by=expense.updated_by,
        ),
    )

    if not created:
        payment.date = expense.date
        payment.party = party
        payment.direction = Payment.OUT
        payment.amount = expense.amount
        payment.description = desc
        payment.payment_source = pay_source
        payment.bank_account = expense.bank_account if pay_source == Payment.BANK else None
        payment.updated_by = expense.updated_by
        payment.full_clean()
        payment.save(update_fields=[
            "date", "party", "direction", "amount", "description",
            "payment_source", "bank_account", "updated_at", "updated_by"
        ])

    # Apply to PO if linked and vendor matches
    if expense.purchase_order_id:
        po = expense.purchase_order
        if po.business_id == expense.business_id and po.supplier_id == party.id:
            remaining = po.balance_due
            if remaining < Decimal("0.00"):
                remaining = Decimal("0.00")
            apply_amt = min(payment.amount, remaining)

            bridge, _ = PurchaseOrderPayment.objects.get_or_create(
                purchase_order=po,
                payment=payment,
                defaults=dict(
                    amount=Decimal("0.00"),
                    created_by=expense.created_by,
                    updated_by=expense.updated_by,
                ),
            )
            bridge.amount = apply_amt
            bridge.updated_by = expense.updated_by
            bridge.full_clean()
            bridge.save(update_fields=["amount", "updated_at", "updated_by"])

    return payment

# -----------------------
# Expense create and edit
class ExpenseCreateView(LoginRequiredMixin, ExpenseBusinessMixin, CreateView):
    model = Expense
    form_class = ExpenseForm
    template_name = "barkat/finance/expense_form.html"
    success_url = reverse_lazy("finance_expense_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        selected = self._get_selected_business(self.request)
        if selected:
            kwargs.setdefault("initial", {})["business"] = selected.pk
        kwargs["selected_business"] = selected
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.updated_by = self.request.user

        # Save expense first so it has an id
        response = super().form_valid(form)

        # Mirror to Payment and apply to PO if present
        try:
            payment = upsert_payment_for_expense(self.object)
            if payment and self.object.purchase_order_id:
                po = self.object.purchase_order
                messages.success(
                    self.request,
                    f"Expense created. Applied ₨ {payment.amount} to PO #{po.id}. Remaining ₨ {po.balance_due}."
                )
            else:
                messages.success(self.request, "Expense created.")
        except ValidationError as e:
            messages.warning(
                self.request,
                f"Expense created, but ledger sync had a validation issue: {e}"
            )
        except Exception:
            messages.warning(
                self.request,
                "Expense created. Ledger sync could not be completed."
            )

        return response

    def get_success_url(self):
        nxt = self.request.GET.get("next") or self.request.POST.get("next")
        return nxt or reverse("finance_expense_list")

class ExpenseUpdateView(LoginRequiredMixin, ExpenseBusinessMixin, UpdateView):
    model = Expense
    form_class = ExpenseForm
    template_name = "barkat/finance/expense_form.html"
    success_url = reverse_lazy("finance_expense_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        obj = self.object  # set by UpdateView before get_form_kwargs
        if obj and obj.business_id:
            selected = obj.business
        else:
            selected = self._get_selected_business(self.request)
        kwargs["selected_business"] = selected
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        form.instance.updated_by = self.request.user

        response = super().form_valid(form)

        # Keep mirrored Payment and PO bridge in sync with edits
        try:
            payment = upsert_payment_for_expense(self.object)
            if payment and self.object.purchase_order_id:
                po = self.object.purchase_order
                messages.success(
                    self.request,
                    f"Expense updated. Applied ₨ {payment.amount} to PO #{po.id}. Remaining ₨ {po.balance_due}."
                )
            else:
                messages.success(self.request, "Expense updated.")
        except ValidationError as e:
            messages.warning(
                self.request,
                f"Expense updated, but ledger sync had a validation issue: {e}"
            )
        except Exception:
            messages.warning(
                self.request,
                "Expense updated. Ledger sync could not be completed."
            )

        return response

    def get_success_url(self):
        nxt = self.request.GET.get("next") or self.request.POST.get("next")
        return nxt or reverse("finance_expense_list")

class ExpenseDeleteView(LoginRequiredMixin, DeleteView):
    model = Expense
    template_name = "barkat/finance/expense_confirm_delete.html"
    success_url = reverse_lazy("finance_expense_list")

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Expense deleted.")
        return super().delete(request, *args, **kwargs)

class ExpenseDetailView(LoginRequiredMixin, DetailView):
    model = Expense
    template_name = "barkat/finance/expense_detail.html"
    context_object_name = "expense"


#------------
#   Sales
#------------

class _SOBaseMixin(LoginRequiredMixin):
    template_name = "barkat/sales/order_form.html"
    login_url = "login"

    def _get_selected_business(self):
        bid = self.request.POST.get("business") or self.request.GET.get("business")
        if bid:
            return Business.objects.filter(pk=bid, is_deleted=False).first()
        # fallback to first active business
        return Business.objects.filter(is_deleted=False, is_active=True).order_by("name").first()

    def _context_lists(self, biz):
        products = Product.objects.filter(
            business=biz, is_deleted=False, is_active=True
        ).select_related("category", "uom").order_by("name")

        categories = ProductCategory.objects.filter(
            business=biz, is_deleted=False
        ).order_by("name")

        customers = Party.objects.filter(
            type__in=["CUSTOMER", "BOTH"], is_deleted=False
        ).order_by("display_name")

        return products, categories, customers

    def _inject_common_ctx(self, ctx, biz):
        products, categories, customers = self._context_lists(biz)
        ctx.update({
            "business": biz,
            "businesses": Business.objects.filter(is_deleted=False, is_active=True).order_by("name"),
            "products": products,
            "categories": categories,
            "customers": customers,
        })
        return ctx

# ---------- Helpers ----------

def _model_has_field(model, field_name: str) -> bool:
    try:
        model._meta.get_field(field_name)
        return True
    except Exception:
        return False

def _selected_business(request: HttpRequest):
    """Pick business by ?business=ID or default to the first one."""
    bid = request.GET.get("business")
    if bid and str(bid).isdigit():
        return get_object_or_404(Business, pk=int(bid))
    return Business.objects.order_by("name", "id").first()

def ensure_party_for_receipt(business, customer, customer_name, customer_phone):
    """
    Return a Party to attach to Payment:
    - If a specific customer is selected, use it.
    - Otherwise use/create a per-business 'Walk-in Customer'.
    """
    if customer:
        return customer

    phone = (customer_phone or "").strip()
    party, _ = Party.objects.get_or_create(
        display_name="Walk-in-Customer",
        default_business=business,
        defaults={
            "type": "CUSTOMER",
            "is_active": True,
            "phone": phone,
        },
    )
    if not party.phone and phone:
        party.phone = phone
        party.save(update_fields=["phone"])
    return party

# ------------------------------------------------------
# Image helper: do NOT use ProductImage in views
# ------------------------------------------------------


def _product_image_url(p):
    """
    Return a safe image URL for a Product.
    Priority: Product.primary_image().image -> p.image (if direct field exists) -> placeholder.
    """
    try:
        # Prefer related ProductImage marked as primary
        if hasattr(p, "primary_image"):
            pim = p.primary_image()
            if pim and getattr(pim, "image", None):
                url = getattr(pim.image, "url", "")
                if url:
                    return url
        # Fallback if product has a direct `image` field (some deployments do)
        direct = getattr(p, "image", None)
        if direct:
            url = getattr(direct, "url", "")
            if url:
                return url
    except Exception:
        pass
    # Final fallback: static placeholder
    return static("img/product-placeholder.png")

# ------------------------------------------------------
# CREATE
# ------------------------------------------------------


# make sure this helper exists in the same file or is imported correctly
# from .utils import _product_image_url

def _get_walkin_party(business):
    """
    Try to find EXISTING 'Walk-in-Customer' Party.
    If business is given, prefer default_business = business.
    Otherwise return any Walk-in party.
    NEVER creates a new one - it must already exist.
    """
    qs = Party.objects.filter(
        is_active=True,
        is_deleted=False,
        display_name__iexact="Walk-in-Customer",
    )

    biz_id = None
    if business is not None:
        try:
            biz_id = getattr(business, "id", None) or getattr(business, "pk", None)
        except Exception:
            biz_id = None
        if biz_id is None and str(business).isdigit():
            biz_id = int(business)

    if biz_id:
        # First try to find one with matching default_business
        p = qs.filter(default_business_id=biz_id).first()
        if p:
            return p

    # Return any Walk-in-Customer (case insensitive)
    return qs.first()


def _q2(v) -> Decimal:
    try:
        return Decimal(str(v or "0")).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _model_has_field(model, field_name: str) -> bool:
    """Return True if `model` has a real DB field named `field_name`."""
    return any(
        getattr(f, "name", None) == field_name
        and getattr(f, "concrete", False)
        and not getattr(f, "many_to_many", False)
        and not getattr(f, "auto_created", False)
        for f in model._meta.get_fields()
    )


def _product_image_url(product):
    """Helper to get product image URL"""
    img = product.primary_image()
    if img and img.image:
        return img.image.url
    return ""

# barkat/views.py (Sales Order section)

class SalesOrderCreateView(LoginRequiredMixin, CreateView):
    model = SalesOrder
    form_class = SalesOrderForm
    template_name = "barkat/sales/order_form.html"

    def get(self, request, *args, **kwargs):
        """Redirect to include business parameter in URL if default business is set"""
        bid = request.GET.get("business")
        
        # If no business parameter in URL, check for default business
        if not bid:
            user_settings = getattr(request.user, "settings", None)
            if user_settings and user_settings.default_sale_business_id:
                # Redirect to same URL with business parameter
                from django.http import HttpResponseRedirect
                return HttpResponseRedirect(f"{request.path}?business={user_settings.default_sale_business_id}")
        
        return super().get(request, *args, **kwargs)

    def get_initial(self):
        init = super().get_initial()
        bid = self.request.GET.get("business")
        if bid and bid.isdigit():
            init["business"] = bid
        else:
            # Check user settings for default business
            user_settings = getattr(self.request.user, "settings", None)
            if user_settings and user_settings.default_sale_business_id:
                init["business"] = user_settings.default_sale_business_id
        init.setdefault("status", SalesOrder.Status.OPEN)
        return init

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        business = None
        if getattr(self, "object", None) and self.object and self.object.business_id:
            business = self.object.business
        else:
            bid = self.request.POST.get("business") or self.request.GET.get("business")
            if bid and str(bid).isdigit():
                business = Business.objects.filter(pk=int(bid)).first()
        kwargs["business"] = business
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        business = None
        form = ctx.get("form")
        if form and getattr(form, "business", None):
            business = form.business
        if not business:
            bid = self.request.GET.get("business")
            if bid and str(bid).isdigit():
                business = Business.objects.filter(pk=int(bid)).first()

        ctx["businesses"] = Business.objects.order_by("name", "id")
        ctx["business"] = business
    
        if self.request.POST:
            ctx["formset"] = SalesOrderItemFormSet(
                self.request.POST,
                form_kwargs={"business": business},
            )
        else:
            ctx["formset"] = SalesOrderItemFormSet(
                form_kwargs={"business": business},
            )

        # Products for UI with UOM data
        products_qs = Product.objects.filter(
            is_active=True, 
            is_deleted=False
        ).select_related("uom", "bulk_uom", "category").order_by("name")
        
        products_cards = []
        for p in products_qs:
            product_data = {
                "id": p.id,
                "name": p.name,
                "sale_price": str(p.sale_price),
                "category_id": p.category_id or "",
                "stock": str(p.stock_qty or 0),
                "barcode": getattr(p, "barcode", "") or "",
                "uom_id": p.uom_id or "",
                "uom_code": p.uom.code if p.uom else "",
                "has_bulk": bool(p.bulk_uom_id and p.default_bulk_size > 0),
            }
            
            # Add bulk unit info if available
            if p.bulk_uom_id and p.default_bulk_size and p.default_bulk_size > 0:
                product_data["bulk_uom_id"] = p.bulk_uom_id
                product_data["bulk_uom_code"] = p.bulk_uom.code
                product_data["bulk_size"] = str(p.default_bulk_size)
            else:
                product_data["bulk_uom_id"] = ""
                product_data["bulk_uom_code"] = ""
                product_data["bulk_size"] = "1"
                
            products_cards.append(product_data)
            
        ctx["products_cards"] = products_cards
        ctx["paid_so_far"] = 0
        ctx["remaining"] = 0
        ctx["previous_receipts"] = []
        ctx["item_unit_data"] = {}  # For new orders, empty
        us = getattr(self.request.user, "settings", None)
        ctx["default_sale_payment_method"] = (
            getattr(us, "default_sale_payment_method", None) or "cash"
        )

        return ctx

    @transaction.atomic
    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx["formset"]
        if not formset.is_valid():
            return self.form_invalid(form)

        # Validate at least one product
        valid_items_count = 0
        for f in formset.forms:
            cd = getattr(f, "cleaned_data", None)
            if not cd or cd.get("DELETE"):
                continue
            prod = cd.get("product")
            qty = cd.get("quantity") or Decimal("0")
            if prod and qty > 0:
                valid_items_count += 1

        if valid_items_count == 0:
            form.add_error(None, "Sales Order must have at least one product item.")
            return self.form_invalid(form)

        # Stock check (convert to base unit)
        requested = {}
        row_map = {}
        for f in formset.forms:
            cd = getattr(f, "cleaned_data", None)
            if not cd or cd.get("DELETE"):
                continue
            prod = cd.get("product")
            qty = cd.get("quantity") or Decimal("0")
            size = cd.get("size_per_unit") or Decimal("1")
            
            if not prod or qty <= 0:
                continue
                
            # Convert to base unit for stock check
            base_qty = qty * size
            requested[prod.id] = requested.get(prod.id, Decimal("0")) + base_qty
            row_map.setdefault(prod.id, []).append(f)

        if requested:
            prods = (
                Product.objects
                .select_for_update()
                .filter(id__in=requested.keys(), is_deleted=False)
            )
            stock_map = {p.id: (p.stock_qty or Decimal("0")) for p in prods}

            any_error = False
            for pid, need in requested.items():
                have = stock_map.get(pid, Decimal("0"))
                if need > have:
                    any_error = True
                    for f in row_map.get(pid, []):
                        prod_name = f.cleaned_data.get("product").name if f.cleaned_data.get("product") else "Product"
                        f.add_error("quantity", f"{prod_name}: Only {have} in stock. You requested {need}.")
            if any_error:
                form.add_error(None, "Insufficient stock for one or more items.")
                return self.form_invalid(form)

        # Save order
        self.object = form.save(commit=False)

        # Set created_at from order_date
        # The form sends naive datetime in Pakistan time, we need to make it timezone-aware
        order_date = form.cleaned_data.get("order_date")
        if order_date:
            if isinstance(order_date, datetime):
                # If it's already timezone-aware, use it as-is
                if timezone.is_aware(order_date):
                    self.object.created_at = order_date
                else:
                    # It's naive datetime in Pakistan time, make it aware
                    # timezone.make_aware() will treat it as being in TIME_ZONE (Asia/Karachi)
                    self.object.created_at = timezone.make_aware(order_date)
            elif isinstance(order_date, date):
                # If it's just a date, convert to datetime at start of day
                order_datetime = datetime.combine(order_date, datetime.min.time())
                self.object.created_at = timezone.make_aware(order_datetime)
            else:
                # Fallback: use current time
                self.object.created_at = timezone.now()
        else:
            # No order_date provided, use current time
            self.object.created_at = timezone.now()

        customer = form.cleaned_data.get("customer")
        cname = (form.cleaned_data.get("customer_name") or "").strip()
        business = form.cleaned_data.get("business")

        # Walk-in customer logic
        if not customer and not cname:
            walkin = _get_walkin_party(business)
            if walkin:
                self.object.customer = walkin
                self.object.customer_name = walkin.display_name
                self.object.customer_phone = walkin.phone or ""
                self.object.customer_address = walkin.address or ""
            else:
                self.object.customer = None
                self.object.customer_name = "Walk-in Customer"
                self.object.customer_phone = ""
                self.object.customer_address = ""
        else:
            self.object.customer = customer
            if cname:
                self.object.customer_name = cname

        self.object.created_by = self.request.user
        self.object.updated_by = self.request.user
        self.object.save()

        # Save items with UOM support
        for item_form in formset:
            if item_form.cleaned_data and not item_form.cleaned_data.get('DELETE'):
                item = item_form.save(commit=False)
                item.sales_order = self.object
                
                # Ensure uom and size_per_unit are set
                if not item.uom_id:
                    item.uom = item.product.uom
                if not item.size_per_unit:
                    item.size_per_unit = Decimal("1.000000")
                    
                item.save()

        self.object.recompute_totals()
        
        # Set status to OPEN initially (will be updated to FULFILLED after payment if fully paid)
        self.object.status = SalesOrder.Status.OPEN
        
        self.object.updated_by = self.request.user
        self.object.save()

        # Stock out (using base units)
        if requested:
            prods = Product.objects.select_for_update().filter(id__in=requested.keys())
            for p in prods:
                need = requested.get(p.id, Decimal("0"))
                if need > 0:
                    p.stock_qty = (p.stock_qty or Decimal("0")) - need
                    try:
                        p.updated_by = self.request.user
                        p.save(update_fields=["stock_qty", "updated_by", "updated_at"])
                    except Exception:
                        p.save(update_fields=["stock_qty"])

        # Receipt on create
        method = form.cleaned_data.get("receipt_method")
        amount = form.cleaned_data.get("received_amount")
        bank = form.cleaned_data.get("bank_account")

        # Extract date from order_date for payment (handle both date and datetime)
        if order_date:
            if isinstance(order_date, datetime):
                pay_date = timezone.localdate(order_date) if timezone.is_aware(order_date) else order_date.date()
            elif isinstance(order_date, date):
                pay_date = order_date
            else:
                pay_date = timezone.localdate()
        else:
            pay_date = timezone.localdate()

        if method in ("cash", "bank", "card") and amount and amount > 0:
            order = self.object
            party = order.customer or _get_walkin_party(order.business)

            if party:
                # Card and Bank both go to bank ledger
                payment_source_value = "bank" if method in ("bank", "card") else "cash"

                payment_kwargs = {
                    "business": order.business,
                    "party": party,
                    "date": pay_date,
                    "amount": amount,
                    "payment_source": payment_source_value,
                    "created_by": self.request.user,
                    "updated_by": self.request.user,
                }

                if _model_has_field(Payment, "direction"):
                    payment_kwargs["direction"] = Payment.IN

                if _model_has_field(Payment, "payment_method"):
                    payment_kwargs["payment_method"] = method

                # Bank and Card both require bank_account
                if method in ("bank", "card") and _model_has_field(Payment, "bank_account") and bank:
                    payment_kwargs["bank_account"] = bank

                pay = Payment.objects.create(**payment_kwargs)

                available = _q2(getattr(order, "balance_due", 0))
                applied_amount = _q2(amount)
                
                if available <= 0:
                    applied_amount = Decimal("0.00")
                elif applied_amount > available:
                    applied_amount = available

                if applied_amount > 0:
                    try:
                        order.apply_receipt(pay, applied_amount)
                        order.recompute_totals()
                        # Auto-update status to fulfilled if fully paid
                        if order.paid_total >= order.net_total and order.net_total > Decimal("0.00"):
                            order.status = SalesOrder.Status.FULFILLED
                        order.updated_by = self.request.user
                        order.save()
                    except ValidationError as ve:
                        messages.error(self.request, str(ve))

                # CashFlow is now automatically handled by Payment.save()


        messages.success(self.request, f"Sales Order #{self.object.pk} created.")
        return redirect(self.get_success_url())

    def get_success_url(self):
        url = reverse("so_add")
        biz_id = self.object.business_id if self.object else None
        return f"{url}?business={biz_id}" if biz_id else url

class SalesOrderUpdateView(LoginRequiredMixin, UpdateView):
    model = SalesOrder
    form_class = SalesOrderForm
    template_name = "barkat/sales/order_form.html"
    pk_url_kwarg = "pk"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["business"] = self.object.business
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        so = self.object
        business = so.business

        ctx["businesses"] = Business.objects.order_by("name", "id")
        ctx["business"] = business

        if self.request.POST:
            ctx["formset"] = SalesOrderItemFormSet(
                self.request.POST,
                instance=so,
                form_kwargs={"business": business},
            )
        else:
            ctx["formset"] = SalesOrderItemFormSet(
                instance=so,
                form_kwargs={"business": business},
            )

        # Products for UI with UOM data
        products_qs = Product.objects.filter(
            is_active=True, 
            is_deleted=False
        ).select_related("uom", "bulk_uom", "category").order_by("name")
        
        products_cards = []
        for p in products_qs:
            product_data = {
                "id": p.id,
                "name": p.name,
                "sale_price": str(p.sale_price),
                "category_id": p.category_id or "",
                "stock": str(p.stock_qty or 0),
                "barcode": getattr(p, "barcode", "") or "",
                "uom_id": p.uom_id or "",
                "uom_code": p.uom.code if p.uom else "",
                "has_bulk": bool(p.bulk_uom_id and p.default_bulk_size > 0),
            }
            
            if p.bulk_uom_id and p.default_bulk_size and p.default_bulk_size > 0:
                product_data["bulk_uom_id"] = p.bulk_uom_id
                product_data["bulk_uom_code"] = p.bulk_uom.code
                product_data["bulk_size"] = str(p.default_bulk_size)
            else:
                product_data["bulk_uom_id"] = ""
                product_data["bulk_uom_code"] = ""
                product_data["bulk_size"] = "1"
                
            products_cards.append(product_data)
            
        ctx["products_cards"] = products_cards
        ctx["paid_so_far"] = so.paid_total
        ctx["previous_receipts"] = so.receipt_applications.select_related(
            "payment", "payment__bank_account"
        ).order_by("created_at")
        
        # Build item unit data for JavaScript
        item_unit_data = {}
        for item in so.items.select_related('product', 'product__uom', 'product__bulk_uom', 'uom').all():
            if item.product_id:
                current_unit = 'lowest'
                if item.uom_id and item.product.bulk_uom_id and item.uom_id == item.product.bulk_uom_id:
                    current_unit = 'bulk'
                
                item_unit_data[str(item.id)] = {
                    'product_id': item.product_id,
                    'uom_id': item.uom_id,
                    'size_per_unit': str(item.size_per_unit or '1.000000'),
                    'current_unit': current_unit,
                }
        
        ctx['item_unit_data'] = item_unit_data
        us = getattr(self.request.user, "settings", None)
        ctx["default_sale_payment_method"] = (
            getattr(us, "default_sale_payment_method", None) or "cash"
        )

        return ctx

    @transaction.atomic
    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx["formset"]
        
        if not formset.is_valid():
            return self.form_invalid(form)

        so = self.object
        
        # Stock reversal with UOM support
        db_items = {
            it.id: (it.product_id, it.quantity, it.size_per_unit or Decimal("1")) 
            for it in so.items.all()
        }
        stock_changes = {}

        for f in formset.forms:
            if not f.cleaned_data: 
                continue
            
            instance_id = f.instance.id
            new_product = f.cleaned_data.get('product')
            new_qty = f.cleaned_data.get('quantity') or Decimal('0')
            new_size = f.cleaned_data.get('size_per_unit') or Decimal('1')
            is_deleted = f.cleaned_data.get('DELETE')

            if instance_id in db_items:
                old_pid, old_qty, old_size = db_items[instance_id]
                old_base = old_qty * old_size
                new_base = new_qty * new_size
                
                if is_deleted:
                    stock_changes[old_pid] = stock_changes.get(old_pid, Decimal('0')) + old_base
                elif new_product and new_product.id != old_pid:
                    stock_changes[old_pid] = stock_changes.get(old_pid, Decimal('0')) + old_base
                    stock_changes[new_product.id] = stock_changes.get(new_product.id, Decimal('0')) - new_base
                else:
                    diff = old_base - new_base
                    stock_changes[old_pid] = stock_changes.get(old_pid, Decimal('0')) + diff
            else:
                if not is_deleted and new_product:
                    new_base = new_qty * new_size
                    stock_changes[new_product.id] = stock_changes.get(new_product.id, Decimal('0')) - new_base

        # Apply stock changes
        for pid, qty_to_change in stock_changes.items():
            if qty_to_change != 0:
                Product.objects.filter(id=pid).update(stock_qty=F('stock_qty') + qty_to_change)

        # Save order (created_at remains unchanged - it's immutable after creation)
        so = form.save(commit=False)
        so.updated_by = self.request.user
        so.save()

        # Save items with UOM
        for item_form in formset:
            if item_form.cleaned_data and not item_form.cleaned_data.get('DELETE'):
                item = item_form.save(commit=False)
                if not item.uom_id:
                    item.uom = item.product.uom
                if not item.size_per_unit:
                    item.size_per_unit = Decimal("1.000000")
                item.save()

        formset.save()
        
        # Recompute totals FIRST before handling payments
        so.recompute_totals()
        so.save()
        
        # Auto-update status based on payment status (before processing new payment)
        if so.paid_total >= so.net_total and so.net_total > Decimal("0.00"):
            so.status = SalesOrder.Status.FULFILLED
        else:
            # If not fully paid, ensure status is OPEN (unless already cancelled)
            if so.status != SalesOrder.Status.CANCELLED:
                so.status = SalesOrder.Status.OPEN
        so.save(update_fields=['status'])
        
        # Handle payment similar to CreateView
        method = form.cleaned_data.get("receipt_method") or "none"
        amount = form.cleaned_data.get("received_amount") or Decimal("0.00")
        bank = form.cleaned_data.get("bank_account")
        order_date = form.cleaned_data.get("order_date")
        
        # Extract date from order_date for payment (handle both date and datetime)
        if order_date:
            if isinstance(order_date, datetime):
                pay_date = timezone.localdate(order_date) if timezone.is_aware(order_date) else order_date.date()
            elif isinstance(order_date, date):
                pay_date = order_date
            else:
                pay_date = timezone.localdate()
        else:
            # Fallback: use order's created_at date if available, otherwise today
            if so.created_at:
                pay_date = timezone.localdate(so.created_at) if timezone.is_aware(so.created_at) else so.created_at.date()
            else:
                pay_date = timezone.localdate()

        if method in ("cash", "bank", "card") and amount and amount > 0:
            # Clean up old payment applications if updating
            # This prevents duplicate payments when editing the order
            for app in so.receipt_applications.all():
                pay_to_del = app.payment
                # Delete associated cash flow if exists
                if hasattr(pay_to_del, 'cashflow') and pay_to_del.cashflow:
                    pay_to_del.cashflow.delete()
                # Delete the application
                app.delete()
                # Delete the payment
                pay_to_del.delete()

            party = so.customer or _get_walkin_party(so.business)

            if party:
                # Card and Bank both go to bank ledger
                payment_source_value = "bank" if method in ("bank", "card") else "cash"

                payment_kwargs = {
                    "business": so.business,
                    "party": party,
                    "date": pay_date,
                    "amount": amount,
                    "payment_source": payment_source_value,
                    "created_by": self.request.user,
                    "updated_by": self.request.user,
                }

                if _model_has_field(Payment, "direction"):
                    payment_kwargs["direction"] = Payment.IN

                if _model_has_field(Payment, "payment_method"):
                    payment_kwargs["payment_method"] = method

                # Bank and Card both require bank_account
                if method in ("bank", "card") and _model_has_field(Payment, "bank_account") and bank:
                    payment_kwargs["bank_account"] = bank

                pay = Payment.objects.create(**payment_kwargs)

                # Use the UPDATED balance_due after recompute_totals
                available = _q2(so.balance_due)
                applied_amount = _q2(amount)
                
                if available <= 0:
                    applied_amount = Decimal("0.00")
                elif applied_amount > available:
                    applied_amount = available

                if applied_amount > 0:
                    try:
                        so.apply_receipt(pay, applied_amount)
                        # Recompute totals again after applying receipt
                        so.recompute_totals()
                        # Auto-update status to fulfilled if fully paid after payment
                        if so.paid_total >= so.net_total and so.net_total > Decimal("0.00"):
                            so.status = SalesOrder.Status.FULFILLED
                        so.updated_by = self.request.user
                        so.save()
                    except ValidationError as ve:
                        messages.error(self.request, str(ve))

                # CashFlow is now automatically handled by Payment.save()

        else:
            # If no payment method selected or amount is 0, clean up any existing payments
            for app in so.receipt_applications.all():
                pay_to_del = app.payment
                if hasattr(pay_to_del, 'cashflow') and pay_to_del.cashflow:
                    pay_to_del.cashflow.delete()
                app.delete()
                pay_to_del.delete()

        messages.success(self.request, f"Sales Order #{so.pk} updated.")
        return redirect(self.get_success_url())

    def get_success_url(self):
        return f"{reverse('so_add')}?business={self.object.business_id}"


# ============================================================================
# PART 2: Complete JavaScript for Template (order_form.html <script> section)
# ============================================================================
"""
Replace the entire <script> section in your order_form.html with this:

NOTE: This includes:
1. Duplicate product prevention with visual warnings
2. Better UX with smooth animations
3. Fixed delete handling
4. Stock warnings
5. Keyboard shortcuts
"""


# --- DELETE & LIST (unchanged from your version) ---



# Import your models (SalesOrder, Product, CashFlow, etc.)

@require_POST
@login_required
def update_sales_order_status_api(request, pk):
    """API endpoint to update sales order status"""
    from django.contrib.auth.hashers import check_password

    try:
        order = get_object_or_404(SalesOrder, pk=pk)
    except SalesOrder.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Sales Order not found'}, status=404)
    
    new_status = request.POST.get('status')
    if not new_status or new_status not in [s[0] for s in SalesOrder.Status.choices]:
        return JsonResponse({'ok': False, 'error': 'Invalid status'}, status=400)
    
    old_status = order.status

    if new_status == SalesOrder.Status.CANCELLED and old_status != SalesOrder.Status.CANCELLED:
        try:
            user_settings = UserSettings.objects.get(user=request.user)
        except UserSettings.DoesNotExist:
            pass
        else:
            stored = (getattr(user_settings, 'cancellation_password', None) or '').strip()
            if stored:
                plain = (request.POST.get('cancellation_password') or '').strip()
                if not plain:
                    return JsonResponse({
                        'ok': False,
                        'error': 'Cancellation password is required to cancel this order.',
                    }, status=400)
                if not check_password(plain, stored):
                    return JsonResponse({
                        'ok': False,
                        'error': 'Incorrect cancellation password.',
                    }, status=400)
    
    with transaction.atomic():
        # If changing to cancelled, reverse stock and delete payments
        if new_status == SalesOrder.Status.CANCELLED and old_status != SalesOrder.Status.CANCELLED:
            # Reverse stock - add quantities back
            for item in order.items.all():
                if item.product and item.quantity:
                    base_qty = item.quantity * (item.size_per_unit or Decimal("1"))
                    Product.objects.filter(pk=item.product_id).update(
                        stock_qty=F('stock_qty') + base_qty
                    )
            
            # Delete receipt applications and associated payments
            receipts = order.receipt_applications.all()
            for receipt in receipts:
                payment = receipt.payment
                # Delete CashFlow if exists
                if hasattr(payment, 'cashflow') and payment.cashflow:
                    payment.cashflow.delete()
                # Delete the payment
                payment.delete()
                # Receipt will be deleted via CASCADE
        
        # If changing from cancelled to another status, deduct stock again
        elif old_status == SalesOrder.Status.CANCELLED and new_status != SalesOrder.Status.CANCELLED:
            # Deduct stock again
            for item in order.items.all():
                if item.product and item.quantity:
                    base_qty = item.quantity * (item.size_per_unit or Decimal("1"))
                    Product.objects.filter(pk=item.product_id).update(
                        stock_qty=F('stock_qty') - base_qty
                    )
        
        # Update status
        order.status = new_status
        order.updated_by = request.user
        order.save(update_fields=['status', 'updated_by', 'updated_at'])
        
        # Auto-set to fulfilled if fully paid
        if order.paid_total >= order.net_total and order.net_total > 0 and new_status != SalesOrder.Status.CANCELLED:
            order.status = SalesOrder.Status.FULFILLED
            order.save(update_fields=['status'])
            new_status = SalesOrder.Status.FULFILLED
    
    return JsonResponse({
        'ok': True,
        'status': new_status,
        'status_display': order.get_status_display(),
        'message': f'Status updated to {order.get_status_display()}'
    })


@require_POST
@login_required
def verify_cancellation_password_api(request):
    """
    Unified API to verify the cancellation/security password.
    Supports both JSON (application/json) and standard FormData (request.POST).
    Used for: supplier ledger, vendors list, party-balances, dashboard reveals.
    """
    from django.contrib.auth.hashers import check_password
    import json
    import time

    # 1. Extract Data (Try JSON first, context: AJAX dashboard reveal)
    data = {}
    if request.content_type == "application/json":
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            pass
    
    # Fallback to request.POST (context: Legacy FormData)
    action = (data.get("action") or request.POST.get("action") or "").strip()
    plain = (
        data.get("password") or 
        request.POST.get("password") or 
        request.POST.get("cancellation_password") or 
        ""
    ).strip()

    if not plain:
        return JsonResponse({"ok": False, "error": "Password is required."}, status=400)

    # 2. Get User Settings
    try:
        user_settings = UserSettings.objects.get(user=request.user)
    except UserSettings.DoesNotExist:
        return JsonResponse({
            "ok": False,
            "error": "Set cancellation password in User Settings first.",
        }, status=400)

    stored = (getattr(user_settings, "cancellation_password", None) or "").strip()
    if not stored:
        return JsonResponse({
            "ok": False,
            "error": "Set cancellation password in User Settings first.",
        }, status=400)

    # 3. Check Password
    if not check_password(plain, stored):
        return JsonResponse({"ok": False, "error": "Incorrect password."}, status=400)

    # 4. Handle Action-based Session Flags
    if action == "supplier_ledger":
        request.session["supplier_ledger_unlocked"] = True
        request.session["supplier_ledger_unlocked_at"] = time.time()
    elif action == "party_balances":
        request.session["party_balances_supplier_unlocked"] = True
        request.session["party_balances_supplier_unlocked_at"] = time.time()
    elif action == "dashboard_reveal":
        # General flag for dashboard if needed, though usually dashboard uses one-time reveal
        request.session["dashboard_revealed_at"] = time.time()
    
    request.session.modified = True
    return JsonResponse({"ok": True})


class SalesOrderDeleteView(LoginRequiredMixin, DeleteView):
    model = SalesOrder
    template_name = "barkat/sales/order_delete.html"

    def get_success_url(self):
        biz = getattr(self.object, "business", None)
        if biz:
            return reverse("so_list_business", args=[biz.pk])
        return reverse("so_list")

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        """
        PERMANENT DELETE LOGIC:
        1. Restore inventory stock levels.
        2. Delete associated CashFlow records (Ledger cleanup).
        3. Delete associated Payment records.
        4. Permanently remove the SalesOrder from the DB.
        """
        self.object = self.get_object()
        oid = self.object.pk

        # --- 1. RESTORE INVENTORY STOCK ---
        # We must do this while the Order object still exists
        order_items = self.object.items.all().only("product_id", "quantity")
        
        restore_map = {}
        for it in order_items:
            if it.product_id and it.quantity > 0:
                restore_map[it.product_id] = restore_map.get(it.product_id, Decimal("0")) + it.quantity

        if restore_map:
            # Lock product rows to prevent stock errors during deletion
            Product.objects.select_for_update().filter(id__in=restore_map.keys())
            for pid, qty in restore_map.items():
                Product.objects.filter(pk=pid).update(stock_qty=F("stock_qty") + qty)

        # --- 2. DELETE CASHFLOW & PAYMENTS ---
        # Find all receipts linked to this order
        receipt_apps = self.object.receipt_applications.select_related('payment')
        
        for app in receipt_apps:
            payment = app.payment
            # Hard delete the linked ledger record first
            if payment.cashflow:
                payment.cashflow.delete()
            
            # Delete the payment record itself
            payment.delete() 

        # --- 3. PERMANENT DELETE ---
        # This removes the SalesOrder and any SalesOrderReceipt bridge rows (due to CASCADE)
        self.object.delete()

        messages.success(request, f"Sales Order #{oid} permanently deleted. Stock restored and ledger entries removed.")
        return redirect(self.get_success_url())


ITEM_TOTAL_PATH = "items__total_amount"
PAYMENT_AMOUNT_PATH = "payments__amount"
# --------------------------------------------------------------------

# barkat/sales/views.py

class SalesOrderListView(LoginRequiredMixin, ListView):
    model = SalesOrder
    template_name = "barkat/sales/order_list.html"
    context_object_name = "orders"
    paginate_by = 20

    # --- heuristics for field names on child models ---
    ITEM_TOTAL_FIELD_CANDIDATES = ["total_amount", "line_total", "subtotal", "amount"]
    PAYMENT_AMOUNT_FIELD_CANDIDATES = ["amount", "received_amount", "paid_amount", "payment_amount", "total_paid", "receipt_amount"]

    def _find_reverse_sum_path(self, model, value_field_candidates):
        """
        Auto-detect a reverse one-to-many relation from `model` that has a numeric field
        named like one of `value_field_candidates`. Returns "<accessor>__<field>" or None.
        """
        for f in model._meta.get_fields():
            if not getattr(f, "one_to_many", False):
                continue
            # Reverse accessor name (e.g. "items", "salesorderitem_set", "payments", etc.)
            accessor = getattr(f, "get_accessor_name", None)
            if not accessor:
                continue
            accessor = f.get_accessor_name()
            child = f.related_model
            if not child:
                continue

            child_field_names = {cf.name for cf in child._meta.get_fields() if hasattr(cf, "attname")}
            for cand in value_field_candidates:
                if cand in child_field_names:
                    return f"{accessor}__{cand}"
        return None

    def _base_filtered_qs(self):
        dec = DecimalField(max_digits=12, decimal_places=2)
        zero_dec = Value(Decimal("0.00"), output_field=dec)

        qs = (
            SalesOrder.objects
            .filter(is_deleted=False, is_active=True)  # Only show active, non-deleted orders
            .select_related("business", "customer", "created_by")
        )

        # text search
        q = (self.request.GET.get("q") or "").strip()
        if q:
            cond = (
                Q(customer__display_name__icontains=q) |
                Q(customer_name__icontains=q) |
                Q(customer_phone__icontains=q) |
                Q(business__name__icontains=q) |
                Q(status__icontains=q)
            )
            if q.isdigit():
                cond |= Q(id=int(q))
            qs = qs.filter(cond)

        # date range (YYYY-MM-DD)
        d_from = (self.request.GET.get("from") or "").strip()
        d_to   = (self.request.GET.get("to") or "").strip()
        if d_from:
            qs = qs.filter(created_at__date__gte=d_from)
        if d_to:
            qs = qs.filter(created_at__date__lte=d_to)

        # --- auto-detect reverse paths ---
        item_sum_path = self._find_reverse_sum_path(SalesOrder, self.ITEM_TOTAL_FIELD_CANDIDATES)
        payment_sum_path = self._find_reverse_sum_path(SalesOrder, self.PAYMENT_AMOUNT_FIELD_CANDIDATES)

        # annotate subtotals from items (fallback to 0 if none found)
        if item_sum_path:
            qs = qs.annotate(subtotal_items=Coalesce(Sum(item_sum_path), zero_dec))
        else:
            qs = qs.annotate(subtotal_items=zero_dec)

        # annotate received from payments (fallback to 0 if none found)
        if payment_sum_path:
            qs = qs.annotate(paid_amount=Coalesce(Sum(payment_sum_path), zero_dec))
        else:
            qs = qs.annotate(paid_amount=zero_dec)

        # prefer stored totals if > 0, else fall back to computed
        qs = qs.annotate(
            subtotal=Case(
                When(total_amount__gt=Decimal("0.00"), then=F("total_amount")),
                default=F("subtotal_items"),
                output_field=dec,
            ),
            net=Case(
                When(net_total__gt=Decimal("0.00"), then=F("net_total")),
                default=F("subtotal"),
                output_field=dec,
            ),
        ).annotate(
            remaining=F("net") - F("paid_amount"),
        )

        # store for debug/context
        self._detected_paths = {
            "item_sum_path": item_sum_path or "(not found)",
            "payment_sum_path": payment_sum_path or "(not found)",
        }
        return qs

    def get_queryset(self):
        self.filtered_qs = self._base_filtered_qs()
        # Order by receipt_number descending (bigger numbers first), then by created_at and id
        # Using nulls_last to handle any NULL receipt_numbers
        return self.filtered_qs.order_by(
            F("id").desc(nulls_last=True),
            "-created_at",
            "-id"
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        dec = DecimalField(max_digits=12, decimal_places=2)
        zero_dec = Value(Decimal("0.00"), output_field=dec)
        # Totals exclude cancelled orders
        non_cancelled = self.filtered_qs.exclude(status=SalesOrder.Status.CANCELLED)
        totals = non_cancelled.aggregate(
            total_subtotal=Coalesce(Sum("subtotal"), zero_dec),
            total_net=Coalesce(Sum("net"), zero_dec),
            total_paid=Coalesce(Sum("paid_amount"), zero_dec),
            total_remaining=Coalesce(Sum("remaining"), zero_dec),
        )

        from barkat.utils.auth_helpers import user_has_cancellation_password
        ctx.update({
            "businesses": Business.objects.order_by("name"),
            "business": None,
            "totals": totals,
            "q": (self.request.GET.get("q") or "").strip(),
            "date_from": (self.request.GET.get("from") or "").strip(),
            "date_to": (self.request.GET.get("to") or "").strip(),
            "debug_paths": getattr(self, "_detected_paths", {}),
            "has_cancellation_password": user_has_cancellation_password(self.request),
        })
        return ctx

class BusinessSalesOrderListView(LoginRequiredMixin, ListView):
    model = SalesOrder
    template_name = "barkat/sales/business_order.html"
    context_object_name = "orders"
    paginate_by = 20

    # ---- heuristics for child fields ----
    ITEM_TOTAL_FIELD_CANDIDATES = ["total_amount", "line_total", "subtotal", "amount"]
    PAYMENT_AMOUNT_FIELD_CANDIDATES = ["amount", "received_amount", "paid_amount", "payment_amount", "total_paid", "receipt_amount"]

    def dispatch(self, request, *args, **kwargs):
        self.business = get_object_or_404(Business, pk=kwargs.get("business_id"))
        return super().dispatch(request, *args, **kwargs)

    def _find_reverse_sum_path(self, model, value_field_candidates):
        """
        Auto-detect a reverse one-to-many relation from `model` that has a numeric field
        named like one of `value_field_candidates`. Returns "<accessor>__<field>" or None.
        """
        for f in model._meta.get_fields():
            if not getattr(f, "one_to_many", False):
                continue
            accessor = getattr(f, "get_accessor_name", None)
            if not accessor:
                continue
            accessor = f.get_accessor_name()
            child = f.related_model
            if not child:
                continue

            child_field_names = {cf.name for cf in child._meta.get_fields() if hasattr(cf, "attname")}
            for cand in value_field_candidates:
                if cand in child_field_names:
                    return f"{accessor}__{cand}"
        return None

    def _base_filtered_qs(self):
        dec = DecimalField(max_digits=12, decimal_places=2)
        zero_dec = Value(Decimal("0.00"), output_field=dec)

        qs = (
            SalesOrder.objects
            .filter(business=self.business, is_deleted=False, is_active=True)  # Only show active, non-deleted orders
            .select_related("business", "customer", "created_by")
        )

        # text search
        q = (self.request.GET.get("q") or "").strip()
        if q:
            cond = (
                Q(customer__display_name__icontains=q) |
                Q(customer_name__icontains=q) |
                Q(customer_phone__icontains=q) |
                Q(status__icontains=q)
            )
            if q.isdigit():
                cond |= Q(id=int(q))
            qs = qs.filter(cond)

        # date range (YYYY-MM-DD)
        d_from = (self.request.GET.get("from") or "").strip()
        d_to   = (self.request.GET.get("to") or "").strip()
        if d_from:
            qs = qs.filter(created_at__date__gte=d_from)
        if d_to:
            qs = qs.filter(created_at__date__lte=d_to)

        # auto-detect reverse paths
        item_sum_path = self._find_reverse_sum_path(SalesOrder, self.ITEM_TOTAL_FIELD_CANDIDATES)
        payment_sum_path = self._find_reverse_sum_path(SalesOrder, self.PAYMENT_AMOUNT_FIELD_CANDIDATES)

        # annotate subtotals from items (fallback to 0 if none found)
        if item_sum_path:
            qs = qs.annotate(subtotal_items=Coalesce(Sum(item_sum_path), zero_dec))
        else:
            qs = qs.annotate(subtotal_items=zero_dec)

        # annotate received from payments (fallback to 0 if none found)
        if payment_sum_path:
            qs = qs.annotate(paid_amount=Coalesce(Sum(payment_sum_path), zero_dec))
        else:
            qs = qs.annotate(paid_amount=zero_dec)

        # prefer stored totals if > 0, else fall back to computed
        qs = qs.annotate(
            subtotal=Case(
                When(total_amount__gt=Decimal("0.00"), then=F("total_amount")),
                default=F("subtotal_items"),
                output_field=dec,
            ),
            net=Case(
                When(net_total__gt=Decimal("0.00"), then=F("net_total")),
                default=F("subtotal"),
                output_field=dec,
            ),
        ).annotate(
            remaining=F("net") - F("paid_amount"),
        )

        # store for optional debug
        self._detected_paths = {
            "item_sum_path": item_sum_path or "(not found)",
            "payment_sum_path": payment_sum_path or "(not found)",
        }
        return qs

    def get_queryset(self):
        self.filtered_qs = self._base_filtered_qs()
        return self.filtered_qs.order_by("-created_at", "-id")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        dec = DecimalField(max_digits=12, decimal_places=2)
        zero_dec = Value(Decimal("0.00"), output_field=dec)
        # Totals exclude cancelled orders
        non_cancelled = self.filtered_qs.exclude(status=SalesOrder.Status.CANCELLED)
        totals = non_cancelled.aggregate(
            total_subtotal=Coalesce(Sum("subtotal"), zero_dec),
            total_net=Coalesce(Sum("net"), zero_dec),
            total_paid=Coalesce(Sum("paid_amount"), zero_dec),
            total_remaining=Coalesce(Sum("remaining"), zero_dec),
        )

        from barkat.utils.auth_helpers import user_has_cancellation_password
        ctx.update({
            "businesses": Business.objects.order_by("name"),
            "business": self.business,
            "totals": totals,
            "q": (self.request.GET.get("q") or "").strip(),
            "date_from": (self.request.GET.get("from") or "").strip(),
            "date_to": (self.request.GET.get("to") or "").strip(),
            "debug_paths": getattr(self, "_detected_paths", {}),
            "has_cancellation_password": user_has_cancellation_password(self.request),
        })
        return ctx

# --------- LIST VIEWS ---------
def ensure_party_for_refund(business, customer, customer_name, customer_phone):
    # Prefer explicit customer; otherwise ensure a per-business "sustomer"
    if customer:
        return customer
    phone = (customer_phone or "").strip()
    party, _ = Party.objects.get_or_create(
        display_name="Walk-in-Customer",
        default_business=business,
        defaults={"type": "CUSTOMER", "is_active": True, "phone": phone},
    )
    if not party.phone and phone:
        party.phone = phone
        party.save(update_fields=["phone"])
    return party

def _product_card_image_url(p):
    try:
        img = p.primary_image()
        if img and img.image:
            return img.image.url
    except Exception:
        pass
    return ""

from django.db.models import Q
from django.views.generic import ListView
from .models import Business, SalesReturn

class _ReturnBaseList(ListView):
    model = SalesReturn
    paginate_by = 25
    context_object_name = "returns"

    def _base_qs(self):
        return (
            SalesReturn.objects
            .select_related("business", "customer")
            .prefetch_related("refund_applications__payment")
            .order_by("-created_at", "-id")
        )

    def _apply_search(self, qs):
        q = (self.request.GET.get("q") or "").strip()
        if not q:
            return qs
        return qs.filter(
            Q(customer__display_name__icontains=q)
            | Q(customer_name__icontains=q)
            | Q(customer_phone__icontains=q)
            | Q(status__icontains=q)
            | Q(id__icontains=q)
        )

class SalesReturnListView(_ReturnBaseList):
    template_name = "barkat/sales/return_order_list.html"

    def get_queryset(self):
        qs = self._base_qs()
        return self._apply_search(qs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["businesses"] = Business.objects.order_by("name", "id")
        return ctx

class SalesReturnBusinessListView(_ReturnBaseList):
    template_name = "barkat/sales/business_return_order.html"

    def get_queryset(self):
        qs = self._base_qs().filter(business_id=self.kwargs["business_id"])
        return self._apply_search(qs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        biz = Business.objects.get(pk=self.kwargs["business_id"])
        ctx["business"] = biz
        ctx["businesses"] = Business.objects.order_by("name", "id")
        return ctx

def _product_image_url(p):
    try:
        img = p.images.first()
        if img and getattr(img, "image", None):
            return img.image.url
    except Exception:
        pass
    return "/static/img/placeholder.png"

def _build_products_cards(qs):
    """
    Build unified product card dicts: id, name, sale_price, category_id, image_url, stock, barcode
    """
    out = []
    for p in qs:
        out.append({
            "id": p.id,
            "name": p.name,
            "sale_price": p.sale_price,
            "category_id": p.category_id or "",
            "image_url": _product_image_url(p),
            "stock": str(p.stock_qty or 0),
            "barcode": getattr(p, "barcode", "") or "",  # Added barcode support
        })
    return out

def _sum_items(items):
    """
    Return {product_id: qty} from an iterable of item objects (SalesReturnItem).
    """
    acc = {}
    for it in items:
        if not it.product_id:
            continue
        acc[it.product_id] = acc.get(it.product_id, 0) + (it.quantity or 0)
    return acc

def _apply_stock_delta(business, delta_by_product_id, user):
    """
    Atomically apply stock delta per product for a given business.
    delta_by_product_id: {product_id: Decimal/float delta}
    """
    if not delta_by_product_id:
        return
    prod_ids = [pid for pid, d in delta_by_product_id.items() if d]
    if not prod_ids:
        return
    # Update one by one with F to be safe and clear; few hundred updates are fine.
    for pid in prod_ids:
        d = delta_by_product_id[pid]
        Product.objects.filter(id=pid, business=business).update(
            stock_qty=F("stock_qty") + d,
            updated_by=user,
        )

def _get_walkin_party(business):
    """Get or return Walk-in-Customer party for refunds."""
    qs = Party.objects.filter(
        is_active=True,
        is_deleted=False,
        display_name__iexact="Walk-in-Customer",
    )
    # Prefer business-specific walk-in party
    if business:
        p = qs.filter(default_business=business).first()
        if p:
            return p
    # Otherwise return global fallback walk-in party
    return qs.first()

@login_required
def sales_order_search_api(request):
    """API endpoint to search sales orders for autocomplete suggestions"""
    q = (request.GET.get("q") or "").strip()
    business_id = request.GET.get("business_id")
    
    qs = SalesOrder.objects.filter(
        status__in=[SalesOrder.Status.OPEN, SalesOrder.Status.FULFILLED]
    ).exclude(status=SalesOrder.Status.CANCELLED).select_related("customer", "business")
    
    if business_id:
        try:
            qs = qs.filter(business_id=int(business_id))
        except ValueError:
            pass
    
    if q:
        # Search by order ID or customer name
        if q.isdigit():
            qs = qs.filter(id=int(q))
        else:
            qs = qs.filter(
                Q(customer_name__icontains=q) |
                Q(customer__display_name__icontains=q) |
                Q(customer_phone__icontains=q)
            )
    
    qs = qs.order_by("-created_at", "-id")[:10]
    
    data = []
    for order in qs:
        customer_name = order.customer_name or (order.customer.display_name if order.customer else "Walk-in")
        data.append({
            "id": order.id,
            "label": f"Order #{order.id} - {customer_name}",
            "order_id": order.id,
            "customer_name": customer_name,
            "date": order.created_at.strftime("%Y-%m-%d") if order.created_at else "",
        })
    
    return JsonResponse(data, safe=False)

@login_required
def sales_order_items_api(request):
    """API endpoint to fetch items from a sales order for return"""
    order_id = request.GET.get("order_id")
    if not order_id:
        return JsonResponse({"error": "order_id required"}, status=400)
    
    try:
        order = SalesOrder.objects.select_related("customer", "business").prefetch_related("items__product").get(pk=order_id)
    except SalesOrder.DoesNotExist:
        return JsonResponse({"error": "Order not found"}, status=404)
    
    items = []
    for item in order.items.all():
        items.append({
            "id": item.id,
            "product_id": item.product_id,
            "product_name": item.product.name,
            "quantity": str(item.quantity),
            "unit_price": str(item.unit_price),
            "line_total": str(item.line_total()),
            "barcode": item.product.barcode or "",
        })
    
    return JsonResponse({
        "order_id": order.id,
        "customer_id": order.customer_id,
        "customer_name": order.customer_name or (order.customer.display_name if order.customer else ""),
        "customer_phone": order.customer_phone or (order.customer.phone if order.customer else ""),
        "business_id": order.business_id,
        "tax_percent": str(order.tax_percent or 0),
        "discount_percent": str(order.discount_percent or 0),
        "items": items,
    })

class SalesReturnCreateView(LoginRequiredMixin, CreateView):
    model = SalesReturn
    form_class = SalesReturnForm
    template_name = "barkat/sales/sales_order_return.html"

    def get_initial(self):
        init = super().get_initial()
        bid = self.request.GET.get("business")
        if bid and bid.isdigit():
            init["business"] = bid
        init.setdefault("status", SalesReturn.Status.PENDING)
        return init

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        business = None
        if getattr(self, "object", None) and self.object.business_id:
            business = self.object.business
        else:
            bid = self.request.POST.get("business") or self.request.GET.get("business")
            if bid and str(bid).isdigit():
                business = Business.objects.filter(pk=int(bid)).first()
        kwargs["business"] = business
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # resolve business
        business = None
        form = ctx.get("form")
        if form and getattr(form, "business", None):
            business = form.business
        if not business:
            bid = self.request.GET.get("business")
            if bid and str(bid).isdigit():
                business = Business.objects.filter(pk=int(bid)).first()

        # tabs
        ctx["businesses"] = Business.objects.order_by("name", "id")
        ctx["business"] = business

        # Global categories and products (matching Sales Order behavior)
        categories = (
            ProductCategory.objects
            .filter(products__is_active=True, products__is_deleted=False)
            .distinct()
            .order_by("name")
        )
        # Products for UI with UOM data (same pattern as sales order)
        products_qs = Product.objects.filter(
            is_active=True, 
            is_deleted=False
        ).select_related("uom", "bulk_uom", "category").order_by("name")
        
        products_cards = []
        for p in products_qs:
            product_data = {
                "id": p.id,
                "name": p.name,
                "sale_price": str(p.sale_price),
                "category_id": p.category_id or "",
                "stock": str(p.stock_qty or 0),
                "barcode": getattr(p, "barcode", "") or "",
                "uom_id": p.uom_id or "",
                "uom_code": p.uom.code if p.uom else "",
                "has_bulk": bool(p.bulk_uom_id and p.default_bulk_size > 0),
            }
            
            # Add bulk unit info if available
            if p.bulk_uom_id and p.default_bulk_size and p.default_bulk_size > 0:
                product_data["bulk_uom_id"] = p.bulk_uom_id
                product_data["bulk_uom_code"] = p.bulk_uom.code if p.bulk_uom else ""
                product_data["bulk_size"] = str(p.default_bulk_size)
            else:
                product_data["bulk_uom_id"] = ""
                product_data["bulk_uom_code"] = ""
                product_data["bulk_size"] = "1"
                
            products_cards.append(product_data)

        ctx["categories"] = categories
        ctx["products_cards"] = products_cards

        # formset
        if self.request.POST:
            ctx["formset"] = SalesReturnItemFormSet(
                self.request.POST,
                form_kwargs={"business": business},
            )
        else:
            ctx["formset"] = SalesReturnItemFormSet(
                form_kwargs={"business": business},
            )

        # totals / refunds (create)
        ctx["refunded_so_far"] = 0
        ctx["remaining_refund"] = 0
        ctx["previous_refunds"] = []
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx["formset"]
        if not formset.is_valid():
            return self.form_invalid(form)

        # save SalesReturn
        self.object = form.save(commit=False)
        
        # Auto-populate customer from source_order if available
        source_order = form.cleaned_data.get("source_order")
        if source_order:
            self.object.source_order = source_order
            # Auto-populate customer info from order
            if source_order.customer:
                self.object.customer = source_order.customer
                self.object.customer_name = source_order.customer.display_name
                self.object.customer_phone = source_order.customer.phone or ""
            elif source_order.customer_name:
                self.object.customer_name = source_order.customer_name
                self.object.customer_phone = source_order.customer_phone or ""
        
        # Handle Walk-in Customer if customer and name are still empty
        customer = form.cleaned_data.get("customer") or self.object.customer
        cname = (form.cleaned_data.get("customer_name") or self.object.customer_name or "").strip()
        if not customer and not cname:
            walkin = _get_walkin_party(form.cleaned_data.get("business"))
            if walkin:
                self.object.customer = walkin
                self.object.customer_name = walkin.display_name
                self.object.customer_phone = walkin.phone or ""
        
        self.object.created_by = self.request.user
        self.object.updated_by = self.request.user
        self.object.save()

        # items (new)
        formset.instance = self.object
        items = formset.save()  # list of saved instances
        
        # Filter out None items (deleted items are already excluded by formset.save())
        items = [item for item in items if item is not None]

        # increase stock for each product by returned quantity
        # Stock should be increased immediately when return is created
        # NOTE: SalesReturnItem.save() only updates stock if status="processed",
        # but we want to update stock immediately regardless of status
        if items:
            for item in items:
                if not item.product_id:
                    continue
                    
                qty = item.quantity or Decimal("0")
                if qty <= 0:
                    continue
                    
                try:
                    # Update stock directly - increase stock when items are returned
                    # Use business filter if available, otherwise update by product_id only
                    if self.object.business_id:
                        updated = Product.objects.filter(
                            id=item.product_id,
                            business=self.object.business
                        ).update(
                            stock_qty=F("stock_qty") + qty,
                            updated_by=self.request.user,
                        )
                        if updated == 0:
                            # Product not found for this business, try without business filter as fallback
                            Product.objects.filter(
                                id=item.product_id
                            ).update(
                                stock_qty=F("stock_qty") + qty,
                                updated_by=self.request.user,
                            )
                    else:
                        # No business set, update by product_id only
                        Product.objects.filter(
                            id=item.product_id
                        ).update(
                            stock_qty=F("stock_qty") + qty,
                            updated_by=self.request.user,
                        )
                except Exception as e:
                    # Log error but don't fail the transaction
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(
                        f"Failed to update stock for product {item.product_id} in sales return {self.object.pk}: {e}",
                        exc_info=True
                    )

        # recompute totals
        self.object.recompute_totals()
        self.object.updated_by = self.request.user
        self.object.save()

        # Handle refund (supports partial and zero refunds)
        method = form.cleaned_data.get("refund_method")   # "cash", "bank", or "card"
        amount = form.cleaned_data.get("refund_amount") or Decimal("0")
        bank   = form.cleaned_data.get("bank_account")
        
        # Convert amount to Decimal if it's not already
        if not isinstance(amount, Decimal):
            try:
                amount = Decimal(str(amount))
            except (ValueError, TypeError):
                amount = Decimal("0")
        
        # Ensure amount doesn't exceed net_total
        if amount > self.object.net_total:
            amount = self.object.net_total

        # Only create Payment and CashFlow if amount > 0 and method is provided
        if method in ("cash", "bank", "card") and amount > 0:
            party = ensure_party_for_refund(
                business=self.object.business,
                customer=form.cleaned_data.get("customer"),
                customer_name=form.cleaned_data.get("customer_name"),
                customer_phone=form.cleaned_data.get("customer_phone"),
            )

            payment_source_value = "bank" if method in ("bank", "card") else "cash"

            payment_kwargs = {
                "business": self.object.business,
                "party": party,
                "date": timezone.now().date(),
                "amount": amount,
                "payment_source": payment_source_value,
                "created_by": self.request.user,
                "updated_by": self.request.user,
                "direction": Payment.OUT,  # refund
            }

            # set payment_method for new Payment model
            if _model_has_field(Payment, "payment_method"):
                payment_kwargs["payment_method"] = method  # "cash" / "bank" / "cheque"

            if method in ("bank", "card") and _model_has_field(Payment, "bank_account") and bank:
                payment_kwargs["bank_account"] = bank

            pay = Payment.objects.create(**payment_kwargs)
            self.object.apply_refund(pay, amount)

            # CashFlow is now automatically handled by Payment.save()

            
            # Check if refund amount equals net_total, then set status to PROCESSED (fulfilled)
            # Reload the object to get updated refunded_total
            self.object.refresh_from_db()
            if self.object.refunded_total >= self.object.net_total:
                self.object.status = SalesReturn.Status.PROCESSED
                self.object.save(update_fields=["status", "updated_by", "updated_at"])
        else:
            # No refund made (amount is 0 or method not provided)
            # Ledger entry for SalesReturn (Cr) will still be created automatically
            # via the ledger.py customer_rows function for registered customers
            # This is correct - customer has a credit balance until refund is made
            pass

        # Note: Ledger entries are handled automatically:
        # - SalesReturn entry (Cr) is created for registered customers via ledger.py
        # - SalesReturnRefund entry (Dr) is created only when apply_refund() is called above
        # This ensures proper ledger tracking even when no refund is made

        messages.success(self.request, f"Sales Return #{self.object.pk} created.")
        return redirect(self.get_success_url())

    def get_success_url(self):
        url = reverse("sr_add")
        biz_id = self.object.business_id if self.object else None
        return f"{url}?business={biz_id}" if biz_id else url

class SalesReturnUpdateView(LoginRequiredMixin, UpdateView):
    model = SalesReturn
    form_class = SalesReturnForm
    template_name = "barkat/sales/sales_order_return.html"
    pk_url_kwarg = "pk"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["business"] = self.object.business
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        sr: SalesReturn = self.object
        business = sr.business

        ctx["businesses"] = Business.objects.order_by("name", "id")
        ctx["business"] = business

        # formset
        if self.request.POST:
            ctx["formset"] = SalesReturnItemFormSet(
                self.request.POST,
                instance=sr,
                form_kwargs={"business": business},
            )
        else:
            ctx["formset"] = SalesReturnItemFormSet(
                instance=sr,
                form_kwargs={"business": business},
            )

        # Global categories and products (matching Sales Order behavior)
        categories = (
            ProductCategory.objects
            .filter(products__is_active=True, products__is_deleted=False)
            .distinct()
            .order_by("name")
        )
        products_qs = (
            Product.objects
            .filter(is_active=True, is_deleted=False)
            .select_related("uom", "category")
            .order_by("name")
        )

        ctx["categories"] = categories
        ctx["products_cards"] = _build_products_cards(products_qs)

        # refund info
        ctx["refunded_so_far"] = sr.refunded_total
        ctx["remaining_refund"] = sr.refund_remaining
        ctx["previous_refunds"] = (
            sr.refund_applications
              .select_related("payment", "payment__bank_account")
              .order_by("created_at")
        )
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx["formset"]
        if not formset.is_valid():
            return self.form_invalid(form)

        # old quantities map before saving changes
        old_items = list(self.object.items.select_related("product").all())
        old_map = _sum_items(old_items)

        # save header
        self.object = form.save(commit=False)
        
        # Handle Walk-in Customer if customer and name are empty
        customer = form.cleaned_data.get("customer")
        cname = (form.cleaned_data.get("customer_name") or "").strip()
        if not customer and not cname:
            walkin = _get_walkin_party(self.object.business)
            if walkin:
                self.object.customer = walkin
                self.object.customer_name = walkin.display_name
                self.object.customer_phone = walkin.phone or ""
        
        self.object.updated_by = self.request.user
        self.object.save()

        # save items
        formset.instance = self.object
        items = formset.save()

        # Compute stock changes: new quantities - old quantities
        # Positive delta = more items returned (stock increases)
        # Negative delta = fewer items returned (stock decreases)
        new_map = _sum_items(items)
        delta_map = {}
        
        # Calculate deltas for all products
        all_pids = set(old_map.keys()) | set(new_map.keys())
        for pid in all_pids:
            old_qty = Decimal(str(old_map.get(pid, 0)))
            new_qty = Decimal(str(new_map.get(pid, 0)))
            delta = new_qty - old_qty
            if delta != 0:
                delta_map[pid] = delta
        
        # Apply stock changes
        for product_id, delta in delta_map.items():
            if not delta:
                continue
                
            try:
                # Update stock with delta (positive = increase, negative = decrease)
                # Use business filter if available, otherwise update by product_id only
                if self.object.business_id:
                    updated = Product.objects.filter(
                        id=product_id,
                        business=self.object.business
                    ).update(
                        stock_qty=F("stock_qty") + delta,
                        updated_by=self.request.user,
                    )
                    if updated == 0:
                        # Product not found for this business, try without business filter as fallback
                        Product.objects.filter(
                            id=product_id
                        ).update(
                            stock_qty=F("stock_qty") + delta,
                            updated_by=self.request.user,
                        )
                else:
                    # No business set, update by product_id only
                    Product.objects.filter(
                        id=product_id
                    ).update(
                        stock_qty=F("stock_qty") + delta,
                        updated_by=self.request.user,
                    )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(
                    f"Failed to update stock for product {product_id} in sales return {self.object.pk}: {e}",
                    exc_info=True
                )

        # recompute totals
        self.object.recompute_totals()
        self.object.updated_by = self.request.user
        self.object.save()

        # optional additional refund (supports partial and zero refunds)
        method = form.cleaned_data.get("refund_method")   # "cash" / "bank" / "card"
        amount = form.cleaned_data.get("refund_amount") or Decimal("0")
        bank   = form.cleaned_data.get("bank_account")
        
        # Convert amount to Decimal if it's not already
        if not isinstance(amount, Decimal):
            try:
                amount = Decimal(str(amount))
            except (ValueError, TypeError):
                amount = Decimal("0")
        
        # Ensure amount doesn't exceed remaining refund
        remaining = self.object.refund_remaining
        if amount > remaining:
            amount = remaining

        # Only create Payment and CashFlow if amount > 0 and method is provided
        if method in ("cash", "bank", "card") and amount > 0:
            party = ensure_party_for_refund(
                business=self.object.business,
                customer=form.cleaned_data.get("customer"),
                customer_name=form.cleaned_data.get("customer_name"),
                customer_phone=form.cleaned_data.get("customer_phone"),
            )
            payment_source_value = "bank" if method in ("bank", "card") else "cash"

            payment_kwargs = {
                "business": self.object.business,
                "party": party,
                "date": timezone.now().date(),
                "amount": amount,
                "payment_source": payment_source_value,
                "created_by": self.request.user,
                "updated_by": self.request.user,
                "direction": Payment.OUT,
            }

            if _model_has_field(Payment, "payment_method"):
                payment_kwargs["payment_method"] = method

            if method in ("bank", "card") and _model_has_field(Payment, "bank_account") and bank:
                payment_kwargs["bank_account"] = bank

            pay = Payment.objects.create(**payment_kwargs)
            self.object.apply_refund(pay, amount)

            # CashFlow is now automatically handled by Payment.save()

            
            # Check if refund amount equals net_total, then set status to PROCESSED (fulfilled)
            # Reload the object to get updated refunded_total
            self.object.refresh_from_db()
            if self.object.refunded_total >= self.object.net_total:
                self.object.status = SalesReturn.Status.PROCESSED
                self.object.save(update_fields=["status", "updated_by", "updated_at"])
        else:
            # No additional refund made (amount is 0 or method not provided)
            # Ledger entries are handled automatically via ledger.py
            pass

        messages.success(self.request, f"Sales Return #{self.object.pk} updated.")
        return redirect(self.get_success_url())

    def get_success_url(self):
        url = reverse("sr_add")
        return f"{url}?business={self.object.business_id}" if self.object.business_id else url

class SalesReturnDeleteView(DeleteView):
    model = SalesReturn
    template_name = "barkat/sales/sales_return_delete.html"

    def get_success_url(self):
        biz = getattr(self.object, "business", None)
        if biz:
            return reverse("sr_list") + f"?business={biz.pk}"
        return reverse("sr_list")

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        oid = self.object.pk
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f"Sales Return #{oid} deleted.")
        return response

class SalesInvoiceListView(LoginRequiredMixin, TemplateView):
    template_name = "barkat/sales/sales_invoice_list.html"
    login_url = "login"

    def get(self, request, *args, **kwargs):
        invoices = (
            SalesInvoice.objects
            .select_related("business", "customer")
            .order_by("-created_at", "-id")
        )
        businesses = Business.objects.order_by("name")
        ctx = {
            "invoices": invoices,
            "businesses": businesses,
        }
        return render(request, self.template_name, ctx)

class BusinessSalesInvoiceListView(LoginRequiredMixin, TemplateView):
    template_name = "barkat/sales/business_sales_invoice.html"
    login_url = "login"

    def get(self, request, business_id, *args, **kwargs):
        business = get_object_or_404(Business, pk=business_id)
        invoices = (
            SalesInvoice.objects
            .select_related("business", "customer")
            .filter(business=business)
            .order_by("-created_at", "-id")
        )
        businesses = Business.objects.order_by("name")
        ctx = {
            "business": business,
            "invoices": invoices,
            "businesses": businesses,
        }
        return render(request, self.template_name, ctx)


# --------- EDIT VIEW ---------

class SalesInvoiceUpdateView(LoginRequiredMixin, UpdateView):
    model = SalesInvoice
    form_class = SalesInvoiceForm
    template_name = "barkat/sales/sales_invoice_form.html"
    login_url = "login"

    def get_success_url(self):
        return reverse("sales_invoice_edit", args=[self.object.pk])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        inv: SalesInvoice = self.object

        if self.request.method == "POST":
            ctx["formset"] = SalesInvoiceItemFormSet(self.request.POST, instance=inv)
        else:
            ctx["formset"] = SalesInvoiceItemFormSet(instance=inv)

        # payments info
        paid_so_far = inv.paid_total
        remaining = inv.balance_due
        previous_receipts = (
            inv.receipt_applications
              .select_related("payment", "payment__bank_account")
              .order_by("-payment__date", "-id")
        )

        ctx.update({
            "paid_so_far": paid_so_far,
            "remaining": remaining,
            "previous_receipts": previous_receipts,
        })
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        ctx = self.get_context_data()
        formset: SalesInvoiceItemFormSet = ctx["formset"]
        if not formset.is_valid():
            return self.form_invalid(form)

        inv: SalesInvoice = form.save(commit=False)
        inv.updated_by = self.request.user
        inv.save()

        formset.instance = inv
        formset.save()

        # recompute totals and persist
        inv.recompute_totals()
        inv.save(update_fields=["total_amount", "net_total", "updated_at", "updated_by"])

        # optional receipt
        method = form.cleaned_data.get("receipt_method") or "none"
        received_amount = form.cleaned_data.get("received_amount") or Decimal("0.00")
        bank = form.cleaned_data.get("bank_account")

        if received_amount and received_amount > 0 and method in {"cash", "bank"}:
            pay = Payment(
                business=inv.business,
                date=inv.created_at.date(),
                party=inv.customer if inv.customer_id else None,
                direction=Payment.IN,
                amount=received_amount,
                description=f"Receipt for Invoice {inv.invoice_no}",
                reference=str(inv.pk),
                payment_source=Payment.BANK if method == "bank" else Payment.CASH,
                bank_account=bank if method == "bank" else None,
                created_by=self.request.user,
                updated_by=self.request.user,
            )
            pay.full_clean()
            pay.save()

            # link to invoice
            inv.apply_receipt(pay, received_amount)

        messages.success(self.request, f"Invoice {inv.invoice_no} saved.")
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        ctx = self.get_context_data()
        ctx["form"] = form
        return self.render_to_response(ctx)
    

#---------- REPORTS ----------------

def make_aware_safe(dt: datetime) -> datetime:
    """Return tz-aware dt if USE_TZ=True; otherwise return dt unchanged."""
    if settings.USE_TZ and is_naive(dt):
        return make_aware(dt, timezone.get_current_timezone())
    return dt

# ---------- small date helpers ----------
def _parse_dt(val: str | None) -> datetime | None:
    """
    Accepts 'YYYY-MM-DD' or 'YYYY-MM-DDTHH:MM'.
    Returns tz-aware datetime when USE_TZ=True.
    """
    if not val:
        return None
    try:
        if "T" in val:
            dt = datetime.strptime(val, "%Y-%m-%dT%H:%M")
        else:
            # date-only → start of day
            dt = datetime.strptime(val, "%Y-%m-%d").replace(hour=0, minute=0, second=0, microsecond=0)
    except Exception:
        return None
    return make_aware_safe(dt)

def _daterange_days(start: date, end: date):
    cur = start
    while cur <= end:
        yield cur
        cur = cur + timedelta(days=1)

def _month_labels(last_n=12, end_month: date | None = None):
    """
    Return (labels[YYYY-MM], month_starts[date], month_ends[date])
    ending at end_month (default: this month).
    """
    if end_month is None:
        now = timezone.localdate()
        end_month = date(year=now.year, month=now.month, day=1)
    labels, starts, ends = [], [], []
    y, m = end_month.year, end_month.month
    for _ in range(last_n):
        starts.insert(0, date(y, m, 1))
        # next month start
        if m == 12:
            ny, nm = y + 1, 1
        else:
            ny, nm = y, m + 1
        # end = next start - 1 day
        next_start = date(ny, nm, 1)
        ends.insert(0, next_start - timedelta(days=1))
        labels.insert(0, f"{y:04d}-{m:02d}")
        # prev month
        if m == 1:
            y, m = y - 1, 12
        else:
            m -= 1
    return labels, starts, ends

# ---------- Decimal ZERO constants with correct output_field ----------
D0  = V(Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2))  # money
DQ0 = V(Decimal("0"),     output_field=DecimalField(max_digits=18, decimal_places=6)) # quantities


# ===========================================
#                VIEW
# ===========================================


# D0 and DQ0 helpers
D0  = Value(Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2))
DQ0 = Value(Decimal("0.000000"), output_field=DecimalField(max_digits=18, decimal_places=6))



# your helpers
# _parse_dt, make_aware_safe, _daterange_days, _month_labels must exist already

# safe Decimal zeros
D0  = Decimal("0.00")
DQ0 = Decimal("0.000000")


@login_required
def finance_reports(request):
    """
    Reports & Analytics
    Query params:
      - from, to . datetime-local or YYYY-MM-DD
      - business . int optional
      - mode . 'simple' or 'recipe'
    """
    now = timezone.localtime()
    from_raw = request.GET.get("from") or ""
    to_raw = request.GET.get("to") or ""

    dt_from = _parse_dt(from_raw) or make_aware_safe(datetime(now.year, now.month, 1, 0, 0, 0))
    dt_to = _parse_dt(to_raw) or make_aware_safe(datetime(now.year, now.month, now.day, 23, 59, 59, 0))
    if dt_from > dt_to:
        dt_from, dt_to = dt_to, dt_from

    d_from = dt_from.date()
    d_to = dt_to.date()

    business_id = request.GET.get("business")
    business = None
    if business_id and str(business_id).isdigit():
        business = Business.objects.filter(pk=int(business_id)).first()

    mode = request.GET.get("mode") or "simple"

    # Sales orders . totals and series
    so_filter = Q(created_at__range=(dt_from, dt_to)) & ~Q(status="cancelled")
    if business:
        so_filter &= Q(business=business)
    orders_qs = SalesOrder.objects.filter(so_filter)

    revenue_total = orders_qs.aggregate(
        s=Coalesce(Sum("net_total", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"]

    so_by_day = (
        orders_qs
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(val=Coalesce(Sum("net_total", output_field=DecimalField(max_digits=18, decimal_places=2)), D0))
        .values("day", "val")
    )
    rev_map = {row["day"]: row["val"] for row in so_by_day}

    cancelled_q = SalesOrder.objects.filter(created_at__range=(dt_from, dt_to), status="cancelled")
    if business:
        cancelled_q = cancelled_q.filter(business=business)
    cancelled_total = cancelled_q.aggregate(
        s=Coalesce(Sum("net_total", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"]
    
    # Receipt number information (using SalesOrder.id as receipt number)
    # Get all sales orders in period (non-cancelled) for receipt count and range
    receipt_ids = list(orders_qs.values_list('id', flat=True).order_by('id'))
    receipt_count = len(receipt_ids)
    receipt_min = min(receipt_ids) if receipt_ids else None
    receipt_max = max(receipt_ids) if receipt_ids else None
    receipt_series = None
    if receipt_min and receipt_max:
        if receipt_min == receipt_max:
            receipt_series = f"Receipt #{receipt_min}"
        else:
            receipt_series = f"Receipt #{receipt_min} to #{receipt_max}"
    
    # Get cancelled receipt numbers
    cancelled_receipt_ids = list(cancelled_q.values_list('id', flat=True).order_by('id'))
    cancelled_receipt_count = len(cancelled_receipt_ids)
    cancelled_receipt_numbers = ", ".join(map(str, cancelled_receipt_ids)) if cancelled_receipt_ids else "None"

    # Purchases . Simple totals for reference only
    po_qs_base = PurchaseOrder.objects.filter(created_at__range=(dt_from, dt_to))
    if business:
        po_qs_base = po_qs_base.filter(business=business)

    # Other expenses
    exp_filter = Q(date__range=(d_from, d_to))
    if business:
        exp_filter &= Q(business=business)

    expense_total_all = Expense.objects.filter(exp_filter).aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"]

    # --- NEW: Landed PO Expenses vs Operating Expenses ---
    landed_po_expenses_total = Expense.objects.filter(exp_filter, purchase_order__isnull=False).aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0

    operating_expenses_total = Expense.objects.filter(exp_filter, purchase_order__isnull=True).aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0

    # For backward compatibility in case old code expects expense_total
    expense_total = expense_total_all

    exp_by_day = (
        Expense.objects.filter(exp_filter)
        .values("date")
        .annotate(amt=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0))
        .values("date", "amt")
    )
    expense_by_day_map = {row["date"]: row["amt"] for row in exp_by_day}

    # ---------------------------------------------------------------------
    # CASH CALCULATIONS - NEW STRUCTURE
    # ---------------------------------------------------------------------
    
    # Get all cash IN payments in period
    all_cash_in_qs = Payment.objects.filter(
        direction=Payment.IN, 
        payment_source=Payment.CASH,
        date__range=(d_from, d_to)
    )
    if business:
        all_cash_in_qs = all_cash_in_qs.filter(business=business)
    
    # Sales Cash: payments linked to sales orders/invoices
    sales_cash_payments = all_cash_in_qs.filter(
        Q(applied_sales_orders__isnull=False) | 
        Q(applied_sales_invoices__isnull=False)
    ).distinct()
    
    # Receipts Cash: payments NOT linked to sales (general receipts)
    receipts_cash_payments = all_cash_in_qs.exclude(
        Q(applied_sales_orders__isnull=False) | 
        Q(applied_sales_invoices__isnull=False)
    )
    
    # Calculate cash totals
    kpi_sales_cash = sales_cash_payments.aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0
    
    kpi_receipts_cash = receipts_cash_payments.aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0
    
    # Cash OUT payments
    cash_out_qs = Payment.objects.filter(
        direction=Payment.OUT, 
        payment_source=Payment.CASH,
        date__range=(d_from, d_to)
    )
    if business:
        cash_out_qs = cash_out_qs.filter(business=business)
    
    cash_out_total = cash_out_qs.aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0
    
    # Bank transactions
    pay_qs = Payment.objects.filter(date__range=(d_from, d_to))
    if business:
        pay_qs = pay_qs.filter(business=business)

    # Bank in and out (payments)
    bank_in_all_qs = pay_qs.filter(direction=Payment.IN, payment_source=Payment.BANK)
    bank_out_all_qs = pay_qs.filter(direction=Payment.OUT, payment_source=Payment.BANK)

    # total bank collected. all IN including cheques
    kpi_bank_collected = bank_in_all_qs.aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0

    # bank cash in only. exclude cheques
    bank_cash_in_qs = bank_in_all_qs.exclude(payment_method=Payment.PaymentMethod.CHEQUE)
    kpi_bank_cash_in = bank_cash_in_qs.aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0

    # Cheques in (bank side)
    cheque_qs = bank_in_all_qs.filter(payment_method=Payment.PaymentMethod.CHEQUE)

    kpi_cheque_in_hand_pending = cheque_qs.filter(
        cheque_status=Payment.ChequeStatus.PENDING
    ).aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0

    cheque_deposited_qs = Payment.objects.filter(
        direction=Payment.IN,
        payment_source=Payment.BANK,
        payment_method=Payment.PaymentMethod.CHEQUE,
        cheque_status=Payment.ChequeStatus.DEPOSITED,
        updated_at__date__range=(d_from, d_to),
    )
    if business:
        cheque_deposited_qs = cheque_deposited_qs.filter(business=business)

    kpi_cheque_in_hand_deposited = cheque_deposited_qs.aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0

    # For amount in hand card. only pending part
    kpi_cheque_in_hand = kpi_cheque_in_hand_pending

    # Amount In card splits
    amount_in_cash = kpi_sales_cash + kpi_receipts_cash  # Combined sales and receipts cash
    amount_in_cheque_deposited = kpi_cheque_in_hand_deposited
    amount_in_bank_cash = kpi_bank_cash_in
    amount_in_total = (amount_in_cash or D0) + (amount_in_cheque_deposited or D0) + (amount_in_bank_cash or D0)

    # Deposit table sources
    deposit_bank_qs = bank_cash_in_qs.select_related("bank_account")
    deposit_cheque_qs = cheque_deposited_qs.select_related("bank_account")

    deposit_bank_rows = []
    for p in deposit_bank_qs:
        deposit_bank_rows.append(
            {
                "date": p.date.strftime("%Y-%m-%d"),
                "account": p.bank_account.name if getattr(p, "bank_account", None) else "Bank",
                "ref": p.reference or p.description or "",
                "amount": p.amount or D0,
            }
        )

    deposit_cheque_rows = []
    for p in deposit_cheque_qs:
        deposit_cheque_rows.append(
            {
                "date": p.updated_at.strftime("%Y-%m-%d"),
                "account": p.bank_account.name if getattr(p, "bank_account", None) else "Bank",
                "ref": p.reference or p.description or "",
                "amount": p.amount or D0,
            }
        )

    # CASH IN HAND CALCULATION (Total cash balance)
    # Starting from sales cash + receipts cash, subtract cash out payments
    # and adjust for BankMovement cash transactions
    
    # Initial calculation from payments
    kpi_cash_in_hand = (kpi_sales_cash + kpi_receipts_cash) - (cash_out_total or D0)
    
    # Adjust for BankMovement cash transactions
    bm_filters = {"date__range": (d_from, d_to)}
    if business and hasattr(BankMovement, "business_id"):
        bm_filters["business"] = business
    bm_qs = BankMovement.objects.select_related("from_bank", "to_bank").filter(**bm_filters).order_by("date", "id")
    
    cash_delta_from_bm = D0
    for mv in bm_qs:
        amt = getattr(mv, "amount", D0) or D0
        mtype = (getattr(mv, "movement_type", "") or "").lower()
        
        if mtype in ("deposit", "cash_deposit"):
            cash_delta_from_bm -= amt  # Cash out to bank
        elif mtype in ("withdraw", "withdrawal", "cash_withdrawal"):
            cash_delta_from_bm += amt  # Cash in from bank
    
    # Adjust for cash expenses without Payment mirror
    cash_exp_filters = {"date__range": (d_from, d_to)}
    if business:
        cash_exp_filters["business"] = business
    exp_cash_qs = Expense.objects.filter(**cash_exp_filters)

    cash_source_val = getattr(Expense, "PAYMENT_SOURCE_CASH", None) or getattr(Expense, "SOURCE_CASH", None)
    if cash_source_val is not None and hasattr(Expense, "payment_source"):
        exp_cash_qs = exp_cash_qs.filter(payment_source=cash_source_val)
    elif hasattr(Expense, "payment_source"):
        exp_cash_qs = exp_cash_qs.filter(payment_source__in=["cash", "CASH"])

    # We exclude expenses that already have a linked Payment object 
    # because those outflows are already counted in 'cash_out_total'.
    exp_cash_qs = exp_cash_qs.filter(payment__isnull=True)

    extra_cash_exp_total = exp_cash_qs.aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0

    # Final cash in hand calculation
    # Start with cash in (sales + receipts), subtract all cash out payments
    # Note: kpi_cash_out_po and kpi_cash_out_sr_refund are calculated below and will be part of cash_out_total
    # But we want to show them separately, so we keep the existing calculation
    kpi_cash_in_hand = (kpi_cash_in_hand or D0) - (extra_cash_exp_total or D0)
    kpi_cash_in_hand = (kpi_cash_in_hand or D0) + (cash_delta_from_bm or D0)
    
    # Note: cash_out_total already includes PO payments and SR refunds
    # So the calculation above is correct - it's (sales cash + receipts cash) - (all cash out) - adjustments

    # Receivable remaining for sales orders in range
    paid_pairs_kpi = orders_qs.annotate(
        paid=Coalesce(
            Sum("receipt_applications__amount", output_field=DecimalField(max_digits=18, decimal_places=2)),
            D0,
        )
    ).values_list("id", "paid")
    so_paid_map_kpi = {sid: (paid or D0) for sid, paid in paid_pairs_kpi}

    remaining_receivable = D0
    for so in orders_qs.only("id", "net_total"):
        paid = so_paid_map_kpi.get(so.id, D0)
        remaining_receivable += (so.net_total or D0) - paid
    kpi_remaining = remaining_receivable

    # Bank rows from Payment (for detailed tables)
    bank_in_qs = bank_in_all_qs.select_related("bank_account").order_by("date", "id")
    bank_out_qs = bank_out_all_qs.select_related("bank_account").order_by("date", "id")

    bank_in_rows = []
    for p in bank_in_qs:
        ref = p.reference or p.description or ""
        bank_in_rows.append(
            {
                "date": p.date.strftime("%Y-%m-%d"),
                "account": (p.bank_account.name if getattr(p, "bank_account", None) else "Bank"),
                "ref": ref,
                "amount": p.amount or D0,
            }
        )

    bank_out_rows = []
    for p in bank_out_qs:
        exp = getattr(p, "expense", None) if hasattr(p, "expense") else None
        if exp is not None:
            parts = ["Expense"]
            cat = getattr(exp, "get_category_display", None)
            if callable(cat):
                parts.append(cat())
            elif getattr(exp, "category", None):
                parts.append(str(exp.category))
            desc = getattr(exp, "description", "") or ""
            if desc:
                parts.append(desc)
            ref = " . ".join(parts)
        else:
            ref = p.reference or p.description or ""

        bank_out_rows.append(
            {
                "date": p.date.strftime("%Y-%m-%d"),
                "account": (p.bank_account.name if getattr(p, "bank_account", None) else "Bank"),
                "ref": ref,
                "amount": p.amount or D0,
            }
        )

    # Fold in BankMovement to bank IN and OUT tables
    def _bank_name(bank):
        try:
            return bank.name or getattr(bank, "bank_name", None) or "Bank"
        except Exception:
            return "Bank"

    def _mv_label(mv):
        mt = getattr(mv, "get_movement_type_display", None)
        mt_text = mt() if callable(mt) else getattr(mv, "movement_type", "") or "Movement"
        parts = [mt_text]
        method = getattr(mv, "method", "") or ""
        if method:
            parts.append(method)
        refno = getattr(mv, "reference_no", "") or ""
        if refno:
            parts.append(f"Ref {refno}")
        notes = getattr(mv, "notes", "") or ""
        if notes:
            parts.append(notes)
        return " . ".join(parts)

    for mv in bm_qs:
        amt = getattr(mv, "amount", D0) or D0
        d_str = mv.date.strftime("%Y-%m-%d")
        ref = _mv_label(mv)
        mtype = (getattr(mv, "movement_type", "") or "").lower()

        if mtype in ("deposit", "cash_deposit"):
            account = _bank_name(getattr(mv, "to_bank", None))
            bank_in_rows.append({"date": d_str, "account": account, "ref": ref, "amount": amt})
            continue

        if mtype in ("withdraw", "withdrawal", "cash_withdrawal"):
            account = _bank_name(getattr(mv, "from_bank", None))
            bank_out_rows.append({"date": d_str, "account": account, "ref": ref, "amount": amt})
            continue

        if mtype in ("transfer", "bank_to_bank", "move"):
            from_acc = _bank_name(getattr(mv, "from_bank", None))
            to_acc = _bank_name(getattr(mv, "to_bank", None))
            bank_out_rows.append({"date": d_str, "account": from_acc, "ref": ref, "amount": amt})
            bank_in_rows.append({"date": d_str, "account": to_acc, "ref": ref, "amount": amt})
            continue

        has_from = getattr(mv, "from_bank_id", None) is not None
        has_to = getattr(mv, "to_bank_id", None) is not None
        if has_to and not has_from:
            account = _bank_name(getattr(mv, "to_bank", None))
            bank_in_rows.append({"date": d_str, "account": account, "ref": ref, "amount": amt})
        elif has_from and not has_to:
            account = _bank_name(getattr(mv, "from_bank", None))
            bank_out_rows.append({"date": d_str, "account": account, "ref": ref, "amount": amt})
        elif has_from and has_to:
            from_acc = _bank_name(getattr(mv, "from_bank", None))
            to_acc = _bank_name(getattr(mv, "to_bank", None))
            bank_out_rows.append({"date": d_str, "account": from_acc, "ref": ref, "amount": amt})
            bank_in_rows.append({"date": d_str, "account": to_acc, "ref": ref, "amount": amt})
        else:
            pass

    # Bank-paid expenses directly
    exp_bank_filters = {"date__range": (d_from, d_to)}
    if business:
        exp_bank_filters["business"] = business
    exp_bank_qs = Expense.objects.filter(**exp_bank_filters)

    bank_source_val = getattr(Expense, "PAYMENT_SOURCE_BANK", None) or getattr(Expense, "SOURCE_BANK", None)
    if bank_source_val is not None and hasattr(Expense, "payment_source"):
        exp_bank_qs = exp_bank_qs.filter(payment_source=bank_source_val)
    elif hasattr(Expense, "payment_source"):
        exp_bank_qs = exp_bank_qs.filter(payment_source__in=["bank", "BANK"])
    elif hasattr(Expense, "bank_account"):
        exp_bank_qs = exp_bank_qs.filter(bank_account__isnull=False)

    # We exclude expenses that already have a linked Payment object
    exp_bank_qs = exp_bank_qs.filter(payment__isnull=True)
    # Bank-paid expenses directly

    for e in exp_bank_qs:
        if not getattr(e, "amount", None):
            continue
        if getattr(e, "bank_account_id", None):
            ba = getattr(e, "bank_account", None)
            account_name = getattr(ba, "name", None) or "Bank"
        else:
            account_name = "Bank"

        parts = ["Expense"]
        cat_disp = getattr(e, "get_category_display", None)
        if callable(cat_disp):
            parts.append(cat_disp())
        elif getattr(e, "category", None):
            parts.append(str(e.category))
        desc = getattr(e, "description", "") or ""
        if desc:
            parts.append(desc)
        ref = " . ".join(parts)

        bank_out_rows.append(
            {
                "date": e.date.strftime("%Y-%m-%d"),
                "account": account_name,
                "ref": ref,
                "amount": e.amount or D0,
            }
        )

    bank_in_total = sum((r["amount"] or D0) for r in bank_in_rows) if bank_in_rows else D0
    bank_out_total = sum((r["amount"] or D0) for r in bank_out_rows) if bank_out_rows else D0

    kpi_bank_deposited = bank_in_total or D0
    kpi_bank_amount = (bank_in_total or D0) - (bank_out_total or D0)
    # kpi_bank_revenue will be calculated below specifically from sales-linked payments

    # ---------------------------------------------------------------------
    # NEW: Cash Out via Purchase Orders (Cash payments for POs)
    # ---------------------------------------------------------------------
    po_cash_payments_qs = Payment.objects.filter(
        direction=Payment.OUT,
        payment_source=Payment.CASH,
        payment_method=Payment.PaymentMethod.CASH,
        date__range=(d_from, d_to)
    )
    if business:
        po_cash_payments_qs = po_cash_payments_qs.filter(business=business)
    
    # Get payments linked to Purchase Orders
    po_cash_payments_qs = po_cash_payments_qs.filter(
        applied_purchase_orders__isnull=False
    ).distinct()
    
    kpi_cash_out_po = po_cash_payments_qs.aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0

    # ---------------------------------------------------------------------
    # NEW: Cash Out via Sales Return Refunds (Cash refunds for Sales Returns)
    # ---------------------------------------------------------------------
    sr_cash_refunds_qs = Payment.objects.filter(
        direction=Payment.OUT,
        payment_source=Payment.CASH,
        payment_method=Payment.PaymentMethod.CASH,
        date__range=(d_from, d_to)
    )
    if business:
        sr_cash_refunds_qs = sr_cash_refunds_qs.filter(business=business)
    
    # Get payments linked to Sales Returns
    sr_cash_refunds_qs = sr_cash_refunds_qs.filter(
        applied_sales_returns__isnull=False
    ).distinct()
    
    kpi_cash_out_sr_refund = sr_cash_refunds_qs.aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0

    # ---------------------------------------------------------------------
    # NEW: Cash Out - General (Standalone payments)
    # ---------------------------------------------------------------------
    # Everything in cash_out_total that isn't already categorized
    kpi_cash_out_general = (cash_out_total or D0) - (kpi_cash_out_po or D0) - (kpi_cash_out_sr_refund or D0)

    # ---------------------------------------------------------------------
    # NEW: Bank Sales Amount per Bank Account
    # ---------------------------------------------------------------------
    # Get bank payments from sales (IN, BANK, linked to sales orders or invoices)
    # This matches the logic used for kpi_sales_cash/kpi_receipts_cash
    bank_sales_qs = Payment.objects.filter(
        direction=Payment.IN,
        payment_source=Payment.BANK,
        date__range=(d_from, d_to)
    ).filter(
        Q(applied_sales_orders__isnull=False) |
        Q(applied_sales_invoices__isnull=False)
    ).distinct()
    
    if business:
        bank_sales_qs = bank_sales_qs.filter(business=business)
    
    bank_sales_by_account = {}
    kpi_bank_revenue = D0  # Reset and sum specifically from sales payments
    for payment in bank_sales_qs.select_related("bank_account"):
        bank_name = payment.bank_account.name if payment.bank_account else "Unknown Bank"
        bank_sales_by_account.setdefault(bank_name, D0)
        bank_sales_by_account[bank_name] += payment.amount or D0
        kpi_bank_revenue += payment.amount or D0

    # ---------------------------------------------------------------------
    # NEW: Bank Deposited Amount per Bank Account (Cash deposits only)
    # ---------------------------------------------------------------------
    # Cash deposits via BankMovement (deposit type)
    bank_deposits_cash_by_account = {}
    for mv in bm_qs:
        mtype = (getattr(mv, "movement_type", "") or "").lower()
        if mtype in ("deposit", "cash_deposit"):
            bank = getattr(mv, "to_bank", None)
            if bank:
                bank_name = getattr(bank, "name", None) or "Unknown Bank"
                amt = getattr(mv, "amount", D0) or D0
                bank_deposits_cash_by_account.setdefault(bank_name, D0)
                bank_deposits_cash_by_account[bank_name] += amt
    
    # Also include bank cash in payments (non-cheque bank payments that are IN)
    # EXCLUDING the ones already counted in bank_sales_qs to avoid double-counting
    bank_general_cash_in_payments = bank_cash_in_qs.exclude(
        Q(applied_sales_orders__isnull=False) |
        Q(applied_sales_invoices__isnull=False)
    ).select_related("bank_account").distinct()

    for payment in bank_general_cash_in_payments:
        bank_name = payment.bank_account.name if payment.bank_account else "Unknown Bank"
        bank_deposits_cash_by_account.setdefault(bank_name, D0)
        bank_deposits_cash_by_account[bank_name] += payment.amount or D0

    # ---------------------------------------------------------------------
    # NEW: Cheque Deposited per Bank Account
    # ---------------------------------------------------------------------
    cheque_deposited_by_account = {}
    for payment in cheque_deposited_qs.select_related("bank_account"):
        bank_name = payment.bank_account.name if payment.bank_account else "Unknown Bank"
        cheque_deposited_by_account.setdefault(bank_name, D0)
        cheque_deposited_by_account[bank_name] += payment.amount or D0

    # ---------------------------------------------------------------------
    # NEW: Calculate Total Deposited and Total Bank Amount per Bank
    # ---------------------------------------------------------------------
    # Get all unique bank accounts
    all_bank_names = set()
    all_bank_names.update(bank_sales_by_account.keys())
    all_bank_names.update(bank_deposits_cash_by_account.keys())
    all_bank_names.update(cheque_deposited_by_account.keys())
    
    bank_summaries = []
    grand_total_banks = D0
    
    for bank_name in sorted(all_bank_names):
        bank_sales_amt = bank_sales_by_account.get(bank_name, D0)
        bank_deposit_cash = bank_deposits_cash_by_account.get(bank_name, D0)
        cheque_deposit = cheque_deposited_by_account.get(bank_name, D0)
        
        total_deposited = (bank_deposit_cash or D0) + (cheque_deposit or D0)
        
        # Calculate bank out for this account (from bank_out_rows)
        bank_out_for_account = sum(
            (r.get("amount") or D0) 
            for r in bank_out_rows 
            if r.get("account") == bank_name
        ) or D0
        
        # Total bank amount = sales + deposits - withdrawals
        total_bank_amount = (bank_sales_amt or D0) + (total_deposited or D0) - (bank_out_for_account or D0)
        
        bank_summaries.append({
            "bank_name": bank_name,
            "bank_sales": bank_sales_amt,
            "bank_deposited_cash": bank_deposit_cash,
            "cheque_deposited": cheque_deposit,
            "total_deposited": total_deposited,
            "total_bank_amount": total_bank_amount,
        })
        
        grand_total_banks += total_bank_amount

    # Purchase Orders table
    po_qs = po_qs_base.select_related("supplier").prefetch_related("items").order_by("-created_at", "-id")

    po_item_rows = (
        PurchaseOrderItem.objects.filter(purchase_order__in=po_qs)
        .values("purchase_order_id", "product__name")
        .annotate(
            qty=Coalesce(Sum("quantity", output_field=DecimalField(max_digits=18, decimal_places=6)), DQ0),
            unit=Coalesce(Sum(F("unit_price"), output_field=DecimalField(max_digits=18, decimal_places=2)), D0),
            line=Coalesce(
                Sum(F("quantity") * F("unit_price"), output_field=DecimalField(max_digits=18, decimal_places=2)),
                D0,
            ),
        )
    )
    po_items_map: dict[int, list] = {}
    for r in po_item_rows:
        po_items_map.setdefault(r["purchase_order_id"], []).append(
            {
                "product": r["product__name"],
                "qty": r["qty"],
                "unit_price": r["unit"],
                "line_total": r["line"],
            }
        )

    po_paid_map = {
        row["purchase_order"]: row["amt"]
        for row in PurchaseOrderPayment.objects
        .filter(purchase_order__in=po_qs, payment__date__range=(d_from, d_to))
        .values("purchase_order")
        .annotate(amt=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0))
    }

    po_rows = []
    po_count = po_qs.count()
    po_paid_total = D0
    po_remaining_total = D0

    for po in po_qs:
        paid = po_paid_map.get(po.id, D0) or D0
        total = po.net_total or D0
        remaining = total - paid
        po_paid_total += paid
        po_remaining_total += remaining
        po_rows.append(
            {
                "id": po.id,
                "date": po.created_at.strftime("%Y-%m-%d"),
                "status": getattr(po, "status", "") or "",
                "supplier": getattr(po.supplier, "display_name", None) or "",
                "items_count": len(po_items_map.get(po.id, [])),
                "items": po_items_map.get(po.id, []),
                "total": total,
                "paid": paid,
                "remaining": remaining,
            }
        )

    # Sales Orders table
    so_qs = orders_qs.select_related("customer").prefetch_related("items").order_by("-created_at", "-id")

    so_item_rows = (
        SalesOrderItem.objects
        .filter(sales_order__in=so_qs)
        .values("sales_order_id", "product__name")
        .annotate(
            qty=Coalesce(Sum("quantity", output_field=DecimalField(max_digits=18, decimal_places=6)), DQ0),
            unit=Coalesce(Sum(F("unit_price"), output_field=DecimalField(max_digits=18, decimal_places=2)), D0),
            line=Coalesce(
                Sum(F("quantity") * F("unit_price"), output_field=DecimalField(max_digits=18, decimal_places=2)),
                D0,
            ),
        )
    )
    so_items_map: dict[int, list] = {}
    for r in so_item_rows:
        so_items_map.setdefault(r["sales_order_id"], []).append(
            {
                "product": r["product__name"],
                "qty": r["qty"],
                "unit_price": r["unit"],
                "line_total": r["line"],
            }
        )

    paid_pairs2 = (
        so_qs.annotate(
            paid=Coalesce(
                Sum("receipt_applications__amount", output_field=DecimalField(max_digits=18, decimal_places=2)),
                D0,
            )
        ).values_list("id", "paid")
    )
    so_paid_map = {sid: (paid or D0) for sid, paid in paid_pairs2}

    so_rows = []
    so_count = so_qs.count()
    for so in so_qs:
        paid = so_paid_map.get(so.id, D0)
        total = so.net_total or D0
        remaining = total - paid
        so_rows.append(
            {
                "id": so.id,
                "date": so.created_at.strftime("%Y-%m-%d"),
                "status": getattr(so, "status", "") or "",
                "customer": getattr(so.customer, "display_name", None),
                "customer_name": getattr(so, "customer_name", "") or "",
                "items_count": len(so_items_map.get(so.id, [])),
                "items": so_items_map.get(so.id, []),
                "paid": paid,
                "remaining": remaining,
                "net_total": total,
            }
        )

    # Purchase Returns table
    pr_qs = PurchaseReturn.objects.filter(created_at__range=(dt_from, dt_to))
    if business:
        pr_qs = pr_qs.filter(business=business)
    pr_qs = pr_qs.select_related("supplier").prefetch_related("items").order_by("-created_at", "-id")

    pr_item_rows = (
        PurchaseReturnItem.objects.filter(purchase_return__in=pr_qs)
        .values("purchase_return_id", "product__name")
        .annotate(
            qty=Coalesce(Sum("quantity", output_field=DecimalField(max_digits=18, decimal_places=6)), DQ0),
            unit=Coalesce(Sum(F("unit_price"), output_field=DecimalField(max_digits=18, decimal_places=2)), D0),
            line=Coalesce(
                Sum(F("quantity") * F("unit_price"), output_field=DecimalField(max_digits=18, decimal_places=2)),
                D0,
            ),
        )
    )
    pr_items_map: dict[int, list] = {}
    for r in pr_item_rows:
        pr_items_map.setdefault(r["purchase_return_id"], []).append(
            {
                "product": r["product__name"],
                "qty": r["qty"],
                "unit_price": r["unit"],
                "line_total": r["line"],
            }
        )

    pr_refund_map = {
        row["purchase_return"]: row["amt"]
        for row in PurchaseReturnRefund.objects
        .filter(purchase_return__in=pr_qs, payment__date__range=(d_from, d_to))
        .values("purchase_return")
        .annotate(amt=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0))
    }

    pr_rows = []
    pr_count = pr_qs.count()
    pr_refunded_total = D0
    pr_remaining_total = D0

    for pr in pr_qs:
        refunded = pr_refund_map.get(pr.id, D0) or D0
        total = pr.net_total or D0
        remaining = total - refunded
        pr_refunded_total += refunded
        pr_remaining_total += remaining
        pr_rows.append(
            {
                "id": pr.id,
                "date": pr.created_at.strftime("%Y-%m-%d"),
                "status": getattr(pr, "status", "") or "",
                "supplier": getattr(pr.supplier, "display_name", None) or "",
                "items_count": len(pr_items_map.get(pr.id, [])),
                "items": pr_items_map.get(pr.id, []),
                "total": total,
                "refunded": refunded,
                "remaining": remaining,
            }
        )

    # Sales Returns table
    sr_qs = SalesReturn.objects.filter(created_at__range=(dt_from, dt_to))
    if business:
        sr_qs = sr_qs.filter(business=business)
    sr_qs = sr_qs.select_related("customer").prefetch_related("items").order_by("-created_at", "-id")

    sr_item_rows = (
        SalesReturnItem.objects
        .filter(sales_return__in=sr_qs)
        .values("sales_return_id", "product__name")
        .annotate(
            qty=Coalesce(Sum("quantity", output_field=DecimalField(max_digits=18, decimal_places=6)), DQ0),
            unit=Coalesce(Sum(F("unit_price"), output_field=DecimalField(max_digits=18, decimal_places=2)), D0),
            line=Coalesce(
                Sum(F("quantity") * F("unit_price"), output_field=DecimalField(max_digits=18, decimal_places=2)),
                D0,
            ),
        )
    )
    sr_items_map: dict[int, list] = {}
    for r in sr_item_rows:
        sr_items_map.setdefault(r["sales_return_id"], []).append(
            {
                "product": r["product__name"],
                "qty": r["qty"],
                "unit_price": r["unit"],
                "line_total": r["line"],
            }
        )

    sr_rows = []
    sr_count = sr_qs.count()
    for sr in sr_qs:
        total = sr.net_total or D0
        sr_rows.append(
            {
                "id": sr.id,
                "date": sr.created_at.strftime("%Y-%m-%d"),
                "status": getattr(sr, "status", "") or "",
                "customer": getattr(sr.customer, "display_name", None) or (getattr(sr, "customer_name", "") or ""),
                "items_count": len(sr_items_map.get(sr.id, [])),
                "items": sr_items_map.get(sr.id, []),
                "total": total,
            }
        )

    # SIMPLIFIED day series (removed profit calculations)
    days = []
    series_revenue, series_expense = [], []
    for d in _daterange_days(d_from, d_to):
        r = Decimal(rev_map.get(d, 0) or 0)
        e = Decimal(expense_by_day_map.get(d, 0) or 0)
        days.append(d.strftime("%Y-%m-%d"))
        series_revenue.append(float(r))
        series_expense.append(float(e))

    # SIMPLIFIED Monthly trend (removed profit calculations)
    month_labels, m_starts, m_ends = _month_labels(12)
    trend_revenue, trend_expense = [], []
    for ms, me in zip(m_starts, m_ends):
        ms_dt = make_aware_safe(datetime(ms.year, ms.month, ms.day, 0, 0, 0))
        me_dt = make_aware_safe(datetime(me.year, me.month, me.day, 23, 59, 59))

        so_m = SalesOrder.objects.filter(created_at__range=(ms_dt, me_dt)).exclude(status="cancelled")
        if business:
            so_m = so_m.filter(business=business)

        rev_m = so_m.aggregate(
            s=Coalesce(Sum("net_total", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
        )["s"] or D0

        exp_m = Expense.objects.filter(date__range=(ms, me))
        if business:
            exp_m = exp_m.filter(business=business)
        exp_m_total = exp_m.aggregate(
            s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
        )["s"] or D0

        trend_revenue.append(float(rev_m))
        trend_expense.append(float(exp_m_total))

    # Sales by category
    order_ids = list(orders_qs.values_list("id", flat=True))
    cat_rows = (
        SalesOrderItem.objects
        .filter(sales_order_id__in=order_ids)
        .values("product__category__id", "product__category__name")
        .annotate(
            qty=Coalesce(Sum("quantity", output_field=DecimalField(max_digits=18, decimal_places=6)), DQ0),
            amt=Coalesce(
                Sum(F("quantity") * F("unit_price"), output_field=DecimalField(max_digits=18, decimal_places=2)),
                D0,
            ),
        )
        .order_by("-amt")
    )
    sales_cat_labels, sales_cat_values, sales_cat_amount_values = [], [], []
    for r in cat_rows:
        label = r["product__category__name"] or "Uncategorized"
        sales_cat_labels.append(label)
        sales_cat_values.append(float(r["qty"] or 0))
        sales_cat_amount_values.append(float(r["amt"] or 0))

    # Expenses by category
    exp_cat = (
        Expense.objects.filter(exp_filter)
        .values("category")
        .annotate(amt=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0))
        .order_by("-amt")
    )
    exp_cat_labels, exp_cat_values = [], []
    label_map = dict(ExpenseCategory.choices)
    for r in exp_cat:
        code = r["category"]
        exp_cat_labels.append(label_map.get(code, code or "Other"))
        exp_cat_values.append(float(r["amt"] or 0))

    # Items sold list
    items_sold_qs = (
        SalesOrderItem.objects
        .filter(sales_order_id__in=order_ids)
        .values("product__name")
        .annotate(qty=Coalesce(Sum("quantity", output_field=DecimalField(max_digits=18, decimal_places=6)), DQ0))
        .order_by("-qty", "product__name")
    )
    items_sold = [{"name": r["product__name"], "qty": r["qty"]} for r in items_sold_qs]

    # ---------------------------------------------------------------------
    # NEW: PROFIT & LOSS CALCULATIONS (Landed Cost based)
    # ---------------------------------------------------------------------
    
    # Calculate COGS from SalesOrderItem snapshots
    # Sum of (unit_cost * quantity) for all non-cancelled sales in period
    cogs_total = SalesOrderItem.objects.filter(
        sales_order_id__in=order_ids
    ).aggregate(
        s=Coalesce(Sum(F("quantity") * F("unit_cost"), output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0
    
    # Gross Profit = Revenue - COGS
    gross_profit = revenue_total - cogs_total
    
    # Net Profit = Gross Profit - Non-PO Operating Expenses
    # (PO-linked expenses are already in COGS via landed cost)
    operating_expenses_total = Expense.objects.filter(
        exp_filter & Q(purchase_order__isnull=True)
    ).aggregate(
        s=Coalesce(Sum("amount", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0
    
    net_profit = gross_profit - operating_expenses_total
    
    # Product-wise Profit Breakdown
    product_profit_rows = []
    product_profit_qs = (
        SalesOrderItem.objects
        .filter(sales_order_id__in=order_ids)
        .values("product__id", "product__name")
        .annotate(
            total_qty=Coalesce(Sum("quantity", output_field=DecimalField(max_digits=18, decimal_places=6)), DQ0),
            total_rev=Coalesce(Sum(F("quantity") * F("unit_price"), output_field=DecimalField(max_digits=18, decimal_places=2)), D0),
            total_cost=Coalesce(Sum(F("quantity") * F("unit_cost"), output_field=DecimalField(max_digits=18, decimal_places=2)), D0),
        )
    )
    
    for r in product_profit_qs:
        profit = r["total_rev"] - r["total_cost"]
        margin = (profit / r["total_rev"] * 100) if r["total_rev"] > 0 else 0
        product_profit_rows.append({
            "name": r["product__name"],
            "qty": r["total_qty"],
            "revenue": r["total_rev"],
            "cost": r["total_cost"],
            "profit": profit,
            "margin": margin,
        })
    product_profit_rows.sort(key=lambda x: x["profit"], reverse=True)
    
    # Cash Sale Profit
    # Profit from orders that were paid (even partially) via Cash
    cash_order_ids = set(Payment.objects.filter(
        direction=Payment.IN,
        payment_source=Payment.CASH,
        applied_sales_orders__id__in=order_ids
    ).values_list("applied_sales_orders__id", flat=True))
    
    cash_sales_revenue = SalesOrder.objects.filter(id__in=cash_order_ids).aggregate(
        s=Coalesce(Sum("net_total", output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0
    
    cash_sales_cogs = SalesOrderItem.objects.filter(
        sales_order_id__in=cash_order_ids
    ).aggregate(
        s=Coalesce(Sum(F("quantity") * F("unit_cost"), output_field=DecimalField(max_digits=18, decimal_places=2)), D0)
    )["s"] or D0
    
    cash_sale_profit = cash_sales_revenue - cash_sales_cogs

    # Context
    businesses = Business.objects.order_by("name", "id")

    context = {
        "from": dt_from.strftime("%Y-%m-%dT%H:%M"),
        "to": dt_to.strftime("%Y-%m-%dT%H:%M"),
        "businesses": businesses,
        "business": business,
        "mode": mode,

        # NEW PROFIT CARDS
        "kpi_revenue": revenue_total,
        "kpi_cogs": cogs_total,
        "kpi_gross_profit": gross_profit,
        "kpi_operating_expenses": operating_expenses_total,
        "kpi_net_profit": net_profit,
        "product_profit_rows": product_profit_rows,
        "kpi_cash_sale_profit": cash_sale_profit,

        # NEW CASH CARDS
        "kpi_sales_cash": kpi_sales_cash,  # Cash from sales
        "kpi_receipts_cash": kpi_receipts_cash,  # Cash from other receipts
        "kpi_cash_in_hand": kpi_cash_in_hand,  # Total cash balance
        
        # NEW: Cash Out breakdown
        "kpi_cash_out_po": kpi_cash_out_po,
        "kpi_cash_out_sr_refund": kpi_cash_out_sr_refund,
        "kpi_cash_out_general": kpi_cash_out_general,
        # Cash out via Sales Return Refunds
        
        # Revenue KPI
        "kpi_revenue": revenue_total or D0,
        "kpi_cancelled": cancelled_total or D0,
        
        # Receipt number information
        "receipt_count": receipt_count,
        "receipt_series": receipt_series or "No receipts",
        "cancelled_receipt_count": cancelled_receipt_count,
        "cancelled_receipt_numbers": cancelled_receipt_numbers,

        # Expense KPI
        "kpi_expenses": expense_total or D0,
        "kpi_landed_po_expenses": landed_po_expenses_total or D0,
        "kpi_operating_expenses": operating_expenses_total or D0,

        # Bank and cheque KPIs
        "kpi_cheque_in_hand": kpi_cheque_in_hand,
        "kpi_cheque_in_hand_pending": kpi_cheque_in_hand_pending,
        "kpi_cheque_in_hand_deposited": kpi_cheque_in_hand_deposited,
        "kpi_bank_revenue": kpi_bank_revenue,
        "kpi_bank_deposited": kpi_bank_deposited,
        "kpi_bank_amount": kpi_bank_amount,
        "kpi_remaining": kpi_remaining,
        
        # NEW: Bank summaries per account
        "bank_summaries": bank_summaries,
        "grand_total_banks": grand_total_banks,

        # Cash out total (for reference)
        "kpi_cash_out_total": cash_out_total,

        # Amount In card data
        "amount_in_cash": amount_in_cash,
        "amount_in_cheque_deposited": amount_in_cheque_deposited,
        "amount_in_bank_cash": amount_in_bank_cash,
        "amount_in_total": amount_in_total,

        # deposit table rows
        "deposit_bank_rows": deposit_bank_rows,
        "deposit_cheque_rows": deposit_cheque_rows,

        # SIMPLIFIED day series (removed profit/cost)
        "days": json.dumps(days),
        "series_revenue": json.dumps(series_revenue),
        "series_expense": json.dumps(series_expense),

        # SIMPLIFIED monthly trend (removed profit)
        "months": json.dumps(month_labels),
        "trend_revenue": json.dumps(trend_revenue),
        "trend_expense": json.dumps(trend_expense),

        # categories
        "sales_cat_labels": json.dumps(sales_cat_labels),
        "sales_cat_values": json.dumps(sales_cat_values),
        "sales_cat_amount_values": json.dumps(sales_cat_amount_values),
        "exp_cat_labels": json.dumps(exp_cat_labels),
        "exp_cat_values": json.dumps(exp_cat_values),

        # bank tables
        "bank_in_rows": bank_in_rows,
        "bank_out_rows": bank_out_rows,
        "bank_in_total": bank_in_total,
        "bank_out_total": bank_out_total,

        # purchases
        "po_rows": po_rows,
        "po_count": po_count,
        "po_paid_total": po_paid_total,
        "po_remaining_total": po_remaining_total,

        # purchase returns
        "pr_rows": pr_rows,
        "pr_count": pr_count,
        "pr_refunded_total": pr_refunded_total,
        "pr_remaining_total": pr_remaining_total,

        # sales
        "so_rows": so_rows,
        "so_count": so_count,

        # sales returns
        "sr_rows": sr_rows,
        "sr_count": sr_count,

        # items sold
        "items_sold": items_sold,

        # expenses table
        "expenses_total": expense_total or D0,
        "expenses_rows": [
            {
                "date": e["date"].strftime("%Y-%m-%d"),
                "category": label_map.get(e["category"], e["category"]),
                "note": e["description"] or "",
                "amount": e["amount"],
            }
            for e in Expense.objects.filter(exp_filter)
            .values("date", "category", "description", "amount")
            .order_by("date", "id")
        ],
    }
    return render(request, "barkat/finance/reports.html", context)



# ============================
#       Warehouse
# ============================


def _with_wh_stock(queryset, warehouse):
    subq = WarehouseStock.objects.filter(
        warehouse=warehouse, product=OuterRef("pk")
    ).values("quantity")[:1]
    return queryset.annotate(
        stock_qty=Coalesce(Subquery(subq), Value(Decimal("0"), output_field=DecimalField(max_digits=18, decimal_places=6)))
    )

@login_required
def warehouse_list(request):
    """
    List of warehouses with search + pagination.
    """
    q = request.GET.get("q", "").strip()
    qs = Warehouse.objects.all().order_by("name", "id")

    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q) | Q(address__icontains=q))

    paginator = Paginator(qs, 12)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    ctx = {
        "page_obj": page_obj,
        "query": q,
    }
    return render(request, "barkat/inventory/warehouse_list.html", ctx)

@login_required
def warehouse_create(request):
    """
    Create a new warehouse.
    """
    if request.method == "POST":
        form = WarehouseForm(request.POST)
        if form.is_valid():
            obj = form.save(user=request.user)  # uses AuditUserMixin from forms.py
            messages.success(request, f"Warehouse “{obj.name}” created.")
            return redirect("warehouse_list")
    else:
        form = WarehouseForm()

    return render(request, "barkat/inventory/warehouse_form.html", {"form": form, "mode": "create"})

@login_required
def warehouse_update(request, pk: int):
    """
    Edit an existing warehouse.
    """
    wh = get_object_or_404(Warehouse, pk=pk)
    if request.method == "POST":
        form = WarehouseForm(request.POST, instance=wh)
        if form.is_valid():
            obj = form.save(user=request.user)
            messages.success(request, f"Warehouse “{obj.name}” updated.")
            return redirect("warehouse_list")
    else:
        form = WarehouseForm(instance=wh)

    return render(request, "barkat/inventory/warehouse_form.html", {"form": form, "mode": "edit", "warehouse": wh})

def _with_wh_stock(queryset, warehouse):
    subq = WarehouseStock.objects.filter(
        warehouse=warehouse, product=OuterRef("pk")
    ).values("quantity")[:1]
    # IMPORTANT: use a different name than any Product field
    return queryset.annotate(
        wh_qty=Coalesce(
            Subquery(subq),
            Value(Decimal("0"), output_field=DecimalField(max_digits=18, decimal_places=6)),
        )
    )

@login_required
def warehouse_detail(request, pk: int):
    wh = get_object_or_404(Warehouse, pk=pk)
    businesses = Business.objects.order_by("name", "id")

    q = (request.GET.get("q") or "").strip()
    products = Product.objects.select_related("business", "uom")
    if q:
        products = products.filter(
            Q(name__icontains=q) | Q(sku__icontains=q) | Q(business__name__icontains=q)
        )

    products = _with_wh_stock(products, wh).order_by("business__name", "name", "id")

    page_obj = Paginator(products, 25).get_page(request.GET.get("page"))
    return render(request, "barkat/inventory/business_warehouse.html", {
        "warehouse": wh,
        "businesses": businesses,
        "page_obj": page_obj,
        "query": q,
        "active_business": None,
    })

@login_required
def business_wise_warehouse(request, pk: int, business_id: int):
    wh = get_object_or_404(Warehouse, pk=pk)
    business = get_object_or_404(Business, pk=business_id)
    businesses = Business.objects.order_by("name", "id")

    q = (request.GET.get("q") or "").strip()
    products = Product.objects.select_related("business", "uom").filter(business=business)
    if q:
        products = products.filter(Q(name__icontains=q) | Q(sku__icontains=q))

    products = _with_wh_stock(products, wh).order_by("name", "id")

    page_obj = Paginator(products, 25).get_page(request.GET.get("page"))
    return render(request, "barkat/inventory/business_wise_warehouse.html", {
        "warehouse": wh,
        "business": business,
        "businesses": businesses,
        "page_obj": page_obj,
        "query": q,
        "active_business": business,
    })
    """
    Shows ONLY products of a particular business, with stock in this warehouse.
    Same header actions and the same tab styling as the all-products page.
    """
    wh = get_object_or_404(Warehouse, pk=pk)
    business = get_object_or_404(Business, pk=business_id)
    businesses = Business.objects.order_by("name", "id")

    q = (request.GET.get("q") or "").strip()
    products = Product.objects.select_related("business", "uom").filter(business=business)
    if q:
        products = products.filter(Q(name__icontains=q) | Q(sku__icontains=q))

    products = _with_wh_stock(products, wh).order_by("name", "id")

    paginator = Paginator(products, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    ctx = {
        "warehouse": wh,
        "business": business,
        "businesses": businesses,
        "page_obj": page_obj,
        "query": q,
        "active_business": business,  # template highlights active tab
    }
    return render(request, "barkat/inventory/business_wise_warehouse.html", ctx)

def _with_wh_stock(queryset, warehouse):
    subq = WarehouseStock.objects.filter(
        warehouse=warehouse, product=OuterRef("pk")
    ).values("quantity")[:1]
    return queryset.annotate(
        wh_qty=Coalesce(
            Subquery(subq),
            Value(Decimal("0"), output_field=DecimalField(max_digits=18, decimal_places=6)),
        )
    )

@login_required
@require_http_methods(["GET", "POST"])
def warehouse_refill(request, pk: int):
    """
    Refill quantities for one warehouse and one business.
    - GET requires ?business=<id>
    - Table lists products for that business with current qty in this warehouse.
    - POST only adds delta amounts typed in "add_<product_id>" fields.
    """
    warehouse = get_object_or_404(Warehouse, pk=pk)

    # business can arrive via GET (initial) or POST (hidden field)
    business_id = request.GET.get("business") or request.POST.get("business")
    if not business_id:
        # If no business specified, show a simple picker
        businesses = Business.objects.order_by("name", "id")
        return render(request, "barkat/inventory/warehouse_refill_pick_business.html", {
            "warehouse": warehouse,
            "businesses": businesses,
        })

    business = get_object_or_404(Business, pk=business_id)

    # search/filter
    q = (request.GET.get("q") or request.POST.get("q") or "").strip()
    products = Product.objects.select_related("uom").filter(business=business)
    if q:
        products = products.filter(Q(name__icontains=q) | Q(sku__icontains=q))

    products = _with_wh_stock(products, warehouse).order_by("name", "id")

    if request.method == "POST":
        updated_rows = 0
        total_added = Decimal("0")
        with transaction.atomic():
            # Lock/adjust per product
            for p in products:
                field = f"add_{p.id}"
                raw = request.POST.get(field, "").strip()
                if not raw:
                    continue
                try:
                    delta = Decimal(raw)
                except InvalidOperation:
                    messages.error(request, f"Invalid number for {p.name}: {raw}")
                    continue
                if delta <= 0:
                    continue

                ws, _ = WarehouseStock.objects.select_for_update().get_or_create(
                    warehouse=warehouse, product=p, defaults={"quantity": Decimal("0")}
                )
                ws.quantity = (ws.quantity or Decimal("0")) + delta
                ws.full_clean()
                ws.save(update_fields=["quantity", "updated_at"])

                updated_rows += 1
                total_added += delta

        if updated_rows:
            messages.success(
                request,
                f"Added {total_added} units across {updated_rows} product(s) for {business.name} in {warehouse.code}."
            )
        else:
            messages.info(request, "No quantities entered.")

        # stay on same page
        return redirect(f"{request.path}?business={business.id}")

    # GET
    page_obj = Paginator(products, 50).get_page(request.GET.get("page"))

    return render(request, "barkat/inventory/warehouse_refill.html", {
        "warehouse": warehouse,
        "business": business,
        "page_obj": page_obj,
        "query": q,
    })

def _current_qty_in_wh(product, warehouse) -> Decimal:
    if not (product and warehouse):
        return Decimal("0")
    row = WarehouseStock.objects.filter(warehouse=warehouse, product=product).values("quantity").first()
    return (row or {}).get("quantity") or Decimal("0")

@login_required
@require_http_methods(["GET", "POST"])
def stock_move_create(request):
    """
    Create and post a stock move:
      - Source: warehouse (required)
      - Destination: warehouse OR business
      - Shows current quantity available in source warehouse for the chosen product.
    """
    initial = {}

    # Allow prefill with query string (e.g., from buttons)
    if "source_warehouse" in request.GET:
        initial["source_warehouse"] = request.GET.get("source_warehouse")
    if "dest_warehouse" in request.GET:
        initial["destination_type"] = "warehouse"
        initial["dest_warehouse"] = request.GET.get("dest_warehouse")
    if "business" in request.GET:
        initial["destination_type"] = "business"
        initial["dest_business"] = request.GET.get("business")

    if request.method == "POST":
        form = StockMoveCreateForm(request.POST)
        if form.is_valid():
            try:
                move = form.create_move(user=request.user)
            except ValidationError as e:
                form.add_error(None, e.message if hasattr(e, "message") else e.messages)
            else:
                # After success, if a warehouse is involved as source we can go back there.
                src_wh = form.cleaned_data.get("source_warehouse")
                messages.success(request, f"Stock moved successfully (Move #{move.pk}).")
                if src_wh:
                    return redirect("warehouse_detail", pk=src_wh.pk)
                return redirect("warehouse_list")
    else:
        form = StockMoveCreateForm(initial=initial)

    # Compute current qty for display
    product = None
    src_wh = None
    current_qty = Decimal("0")

    try:
        product_id = request.POST.get("product") or request.GET.get("product")
        if product_id:
            product = get_object_or_404(Product, pk=product_id)
    except Exception:
        product = None

    try:
        src_id = request.POST.get("source_warehouse") or request.GET.get("source_warehouse") or initial.get("source_warehouse")
        if src_id:
            src_wh = get_object_or_404(Warehouse, pk=src_id)
    except Exception:
        src_wh = None

    if product and src_wh:
        current_qty = _current_qty_in_wh(product, src_wh)

    context = {
        "form": form,
        "product": product,
        "source_wh": src_wh,
        "current_qty": current_qty,
    }
    return render(request, "barkat/inventory/stock_move_form.html", context)

def _annotate_wh_qty(qs, warehouse: Warehouse):
    """
    Annotate Product queryset with wh_qty (current qty in given warehouse).
    """
    subq = (WarehouseStock.objects
            .filter(warehouse=warehouse, product=OuterRef("pk"))
            .values("quantity")[:1])
    return qs.annotate(
        wh_qty=Coalesce(
            Subquery(subq),
            Value(Decimal("0"), output_field=DecimalField(max_digits=18, decimal_places=6))
        )
    )

@login_required
def stock_move_bulk(request):
    # Active context
    business_id = request.GET.get("business") or request.POST.get("business")
    source_wh_id = request.GET.get("source_warehouse") or request.POST.get("source_warehouse")
    dest_type = request.GET.get("dest_type") or request.POST.get("dest_type") or "warehouse"

    business = get_object_or_404(Business, pk=business_id) if business_id else None
    source_wh = get_object_or_404(Warehouse, pk=source_wh_id)

    # Listing data (unchanged idea)
    warehouses = Warehouse.objects.filter(is_active=True).order_by("name")
    businesses = Business.objects.filter(is_active=True).order_by("name")

    # Products scoped to selected business
    products = Product.objects.none()
    if business:
        products = (
            Product.objects.filter(business=business, is_active=True, is_deleted=False)
            .select_related("uom")
            .order_by("name")
        )

    # Handle submit
    if request.method == "POST":
        dest_type = request.POST.get("dest_type", "warehouse")
        reference = (request.POST.get("reference") or "").strip()

        dest_wh = None
        dest_biz = None

        if dest_type == "business":
            # Auto-lock destination business to the active tab
            if not business:
                messages.error(request, "Please select a business from the tabs.")
                return redirect(f"{request.path}?business={business_id}&source_warehouse={source_wh_id}&dest_type=business")
            dest_biz = business
        else:
            dest_wh_id = request.POST.get("dest_warehouse")
            if not dest_wh_id:
                messages.error(request, "Please select a destination warehouse.")
                return redirect(f"{request.path}?business={business_id}&source_warehouse={source_wh_id}&dest_type=warehouse")
            dest_wh = get_object_or_404(Warehouse, pk=dest_wh_id)
            if dest_wh.id == source_wh.id:
                messages.error(request, "Source and destination warehouse cannot be the same.")
                return redirect(f"{request.path}?business={business_id}&source_warehouse={source_wh_id}&dest_type=warehouse")

        # Collect quantities
        qty_entries = []
        for p in products:
            raw = request.POST.get(f"qty_{p.id}", "").strip()
            if not raw:
                continue
            try:
                q = Decimal(raw)
            except Exception:
                messages.error(request, f"Invalid quantity for {p.name}.")
                return redirect(f"{request.path}?business={business_id}&source_warehouse={source_wh_id}&dest_type={dest_type}")
            if q <= 0:
                continue
            qty_entries.append((p, q))

        if not qty_entries:
            messages.error(request, "Enter at least one positive quantity.")
            return redirect(f"{request.path}?business={business_id}&source_warehouse={source_wh_id}&dest_type={dest_type}")

        # Validate source balances once, then post
        with transaction.atomic():
            # Build or get all source rows to check availability first
            src_rows = {
                p.id: WarehouseStock.objects.select_for_update().get_or_create(
                    warehouse=source_wh, product=p, defaults={"quantity": Decimal("0")}
                )[0]
                for p, _ in qty_entries
            }
            # Check insufficiency upfront
            for p, q in qty_entries:
                if src_rows[p.id].quantity < q:
                    messages.error(request, f"Insufficient stock of {p.name} in {source_wh.code}.")
                    return redirect(f"{request.path}?business={business_id}&source_warehouse={source_wh_id}&dest_type={dest_type}")

            # Post moves
            for p, q in qty_entries:
                mv = StockMove(
                    product=p,
                    source_warehouse=source_wh,
                    dest_warehouse=dest_wh if dest_type == "warehouse" else None,
                    dest_business=dest_biz if dest_type == "business" else None,
                    quantity=q,
                    reference=reference,
                    status=StockMove.Status.DRAFT,
                    created_by=getattr(request, "user", None),
                    updated_by=getattr(request, "user", None),
                )
                mv.post(user=getattr(request, "user", None))

        messages.success(request, "Stock moved successfully.")
        return redirect("warehouse_detail", pk=source_wh.id)

    # GET
    query = (request.GET.get("q") or "").strip()
    if business:
        qs = products
        if query:
            qs = qs.filter(models.Q(name__icontains=query) | models.Q(sku__icontains=query))
        # attach current WH qty
        wh_map = {
            r["product_id"]: r["qty"]
            for r in WarehouseStock.objects.filter(warehouse=source_wh, product__in=qs).values("product_id").annotate(qty=models.Sum("quantity"))
        }
        products = [
            # attach wh_qty attribute for display
            (lambda p: (setattr(p, "wh_qty", wh_map.get(p.id, Decimal("0"))), p)[1])(p)
            for p in qs
        ]

    ctx = {
        "business": business,
        "businesses": businesses,
        "source_wh": source_wh,
        "dest_type": dest_type,
        "warehouses": warehouses,
        "products": products,
        "query": request.GET.get("q", ""),
    }
    return render(request, "barkat/inventory/stock_move_bulk.html", ctx)


# ---------- Stock status (ALL) from Product.stock_qty ----------

def _stock_status_data(request) -> Tuple[Any, Optional[date], Optional[date], str, Any]:
    """Shared data for stock_status HTML and Excel export. Returns (rows, date_from, date_to, q, businesses)."""
    businesses = Business.objects.filter(is_active=True, is_deleted=False).order_by("name")
    q = (request.GET.get("q") or "").strip()
    today = timezone.localdate()
    date_from_raw = (request.GET.get("date_from") or "").strip()
    date_to_raw = (request.GET.get("date_to") or "").strip()
    date_from = None
    date_to = None
    if date_from_raw:
        try:
            date_from = datetime.strptime(date_from_raw, "%Y-%m-%d").date()
        except ValueError:
            date_from = None
    if date_to_raw:
        try:
            date_to = datetime.strptime(date_to_raw, "%Y-%m-%d").date()
        except ValueError:
            date_to = None
    if not date_from and not date_to:
        first_of_month = today.replace(day=1)
        if today.month == 12:
            next_month = date(today.year + 1, 1, 1)
        else:
            next_month = date(today.year, today.month + 1, 1)
        last_of_month = next_month - timedelta(days=1)
        date_from = first_of_month
        date_to = last_of_month
    po_date_filters = {}
    if date_from:
        po_date_filters["purchase_order__created_at__date__gte"] = date_from
    if date_to:
        po_date_filters["purchase_order__created_at__date__lte"] = date_to
    po_items_qs = PurchaseOrderItem.objects.filter(
        product=OuterRef("pk"),
        purchase_order__is_active=True,
        purchase_order__is_deleted=False,
        **po_date_filters,
    )
    po_avg_subquery = (
        po_items_qs
        .values("product")
        .annotate(
            total_cost=Sum(
                ExpressionWrapper(
                    F("quantity") * F("unit_price"),
                    output_field=DecimalField(max_digits=18, decimal_places=6),
                )
            ),
            total_qty=Sum("quantity"),
        )
        .annotate(
            avg_price=Case(
                When(
                    total_qty__gt=0,
                    then=ExpressionWrapper(
                        F("total_cost") / F("total_qty"),
                        output_field=DecimalField(max_digits=12, decimal_places=4),
                    ),
                ),
                default=Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=4),
            )
        )
        .values("avg_price")[:1]
    )
    rows = (
        Product.objects
        .select_related("business", "uom", "bulk_uom")
        .filter(is_active=True, is_deleted=False)
        .annotate(
            avg_purchase_rate=Subquery(
                po_avg_subquery,
                output_field=DecimalField(max_digits=12, decimal_places=4),
            )
        )
        .order_by("business__name", "name")
    )
    if q:
        rows = rows.filter(
            models.Q(name__icontains=q)
            | models.Q(company_name__icontains=q)
            | models.Q(business__name__icontains=q)
        )
    return (rows, date_from, date_to, q, businesses)


@login_required
def stock_status(request):
    rows, date_from, date_to, q, businesses = _stock_status_data(request)
    ctx = {
        "businesses": businesses,
        "rows": rows,
        "query": q,
        "date_from": date_from,
        "date_to": date_to,
    }
    return render(request, "barkat/inventory/stock_status.html", ctx)


@login_required
def stock_status_excel(request):
    """Export stock status as Excel (.xlsx). Uses same filters as stock_status (q, date_from, date_to)."""
    rows, date_from, date_to, q, _ = _stock_status_data(request)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        return HttpResponse("Excel export requires openpyxl. Install with: pip install openpyxl", status=500)
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Status"
    headers = ["Business", "Product", "Company", "Base Unit", "Stock Qty", "Bulk (if any)", "Avg purchase rate"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_num = 2
    for p in rows:
        ws.cell(row=row_num, column=1, value=getattr(p.business, "name", "") or "")
        ws.cell(row=row_num, column=2, value=getattr(p, "name", "") or "")
        ws.cell(row=row_num, column=3, value=getattr(p, "company_name", "") or "")
        uom = getattr(p, "uom", None)
        ws.cell(row=row_num, column=4, value=(getattr(uom, "symbol", None) or getattr(uom, "code", None) or "") if uom else "")
        stock = getattr(p, "stock_qty", None)
        ws.cell(row=row_num, column=5, value=float(stock) if stock is not None else 0)
        bulk = getattr(p, "bulk_stock_status", None) or ""
        ws.cell(row=row_num, column=6, value=bulk)
        avg = getattr(p, "avg_purchase_rate", None)
        ws.cell(row=row_num, column=7, value=float(avg) if avg is not None else None)
        row_num += 1
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"stock_status_{timezone.localdate().strftime('%Y-%m-%d')}.xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _item_base_qty_and_unit(item, product, has_uom_size):
    """Return (base_qty, display_qty, display_unit) for a line item.
    has_uom_size: True for PO/PR/SO items (uom, size_per_unit); False for SI/SR.
    """
    qty = item.quantity or Decimal("0")
    if has_uom_size:
        size = getattr(item, "size_per_unit", None) or Decimal("1")
        base = qty * size
    else:
        base = qty
    uom = None
    if has_uom_size and getattr(item, "uom_id", None) and getattr(item, "uom", None):
        uom = item.uom
    if not uom and product and getattr(product, "uom", None):
        uom = product.uom
    unit_code = (getattr(uom, "code", None) or getattr(uom, "symbol", None) or "") if uom else ""
    if not unit_code and product and getattr(product, "uom", None):
        unit_code = getattr(product.uom, "code", None) or getattr(product.uom, "symbol", None) or "—"
    if not unit_code:
        unit_code = "—"
    return (base, qty, unit_code)


@login_required
def product_stock_detail(request, product_id):
    # product with basic info
    product = get_object_or_404(
        Product.objects.select_related("business", "uom", "bulk_uom"),
        pk=product_id,
        is_active=True,
        is_deleted=False,
    )

    # date filters
    start_raw = (request.GET.get("start") or "").strip()
    end_raw = (request.GET.get("end") or "").strip()
    start_date = None
    end_date = None

    try:
        if start_raw:
            start_date = datetime.strptime(start_raw, "%Y-%m-%d").date()
    except ValueError:
        start_date = None

    try:
        if end_raw:
            end_date = datetime.strptime(end_raw, "%Y-%m-%d").date()
    except ValueError:
        end_date = None

    def include_dt(dt):
        """
        Accepts both datetime and date, and applies start/end filter.
        """
        if dt is None:
            return True

        if isinstance(dt, datetime):
            d = dt.date()
        else:
            d = dt

        if start_date and d < start_date:
            return False
        if end_date and d > end_date:
            return False
        return True

    movements = []

    # 1. Purchase Orders. STOCK IN. vendor name
    po_items = (
        PurchaseOrderItem.objects
        .select_related("purchase_order", "purchase_order__supplier", "uom")
        .filter(
            product_id=product.id,
            purchase_order__status="received",
            purchase_order__is_active=True,
            purchase_order__is_deleted=False,
        )
    )
    for it in po_items:
        po = it.purchase_order
        if not include_dt(po.created_at):
            continue
        base, disp_qty, unit_code = _item_base_qty_and_unit(it, product, True)
        vendor_name = getattr(po.supplier, "display_name", None) or "Vendor"
        movements.append({
            "date": po.created_at,
            "type": "Purchase",
            "party": vendor_name,
            "qty_in": base,
            "qty_out": Decimal("0"),
            "display_qty_in": disp_qty,
            "display_unit_in": unit_code,
            "display_qty_out": None,
            "display_unit_out": None,
        })

    # 2. Purchase Returns. STOCK OUT. vendor name
    pr_items = (
        PurchaseReturnItem.objects
        .select_related("purchase_return", "purchase_return__supplier", "uom")
        .filter(
            product_id=product.id,
            purchase_return__status="processed",
            purchase_return__is_active=True,
            purchase_return__is_deleted=False,
        )
    )
    for it in pr_items:
        pr = it.purchase_return
        if not include_dt(pr.created_at):
            continue
        base, disp_qty, unit_code = _item_base_qty_and_unit(it, product, True)
        vendor_name = getattr(pr.supplier, "display_name", None) or "Vendor"
        movements.append({
            "date": pr.created_at,
            "type": "Purchase return",
            "party": vendor_name,
            "qty_in": Decimal("0"),
            "qty_out": base,
            "display_qty_in": None,
            "display_unit_in": None,
            "display_qty_out": disp_qty,
            "display_unit_out": unit_code,
        })

    # 3. Sales Invoices. STOCK OUT. customer name (no uom/size on item; base unit only)
    si_items = (
        SalesInvoiceItem.objects
        .select_related("sales_invoice", "sales_invoice__customer")
        .filter(
            product_id=product.id,
            # optionally filter by posted status
            # sales_invoice__status=SalesInvoice.Status.POSTED,
        )
    )
    for it in si_items:
        inv = it.sales_invoice
        if not include_dt(inv.created_at):
            continue
        base, disp_qty, unit_code = _item_base_qty_and_unit(it, product, False)
        if inv.customer_id and getattr(inv, "customer", None):
            customer_name = inv.customer.display_name or "Customer"
        elif inv.customer_name:
            customer_name = inv.customer_name
        else:
            customer_name = "Customer"
        movements.append({
            "date": inv.created_at,
            "type": "Sale (Invoice)",
            "party": customer_name,
            "qty_in": Decimal("0"),
            "qty_out": base,
            "display_qty_in": None,
            "display_unit_in": None,
            "display_qty_out": disp_qty,
            "display_unit_out": unit_code,
        })

    # 4. Sales Orders. STOCK OUT. customer name
    so_items = (
        SalesOrderItem.objects
        .select_related("sales_order", "sales_order__customer", "uom")
        .filter(
            product_id=product.id,
            # add status filters if needed
            # sales_order__status=SalesOrder.Status.FULFILLED,
        )
    )
    for it in so_items:
        so = it.sales_order
        if hasattr(SalesOrder, "Status") and hasattr(SalesOrder.Status, "CANCELLED"):
            if so.status == SalesOrder.Status.CANCELLED:
                continue
        if not include_dt(so.created_at):
            continue
        base, disp_qty, unit_code = _item_base_qty_and_unit(it, product, True)
        if so.customer_id and getattr(so, "customer", None):
            customer_name = so.customer.display_name or "Customer"
        elif so.customer_name:
            customer_name = so.customer_name
        else:
            customer_name = "Customer"
        movements.append({
            "date": so.created_at,
            "type": "Sale (Order)",
            "party": customer_name,
            "qty_in": Decimal("0"),
            "qty_out": base,
            "display_qty_in": None,
            "display_unit_in": None,
            "display_qty_out": disp_qty,
            "display_unit_out": unit_code,
        })

    # 5. Sales Returns. STOCK IN. customer name (no uom/size on item; base unit only)
    sr_items = (
        SalesReturnItem.objects
        .select_related("sales_return", "sales_return__customer")
        .filter(
            product_id=product.id,
            # if you only want processed returns, filter by status
            # sales_return__status=SalesReturn.Status.PROCESSED,
        )
    )
    for it in sr_items:
        sr = it.sales_return
        if not include_dt(sr.created_at):
            continue
        base, disp_qty, unit_code = _item_base_qty_and_unit(it, product, False)
        if sr.customer_id and getattr(sr, "customer", None):
            customer_name = sr.customer.display_name or "Customer"
        elif sr.customer_name:
            customer_name = sr.customer_name
        else:
            customer_name = "Customer"
        movements.append({
            "date": sr.created_at,
            "type": "Sales return",
            "party": customer_name,
            "qty_in": base,
            "qty_out": Decimal("0"),
            "display_qty_in": disp_qty,
            "display_unit_in": unit_code,
            "display_qty_out": None,
            "display_unit_out": None,
        })

    # sort by date then type
    movements.sort(key=lambda m: (m["date"], m["type"]))

    # running balance for the ledger
    balance = Decimal("0")
    total_in = Decimal("0")
    total_out = Decimal("0")
    for m in movements:
        qty_in = m["qty_in"] or Decimal("0")
        qty_out = m["qty_out"] or Decimal("0")
        total_in += qty_in
        total_out += qty_out
        balance += qty_in - qty_out
        m["balance"] = balance

    ctx = {
        "product": product,
        "movements": movements,
        "total_in": total_in,
        "total_out": total_out,
        "start": start_raw,
        "end": end_raw,
    }
    return render(request, "barkat/inventory/product_stock_detail.html", ctx)

# ---------- Business stock status from Product.stock_qty ----------
@login_required
def business_stock_status(request, business_id: int):
    business = get_object_or_404(Business, pk=business_id, is_active=True, is_deleted=False)
    businesses = Business.objects.filter(is_active=True, is_deleted=False).order_by("name")
    q = (request.GET.get("q") or "").strip()

    rows = (
        Product.objects
        .filter(business=business, is_active=True, is_deleted=False)
        .select_related("uom")
        .order_by("name")
    )
    if q:
        rows = rows.filter(
            models.Q(name__icontains=q) |
            models.Q(company_name__icontains=q)
        )

    warehouses = Warehouse.objects.filter(is_active=True).order_by("name")

    ctx = {
        "business": business,
        "businesses": businesses,
        "rows": rows,
        "warehouses": warehouses,
        "query": q,
    }
    return render(request, "barkat/inventory/business_stock_status.html", ctx)

# ---------- Business refill: increase Product.stock_qty ----------
@login_required
def business_refill(request, business_id: int):
    business = get_object_or_404(Business, pk=business_id, is_active=True, is_deleted=False)
    products = (
        Product.objects.filter(business=business, is_active=True, is_deleted=False)
        .select_related("uom")
        .order_by("name")
    )

    if request.method == "POST":
        reference = (request.POST.get("reference") or "").strip()
        qty_entries = []
        for p in products:
            raw = (request.POST.get(f"qty_{p.id}") or "").strip()
            if not raw:
                continue
            try:
                q = Decimal(raw)
            except Exception:
                messages.error(request, f"Invalid quantity for {p.name}.")
                return redirect("business_refill", business_id=business.id)
            if q <= 0:
                continue
            qty_entries.append((p.id, q))

        if not qty_entries:
            messages.error(request, "Enter at least one positive quantity.")
            return redirect("business_refill", business_id=business.id)

        with transaction.atomic():
            # lock & update Product.stock_qty
            for pid, q in qty_entries:
                prod = Product.objects.select_for_update().get(pk=pid)
                prod.stock_qty = (prod.stock_qty or Decimal("0")) + q
                prod.full_clean()
                prod.save(update_fields=["stock_qty", "updated_at"])

        messages.success(request, f"Stock refilled for {business.name}.")
        return redirect("business_stock_status", business_id=business.id)

    ctx = {
        "business": business,
        "products": products,
    }
    return render(request, "barkat/inventory/business_refill.html", ctx)

# ---------- Business → Warehouse: use Product.stock_qty as source ----------
@login_required
def stock_move_b2w(request):
    """
    Move stock from a Business (Product.stock_qty) to a Warehouse.
    GET: ?business=<id>&dest_warehouse=<id>
    """
    business_id = request.GET.get("business") or request.POST.get("business")
    business = get_object_or_404(Business, pk=business_id) if business_id else None
    dest_wh_id = request.GET.get("dest_warehouse") or request.POST.get("dest_warehouse")
    dest_wh = get_object_or_404(Warehouse, pk=dest_wh_id) if dest_wh_id else None

    businesses = Business.objects.filter(is_active=True, is_deleted=False).order_by("name")
    warehouses = Warehouse.objects.filter(is_active=True).order_by("name")

    products = Product.objects.none()
    if business:
        products = (
            Product.objects.filter(business=business, is_active=True, is_deleted=False)
            .select_related("uom")
            .order_by("name")
        )

    if request.method == "POST":
        if not business:
            messages.error(request, "Please select a Business.")
            return redirect(request.path)
        if not dest_wh:
            messages.error(request, "Please select a destination Warehouse.")
            return redirect(f"{request.path}?business={business.id}")

        reference = (request.POST.get("reference") or "").strip()
        qty_entries = []
        for p in products:
            raw = (request.POST.get(f"qty_{p.id}") or "").strip()
            if not raw:
                continue
            try:
                q = Decimal(raw)
            except Exception:
                messages.error(request, f"Invalid quantity for {p.name}.")
                return redirect(f"{request.path}?business={business.id}&dest_warehouse={dest_wh.id}")
            if q <= 0:
                continue
            qty_entries.append((p.id, q))

        if not qty_entries:
            messages.error(request, "Enter at least one positive quantity.")
            return redirect(f"{request.path}?business={business.id}&dest_warehouse={dest_wh.id}")

        with transaction.atomic():
            # 1) Validate availability on Product.stock_qty (business scope)
            locked_products = {
                pid: Product.objects.select_for_update().select_related("business").get(pk=pid)
                for pid, _ in qty_entries
            }
            for pid, q in qty_entries:
                prod = locked_products[pid]
                if prod.business_id != business.id:
                    messages.error(request, f"{prod.name}: not part of selected business.")
                    return redirect(f"{request.path}?business={business.id}&dest_warehouse={dest_wh.id}")
                if (prod.stock_qty or Decimal("0")) < q:
                    messages.error(request, f"Insufficient stock of {prod.name} in {business.name}.")
                    return redirect(f"{request.path}?business={business.id}&dest_warehouse={dest_wh.id}")

            # 2) Apply: decrement Product.stock_qty, increment WarehouseStock, create POSTED StockMove
            for pid, q in qty_entries:
                prod = locked_products[pid]

                # decrement product stock
                prod.stock_qty = (prod.stock_qty or Decimal("0")) - q
                prod.full_clean()
                prod.save(update_fields=["stock_qty", "updated_at"])

                # increment warehouse stock row
                wh_row, _ = WarehouseStock.objects.select_for_update().get_or_create(
                    warehouse=dest_wh, product=prod, defaults={"quantity": Decimal("0")}
                )
                wh_row.quantity = (wh_row.quantity or Decimal("0")) + q
                wh_row.full_clean()
                wh_row.save(update_fields=["quantity", "updated_at"])

                # log move (mark as POSTED directly, since we applied amounts here)
                mv = StockMove.objects.create(
                    product=prod,
                    source_business=business,
                    dest_warehouse=dest_wh,
                    quantity=q,
                    reference=reference,
                    status=StockMove.Status.POSTED,
                    created_by=getattr(request, "user", None),
                    updated_by=getattr(request, "user", None),
                )

        messages.success(request, "Stock moved from Business to Warehouse successfully.")
        return redirect("business_stock_status", business_id=business.id)

    ctx = {
        "business": business,
        "businesses": businesses,
        "warehouses": warehouses,
        "dest_wh": dest_wh,
        "products": products,
    }
    return render(request, "barkat/inventory/stock_move_b2w.html", ctx)





class QuickReceiptListView(LoginRequiredMixin, ListView):
    model = Payment
    template_name = "barkat/finance/quick_receipt_list.html"
    context_object_name = "receipts"
    paginate_by = 25

    def _parse_date(self, s):
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return None

    def get_queryset(self):
        request = self.request

        # base queryset. only receipts (direction IN)
        base_qs = (
            Payment.objects
            .filter(direction=Payment.IN)
            .select_related("business", "party", "bank_account")
            .order_by("-date", "-id")
        )

        q_text = (request.GET.get("q") or "").strip()
        date_from_str = (request.GET.get("date_from") or "").strip()
        date_to_str = (request.GET.get("date_to") or "").strip()
        period = (request.GET.get("period") or "").strip().lower()
        pending_only = (request.GET.get("pending") == "1")

        # store for context
        self.q_text = q_text
        self.period = period
        self.pending_only = pending_only

        today = date.today()
        if period == "today":
            date_from = date_to = today
        elif period == "yesterday":
            d = today - timedelta(days=1)
            date_from = date_to = d
        elif period == "month":
            date_from = today.replace(day=1)
            date_to = today
        elif period == "year":
            date_from = today.replace(month=1, day=1)
            date_to = today
        else:
            date_from = self._parse_date(date_from_str)
            date_to = self._parse_date(date_to_str)

        self.date_from = date_from
        self.date_to = date_to

        # text search
        if q_text:
            base_qs = base_qs.filter(
                Q(party__display_name__icontains=q_text)
                | Q(reference__icontains=q_text)
                | Q(description__icontains=q_text)
                | Q(amount__icontains=q_text)
            )

        # date filter
        if date_from:
            base_qs = base_qs.filter(date__gte=date_from)
        if date_to:
            base_qs = base_qs.filter(date__lte=date_to)

        # keep for totals
        self.base_qs = base_qs

        # pending cheques only
        if pending_only:
            return base_qs.filter(
                payment_method=Payment.PaymentMethod.CHEQUE,
                cheque_status=Payment.ChequeStatus.PENDING,
            )

        return base_qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        base_qs = getattr(self, "base_qs", Payment.objects.none())

        # 1. cash total. all receipts with source CASH
        cash_qs = base_qs.filter(payment_source=Payment.CASH)
        cash_agg = cash_qs.aggregate(cash_total=Sum("amount"))
        cash_total = cash_agg["cash_total"] or Decimal("0.00")

        # 2. pending cheques total
        pending_qs = base_qs.filter(
            payment_method=Payment.PaymentMethod.CHEQUE,
            cheque_status=Payment.ChequeStatus.PENDING,
        )
        pending_agg = pending_qs.aggregate(pending_total=Sum("amount"))
        pending_total = pending_agg["pending_total"] or Decimal("0.00")

        # 3. bank received total. source BANK but exclude pending cheques
        bank_received_qs = base_qs.filter(
            payment_source=Payment.BANK
        ).exclude(
            payment_method=Payment.PaymentMethod.CHEQUE,
            cheque_status=Payment.ChequeStatus.PENDING,
        )
        bank_agg = bank_received_qs.aggregate(bank_total=Sum("amount"))
        bank_total = bank_agg["bank_total"] or Decimal("0.00")

        # 4. overall received = cash + bank received
        total_received = cash_total + bank_total

        ctx["cash_total"] = cash_total
        ctx["bank_total"] = bank_total
        ctx["total_received"] = total_received
        ctx["pending_total"] = pending_total

        ctx["q"] = self.q_text
        ctx["period"] = self.period
        ctx["pending_only"] = self.pending_only
        ctx["date_from"] = self.date_from.isoformat() if self.date_from else ""
        ctx["date_to"] = self.date_to.isoformat() if self.date_to else ""

        return ctx

class QuickReceiptCreateView(LoginRequiredMixin, FormView):
    template_name = "barkat/finance/quick_receipt.html"
    form_class = QuickReceiptForm
    success_url = reverse_lazy("quick_receipt_list")

    def form_valid(self, form):
        try:
            # create_payment already uses the date from form.cleaned_data['date']
            payment = form.create_payment(self.request.user)
        except Exception as e:
            form.add_error(None, str(e))
            return self.form_invalid(form)

        # CashFlow is now automatically handled by Payment.save()


        messages.success(
            self.request,
            f"Receipt saved for {payment.date}. {payment.party.display_name} gave {payment.amount}.",
        )
        return super().form_valid(form)

import datetime
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import FormView

import datetime
from django.utils import timezone
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import FormView
from datetime import datetime
from .models import Payment, CashFlow, Party, BankAccount

class QuickReceiptUpdateView(LoginRequiredMixin, FormView):
    template_name = "barkat/finance/quick_receipt.html"
    form_class = QuickReceiptForm
    payment: Payment  # Type hint for Pylance

    def dispatch(self, request, *args, **kwargs):
        self.payment = get_object_or_404(Payment, pk=kwargs.get("pk"), direction=Payment.IN)
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        p = self.payment
        # Use .pk instead of .id to satisfy Pylance
        initial = {
            "party_id": p.party.pk if p.party else None,
            "party_name": p.party.display_name if p.party else "",
            "date": p.date, 
            "amount": p.amount,
            "ref_no": p.reference,
            "note": p.description,
            "type": p.payment_method,
            "bank_account": p.bank_account.pk if p.bank_account else None,
        }
        if hasattr(p, "cheque_status"):
            initial["cheque_status"] = p.cheque_status
        return initial

    @transaction.atomic
    def form_valid(self, form):
        p = self.payment
        new_date = form.cleaned_data["date"]
        
        # --- 1. SYNC THE TIMESTAMP (created_at) ---
        # Combine selected date with current time
        # We use midday (12:00) to ensure the date doesn't shift due to Timezone offsets
        new_dt = datetime.datetime.combine(new_date, datetime.time(12, 0, 0))
        if timezone.is_aware(timezone.now()):
            new_dt = timezone.make_aware(new_dt)

        # --- 2. UPDATE PAYMENT ---
        p.party = form.party
        p.amount = form.cleaned_data["amount"]
        p.description = form.cleaned_data.get("note") or ""
        p.reference = form.cleaned_data.get("ref_no") or ""
        p.business = form._infer_business(self.request.user)
        
        # Update both date fields
        p.date = new_date           # Logical Date (DateField)
        p.created_at = new_dt       # Record Timestamp (DateTimeField)

        # Payment Method Logic
        method = form.cleaned_data["type"]
        p.payment_method = method
        if method == "cash":
            p.bank_account = None
        else:
            p.bank_account = form.cleaned_data.get("bank_account")
            if method == "cheque":
                p.cheque_status = form.cleaned_data.get("cheque_status") or Payment.ChequeStatus.PENDING

        p.updated_by = self.request.user
        # Save forces both 'date' and 'created_at' to update in DB
        p.save()

        # CashFlow is now automatically handled by Payment.save()


        messages.success(self.request, f"Receipt and Ledger successfully moved to {p.date}")
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse("quick_receipt_list")

from django.views.generic import DeleteView
from django.db import transaction

class QuickReceiptDeleteView(LoginRequiredMixin, DeleteView):
    model = Payment
    success_url = reverse_lazy("quick_receipt_list")

    def get_queryset(self):
        # Only allow deleting receipts (IN)
        return Payment.objects.filter(direction=Payment.IN)

    @transaction.atomic
    def form_valid(self, form):
        payment = self.get_object()
        
        # CashFlow deletion handled by Payment.delete()

        
        messages.success(self.request, "Receipt and ledger entry permanently deleted.")
        return super().form_valid(form)


@require_GET
@login_required
def party_search(request):
    """
    Lightweight JSON search for Party suggestions.
    Accepts:
      q   . search text
      type . optional filter . customer, vendor, both
    """
    q = (request.GET.get("q") or "").strip()
    type_filter = (request.GET.get("type") or "").strip().upper()

    qs = Party.objects.all()

    if type_filter in {"CUSTOMER", "VENDOR", "BOTH"}:
        qs = qs.filter(type=type_filter)

    if q:
        qs = qs.filter(display_name__icontains=q)

    qs = qs.order_by("display_name")[:10]

    data = [
        {
            "id": p.id,
            "name": p.display_name,
            "type": p.get_type_display(),  # eg . Customer or Vendor
        }
        for p in qs
    ]
    return JsonResponse(data, safe=False)


# --------------------------------
# User Settings
# --------------------------------
class UserSettingsUpdateView(LoginRequiredMixin, UpdateView):
    model = UserSettings
    form_class = UserSettingsForm
    template_name = "barkat/settings/user_settings.html"
    success_url = reverse_lazy("user_settings")
    
    def get_object(self, queryset=None):
        """Get or create user settings."""
        obj, created = UserSettings.objects.get_or_create(user=self.request.user)
        return obj
    
    def form_valid(self, form):
        from django.contrib.auth.hashers import make_password
        remove = form.cleaned_data.get("remove_cancellation_password") is True
        new_pw = (form.cleaned_data.get("cancellation_password_new") or "").strip()
        self.object = form.save(commit=False)
        if remove:
            self.object.cancellation_password = ""
        elif new_pw:
            self.object.cancellation_password = make_password(new_pw)
        self.object.save()
        messages.success(self.request, "Settings saved successfully.")
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["businesses"] = Business.objects.filter(is_active=True, is_deleted=False).order_by("name")
        ctx["has_cancellation_password"] = bool(
            getattr(self.object, "cancellation_password", None) or ""
        )
        return ctx


# --------------------------------
# Barcode Printing
# --------------------------------
# Use same tmp folder pattern as POS receipts
from pathlib import Path
TMP_DIR: Path = Path(
    getattr(settings, "RECEIPT_TMP_DIR", Path(settings.BASE_DIR) / "tmp_receipts")
).resolve()
TMP_DIR.mkdir(parents=True, exist_ok=True)

@method_decorator(csrf_exempt, name="dispatch")
class PrintBarcodeLabelsView(LoginRequiredMixin, View):
    """Print barcode labels for products - silent print like quick receipt."""
    
    def post(self, request: HttpRequest):
        try:
            import json
            from .utils.barcode_label_render import render_barcode_labels
            from .utils.pos_print import raw_print_bitmap, PosPrintError
            
            data = json.loads(request.body)
            product_ids = data.get("product_ids", [])
            quantities = data.get("quantities", {})  # {product_id: quantity}
            
            if not product_ids:
                return JsonResponse({"ok": False, "error": "No products selected."}, status=400)
            
            # Get products
            products = Product.objects.filter(id__in=product_ids, is_deleted=False)
            if not products.exists():
                return JsonResponse({"ok": False, "error": "No valid products found."}, status=400)
            
            # Get user settings
            try:
                user_settings = UserSettings.objects.get(user=request.user)
            except UserSettings.DoesNotExist:
                return JsonResponse({
                    "ok": False, 
                    "error": "Please configure your barcode printer in Settings first."
                }, status=400)
            
            printer_name = (user_settings.barcode_printer_name or "").strip()
            if not printer_name:
                return JsonResponse({
                    "ok": False,
                    "error": "Barcode printer not configured. Please set it in Settings."
                }, status=400)
            
            # Get business name from user settings (custom text field)
            business_name = (user_settings.business_name or "").strip()
            
            # Prepare product data
            product_list = []
            quantities_dict = {}
            for product in products:
                if not product.barcode:
                    continue
                product_list.append({
                    "id": product.id,
                    "name": product.name,
                    "barcode": product.barcode,
                    "company_name": product.company_name or "",
                    # Ensure label renderer can show SALE price
                    "sale_price": product.sale_price,
                })
                # Use provided quantity or default to 1
                quantities_dict[product.id] = int(quantities.get(str(product.id), 1))
            
            if not product_list:
                return JsonResponse({
                    "ok": False,
                    "error": "No products with barcodes found."
                }, status=400)
            
            # Generate barcode labels image - use same TMP_DIR as receipts
            bmp_path = render_barcode_labels(
                products=product_list,
                quantities=quantities_dict,
                business_name=business_name,
                out_dir=str(TMP_DIR),
                debug=False,
            )
            
            # Print silently - exactly like quick receipt and sales order
            # Windows print spooler automatically handles queueing if printer is offline
            # Total media width: 82mm at 203 DPI = 655px
            from .utils.barcode_label_render import TOTAL_MEDIA_WIDTH_PX
            raw_print_bitmap(
                printer_name=printer_name,
                bmp_path=bmp_path,
                width_px=TOTAL_MEDIA_WIDTH_PX,  # 82mm total media width at 203 DPI
            )
            
            return JsonResponse({
                "ok": True,
                "message": f"Printed {sum(quantities_dict.values())} barcode label(s) successfully.",
            })
            
        except PosPrintError as e:
            return JsonResponse({"ok": False, "error": str(e)}, status=500)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"ok": False, "error": f"Error: {str(e)}"}, status=500)


# verify_cancellation_password_api is used instead.
